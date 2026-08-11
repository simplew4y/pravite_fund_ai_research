#!/usr/bin/env python3
"""Private-fund research directory ingestion pipeline.

This pipeline treats a local directory as one research dataset. It stores a
unified evidence skeleton in SQLite while keeping Excel-specific structure in
dedicated fact tables:

    workspace_root/
      datasets.sqlite3
      <dataset_id>/
        raw/
        meta/collection.sqlite3

PDF files are parsed deterministically with PyMuPDF. Documents that fail the
text-quality gate are recorded as ``needs_ocr`` with no searchable chunks, so
scanned files cannot silently look indexed. Office/text formats are dispatched
through format-specific adapters, while document versions and tombstones keep
incremental runs auditable.
"""

from __future__ import annotations

import argparse
import bisect
import hashlib
import json
import os
import re
import shutil
import sqlite3
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Optional

try:
    from .private_fund_format_adapters import (  # type: ignore
        SUPPORTED_EXTENSIONS as ADAPTER_EXTENSIONS,
        adapt_document,
    )
except ImportError:
    from private_fund_format_adapters import (  # type: ignore
        SUPPORTED_EXTENSIONS as ADAPTER_EXTENSIONS,
        adapt_document,
    )

try:
    from .document_classifier import (  # type: ignore
        CLASSIFIER_VERSION,
        LEGACY_DOCUMENT_TYPE_MAP,
        TAXONOMY_VERSION,
        ClassificationChatClient,
        DocumentClassification,
        build_document_preview,
        classify_document,
    )
except ImportError:
    from document_classifier import (  # type: ignore
        CLASSIFIER_VERSION,
        LEGACY_DOCUMENT_TYPE_MAP,
        TAXONOMY_VERSION,
        ClassificationChatClient,
        DocumentClassification,
        build_document_preview,
        classify_document,
    )

CORE_EXTENSIONS = {".pdf", ".xlsx", ".xlsm"}
SUPPORTED_EXTENSIONS = CORE_EXTENSIONS | set(ADAPTER_EXTENSIONS)
DEFAULT_MAX_PDF_CHARS = 2200
DEFAULT_MAX_REGION_LABELS = 30
MIN_PDF_MEANINGFUL_CHARS = 40
MIN_PDF_MEANINGFUL_CHARS_PER_PAGE = 15
MIN_PDF_TEXT_PAGE_CHARS = 20
MIN_PDF_TEXT_PAGE_COVERAGE = 0.60


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8", errors="replace")).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_safe(v) for v in value]
    # openpyxl's ArrayFormula/DataTableFormula (and a few other library
    # objects) use the default object repr, which embeds a process-specific
    # memory address.  Persist a deterministic structural description instead.
    attributes = getattr(value, "__dict__", None)
    if isinstance(attributes, dict):
        return {
            "type": type(value).__name__,
            "attributes": {str(k): json_safe(v) for k, v in sorted(attributes.items())},
        }
    return {"type": type(value).__name__}


def dumps_json(value: Any) -> str:
    return json.dumps(json_safe(value), ensure_ascii=False, sort_keys=True)


def normalize_text(value: Any) -> str:
    if value is None:
        text = ""
    elif isinstance(value, str):
        text = value
    elif isinstance(value, (int, float, bool, datetime, date, Path)):
        text = str(json_safe(value))
    else:
        text = dumps_json(value)
    text = unicodedata.normalize("NFKC", text)
    return re.sub(r"\s+", " ", text).strip()


def cell_display(value: Any, max_len: int = 160) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        text = value.isoformat()
    else:
        text = normalize_text(value)
    if len(text) > max_len:
        return text[: max_len - 1] + "..."
    return text


def safe_slug(value: str, fallback: str = "dataset") -> str:
    value = unicodedata.normalize("NFKC", value or "")
    value = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", value, flags=re.UNICODE)
    value = value.strip("._-")
    return value or fallback


def default_workspace_root() -> Path:
    override = os.environ.get("PRIVATE_FUND_DATASET_WORKSPACE")
    if override:
        return Path(override).expanduser().resolve()
    return Path.cwd().resolve() / "output" / "private_fund_datasets"


@dataclass
class IngestOptions:
    source_dir: Path
    workspace_root: Path
    dataset_id: str
    dataset_name: str
    company_name: str = ""
    company_ticker: str = ""
    recursive: bool = True
    reset: bool = False
    job_id: Optional[str] = None


@dataclass
class DocumentIngestResult:
    doc_id: str
    filename: str
    file_type: str
    status: str
    chunk_count: int = 0
    location_count: int = 0
    pdf_page_count: int = 0
    excel_sheet_count: int = 0
    excel_region_count: int = 0
    excel_cell_count: int = 0
    metric_fact_count: int = 0
    error_message: Optional[str] = None
    logical_doc_id: str = ""
    version_no: int = 0
    supersedes_doc_id: Optional[str] = None
    reused: bool = False
    parser_name: str = ""
    parser_version: str = ""
    parser_metadata: dict[str, Any] = field(default_factory=dict)
    lifecycle_state: str = "active"
    doc_type: str = "other"
    doc_subtype: str = ""
    doc_type_confidence: float = 0.0
    classification_status: str = "needs_review"
    classification_method: str = ""
    company_name: str = ""
    company_ticker: str = ""
    company_confidence: float = 0.0


@dataclass
class IngestResult:
    job_id: str
    dataset_id: str
    dataset_name: str
    source_dir: str
    workspace_root: str
    dataset_root: str
    collection_db_path: str
    global_db_path: str
    status: str
    file_count: int
    discovered_file_count: int = 0
    supported_file_count: int = 0
    unsupported_file_count: int = 0
    removed_file_count: int = 0
    warning_count: int = 0
    documents: list[DocumentIngestResult] = field(default_factory=list)
    started_at: str = ""
    finished_at: str = ""
    message: str = ""


def connect_sqlite(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(path), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def ensure_global_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS datasets (
            dataset_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            status TEXT NOT NULL,
            source_dir TEXT,
            dataset_root TEXT NOT NULL,
            company_name TEXT,
            company_ticker TEXT,
            file_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS dataset_state (
            id INTEGER PRIMARY KEY CHECK(id = 1),
            active_dataset_id TEXT,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS private_fund_upload_batches (
            batch_id TEXT PRIMARY KEY,
            status TEXT NOT NULL,
            file_count INTEGER NOT NULL DEFAULT 0,
            message TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            finished_at TEXT
        );

        CREATE TABLE IF NOT EXISTS private_fund_upload_items (
            item_id TEXT PRIMARY KEY,
            batch_id TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            staged_path TEXT NOT NULL,
            file_type TEXT NOT NULL,
            file_size INTEGER NOT NULL,
            checksum TEXT NOT NULL,
            status TEXT NOT NULL,
            company_name TEXT,
            company_ticker TEXT,
            company_confidence REAL NOT NULL DEFAULT 0,
            company_detection_method TEXT,
            matched_dataset_id TEXT,
            project_match_confidence REAL NOT NULL DEFAULT 0,
            project_match_method TEXT,
            candidate_projects_json TEXT,
            pipeline_job_id TEXT,
            error_message TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(batch_id) REFERENCES private_fund_upload_batches(batch_id)
                ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_private_fund_upload_items_batch
            ON private_fund_upload_items(batch_id, created_at);
        CREATE INDEX IF NOT EXISTS idx_private_fund_upload_items_status
            ON private_fund_upload_items(status, updated_at);
        """
    )
    conn.execute(
        "INSERT OR IGNORE INTO dataset_state (id, active_dataset_id, updated_at) VALUES (1, NULL, ?)",
        (now_iso(),),
    )
    conn.commit()


def ensure_collection_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS documents (
            doc_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            logical_doc_id TEXT,
            version_no INTEGER NOT NULL DEFAULT 1,
            supersedes_doc_id TEXT,
            is_current INTEGER NOT NULL DEFAULT 1,
            lifecycle_state TEXT NOT NULL DEFAULT 'active',
            title TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            source_root TEXT,
            source_relpath TEXT,
            stored_path TEXT NOT NULL,
            file_type TEXT NOT NULL,
            doc_type TEXT,
            doc_subtype TEXT,
            doc_type_confidence REAL NOT NULL DEFAULT 0,
            classification_status TEXT NOT NULL DEFAULT 'needs_review',
            classification_method TEXT,
            classification_taxonomy_version TEXT,
            classifier_version TEXT,
            classification_metadata_json TEXT,
            source_type TEXT,
            source_name TEXT,
            company_name TEXT,
            company_ticker TEXT,
            company_confidence REAL NOT NULL DEFAULT 0,
            company_detection_method TEXT,
            document_date TEXT,
            checksum TEXT NOT NULL,
            file_size INTEGER NOT NULL,
            status TEXT NOT NULL,
            chunk_count INTEGER NOT NULL DEFAULT 0,
            error_message TEXT,
            metadata_json TEXT,
            parser_name TEXT,
            parser_version TEXT,
            parser_metadata_json TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            deleted_at TEXT
        );

        CREATE TABLE IF NOT EXISTS chunks (
            chunk_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            chunk_index INTEGER NOT NULL,
            content TEXT NOT NULL,
            content_type TEXT NOT NULL,
            title_path TEXT,
            summary TEXT,
            token_count INTEGER,
            content_hash TEXT NOT NULL,
            prev_chunk_id TEXT,
            next_chunk_id TEXT,
            source_ref TEXT,
            metadata_json TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS chunk_locations (
            location_id TEXT PRIMARY KEY,
            chunk_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            location_index INTEGER NOT NULL DEFAULT 0,
            page_start INTEGER,
            page_end INTEGER,
            page_numbers_json TEXT,
            slide_start INTEGER,
            slide_end INTEGER,
            sheet_name TEXT,
            cell_range TEXT,
            heading_path TEXT,
            bbox_json TEXT,
            source_refs_json TEXT,
            display_text TEXT NOT NULL,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS index_registry (
            index_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            index_type TEXT NOT NULL,
            collection_name TEXT,
            index_path TEXT NOT NULL,
            source_doc_ids_json TEXT,
            source_chunk_count INTEGER,
            status TEXT NOT NULL,
            built_at TEXT,
            error_message TEXT,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS ingest_jobs (
            job_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            job_type TEXT NOT NULL,
            status TEXT NOT NULL,
            doc_ids_json TEXT,
            file_count INTEGER NOT NULL DEFAULT 0,
            log_path TEXT,
            message TEXT,
            returncode INTEGER,
            created_at TEXT NOT NULL,
            started_at TEXT,
            finished_at TEXT,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS pdf_pages (
            page_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            page_number INTEGER NOT NULL,
            text TEXT NOT NULL,
            char_count INTEGER NOT NULL,
            word_count INTEGER NOT NULL,
            extraction_method TEXT NOT NULL,
            bbox_json TEXT,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS excel_workbooks (
            workbook_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            workbook_type TEXT NOT NULL,
            sheet_count INTEGER NOT NULL,
            visible_sheet_count INTEGER NOT NULL,
            formula_count INTEGER NOT NULL,
            non_empty_cell_count INTEGER NOT NULL,
            formula_density REAL NOT NULL,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS excel_sheets (
            sheet_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            sheet_index INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            sheet_role TEXT NOT NULL,
            sheet_state TEXT,
            used_range TEXT,
            row_count INTEGER NOT NULL,
            col_count INTEGER NOT NULL,
            non_empty_cell_count INTEGER NOT NULL,
            formula_count INTEGER NOT NULL,
            formula_density REAL NOT NULL,
            summary TEXT,
            header_json TEXT,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS excel_regions (
            region_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            region_index INTEGER NOT NULL,
            region_type TEXT NOT NULL,
            cell_range TEXT NOT NULL,
            row_count INTEGER NOT NULL,
            col_count INTEGER NOT NULL,
            non_empty_cell_count INTEGER NOT NULL,
            formula_count INTEGER NOT NULL,
            formula_density REAL NOT NULL,
            summary TEXT,
            header_json TEXT,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS excel_cells (
            cell_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            sheet_name TEXT NOT NULL,
            cell_ref TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            col_index INTEGER NOT NULL,
            value_type TEXT NOT NULL,
            display_value TEXT,
            raw_value TEXT,
            numeric_value REAL,
            formula TEXT,
            cached_value TEXT,
            number_format TEXT,
            row_label TEXT,
            col_label TEXT,
            period TEXT,
            unit TEXT,
            is_formula INTEGER NOT NULL DEFAULT 0,
            formula_type TEXT,
            formula_cache_status TEXT NOT NULL DEFAULT 'not_applicable',
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS metric_facts (
            fact_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            metric_name TEXT NOT NULL,
            metric_alias TEXT,
            period TEXT,
            value_text TEXT,
            value_numeric REAL,
            unit TEXT,
            sheet_name TEXT NOT NULL,
            cell_ref TEXT NOT NULL,
            source_range TEXT,
            formula TEXT,
            confidence REAL NOT NULL DEFAULT 0.5,
            fact_status TEXT NOT NULL DEFAULT 'candidate',
            quality_status TEXT NOT NULL DEFAULT 'review_required',
            quality_issues_json TEXT,
            metadata_json TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_documents_dataset_status ON documents(dataset_id, status);
        CREATE INDEX IF NOT EXISTS idx_chunks_doc_index ON chunks(doc_id, chunk_index);
        CREATE INDEX IF NOT EXISTS idx_locations_chunk ON chunk_locations(chunk_id);
        CREATE INDEX IF NOT EXISTS idx_pdf_pages_doc_page ON pdf_pages(doc_id, page_number);
        CREATE INDEX IF NOT EXISTS idx_excel_sheets_doc ON excel_sheets(doc_id, sheet_name);
        CREATE INDEX IF NOT EXISTS idx_excel_regions_doc ON excel_regions(doc_id, sheet_name, cell_range);
        CREATE INDEX IF NOT EXISTS idx_excel_cells_doc_sheet ON excel_cells(doc_id, sheet_name, cell_ref);
        CREATE INDEX IF NOT EXISTS idx_excel_cells_doc_sheet_position
            ON excel_cells(doc_id, sheet_name, row_index, col_index);
        CREATE INDEX IF NOT EXISTS idx_metric_facts_metric ON metric_facts(doc_id, metric_name, period);
        CREATE INDEX IF NOT EXISTS idx_metric_facts_source ON metric_facts(doc_id, sheet_name, cell_ref);
        CREATE INDEX IF NOT EXISTS idx_index_registry_dataset_type ON index_registry(dataset_id, index_type);
        """
    )
    _ensure_collection_schema_migrations(conn)
    conn.commit()


def _table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {str(row["name"]) for row in conn.execute(f"PRAGMA table_info({table})")}


def _ensure_columns(conn: sqlite3.Connection, table: str, definitions: dict[str, str]) -> None:
    existing = _table_columns(conn, table)
    for column, definition in definitions.items():
        if column not in existing:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def _normalized_source_relpath(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).replace("\\", "/")
    text = re.sub(r"/+", "/", text).strip("/ ")
    return text or "document"


def _logical_doc_id(dataset_id: str, source_relpath: str) -> str:
    normalized = _normalized_source_relpath(source_relpath)
    return sha256_text(f"{dataset_id}\0{normalized}")[:40]


def _migrate_document_versions(conn: sqlite3.Connection) -> None:
    rows = list(
        conn.execute(
            """
            SELECT doc_id, dataset_id, logical_doc_id, version_no, original_filename,
                   stored_path, source_root, source_relpath, status, lifecycle_state,
                   file_type, chunk_count, error_message, metadata_json,
                   created_at, updated_at, deleted_at
            FROM documents
            ORDER BY dataset_id, COALESCE(created_at, ''), doc_id
            """
        )
    )
    legacy_identity_counts: dict[tuple[str, str], int] = {}
    for row in rows:
        if row["logical_doc_id"]:
            continue
        fallback = _normalized_source_relpath(
            row["source_relpath"] or row["original_filename"]
        )
        key = (str(row["dataset_id"]), fallback)
        legacy_identity_counts[key] = legacy_identity_counts.get(key, 0) + 1

    groups: dict[tuple[str, str], list[sqlite3.Row]] = {}
    for row in rows:
        fallback = _normalized_source_relpath(
            row["source_relpath"] or row["original_filename"]
        )
        legacy_key = (str(row["dataset_id"]), fallback)
        is_ambiguous_legacy_identity = (
            not row["logical_doc_id"] and legacy_identity_counts.get(legacy_key, 0) > 1
        )
        if is_ambiguous_legacy_identity:
            # The legacy schema discarded the original relative source path.
            # Two active rows sharing a basename may be independent files from
            # different subdirectories, not versions of one document.  Use a
            # stable synthetic path rather than destructively merging them.
            relpath = _normalized_source_relpath(
                f"__legacy_ambiguous__/{str(row['doc_id'])[:12]}/{Path(str(row['original_filename'])).name}"
            )
        else:
            relpath = fallback
        logical_id = str(row["logical_doc_id"] or _logical_doc_id(row["dataset_id"], relpath))
        try:
            metadata = json.loads(row["metadata_json"] or "{}")
            if not isinstance(metadata, dict):
                metadata = {}
        except (TypeError, ValueError, json.JSONDecodeError):
            metadata = {}
        if is_ambiguous_legacy_identity:
            metadata.update(
                {
                    "legacy_identity_disambiguated": True,
                    "legacy_original_relpath": fallback,
                    "legacy_stored_path": str(row["stored_path"] or ""),
                }
            )
        conn.execute(
            """
            UPDATE documents
            SET logical_doc_id = ?, source_relpath = ?, metadata_json = ?,
                parser_name = COALESCE(NULLIF(parser_name, ''), 'legacy_unknown'),
                parser_version = COALESCE(NULLIF(parser_version, ''), 'unknown')
            WHERE doc_id = ?
            """,
            (logical_id, relpath, dumps_json(metadata), row["doc_id"]),
        )
        groups.setdefault((str(row["dataset_id"]), logical_id), []).append(row)

    migration_time = now_iso()
    for (_, _), versions in groups.items():
        ordered = sorted(
            versions,
            key=lambda row: (
                int(row["version_no"] or 0),
                str(row["created_at"] or ""),
                str(row["doc_id"]),
            ),
        )
        active = [row for row in ordered if not row["deleted_at"]]
        current_doc_id = str(active[-1]["doc_id"]) if active else ""
        previous_doc_id: Optional[str] = None
        for version_no, row in enumerate(ordered, start=1):
            doc_id = str(row["doc_id"])
            is_current = 1 if doc_id == current_doc_id else 0
            status = str(row["status"] or "")
            lifecycle_state = str(row["lifecycle_state"] or "active")
            deleted_at = row["deleted_at"]
            if status in {"removed", "superseded"}:
                lifecycle_state = status
                if int(row["chunk_count"] or 0) > 0:
                    status = "indexed"
                elif str(row["file_type"] or "") == "pdf" and conn.execute(
                    "SELECT COUNT(*) FROM pdf_pages WHERE doc_id = ?", (doc_id,)
                ).fetchone()[0]:
                    status = "needs_ocr"
                else:
                    status = "failed"
            if active and not is_current and not deleted_at:
                deleted_at = row["updated_at"] or migration_time
                lifecycle_state = "superseded"
            elif is_current and not deleted_at:
                lifecycle_state = "active"
            elif not is_current and lifecycle_state == "active":
                lifecycle_state = "superseded"
            conn.execute(
                """
                UPDATE documents
                SET version_no = ?, supersedes_doc_id = ?, is_current = ?,
                    status = ?, lifecycle_state = ?, deleted_at = ?
                WHERE doc_id = ?
                """,
                (
                    version_no,
                    previous_doc_id,
                    is_current,
                    status,
                    lifecycle_state,
                    deleted_at,
                    doc_id,
                ),
            )
            previous_doc_id = doc_id


def _ensure_collection_schema_migrations(conn: sqlite3.Connection) -> None:
    _ensure_columns(
        conn,
        "documents",
        {
            "logical_doc_id": "TEXT",
            "version_no": "INTEGER NOT NULL DEFAULT 1",
            "supersedes_doc_id": "TEXT",
            "is_current": "INTEGER NOT NULL DEFAULT 1",
            "lifecycle_state": "TEXT NOT NULL DEFAULT 'active'",
            "source_root": "TEXT",
            "source_relpath": "TEXT",
            "parser_name": "TEXT",
            "parser_version": "TEXT",
            "parser_metadata_json": "TEXT",
            "doc_subtype": "TEXT",
            "doc_type_confidence": "REAL NOT NULL DEFAULT 0",
            "classification_status": "TEXT NOT NULL DEFAULT 'needs_review'",
            "classification_method": "TEXT",
            "classification_taxonomy_version": "TEXT",
            "classifier_version": "TEXT",
            "classification_metadata_json": "TEXT",
            "company_confidence": "REAL NOT NULL DEFAULT 0",
            "company_detection_method": "TEXT",
        },
    )
    _ensure_columns(
        conn,
        "excel_cells",
        {
            "formula_type": "TEXT",
            "formula_cache_status": "TEXT NOT NULL DEFAULT 'not_applicable'",
        },
    )
    _ensure_columns(
        conn,
        "metric_facts",
        {
            "fact_status": "TEXT NOT NULL DEFAULT 'candidate'",
            "quality_status": "TEXT NOT NULL DEFAULT 'review_required'",
            "quality_issues_json": "TEXT",
        },
    )
    _migrate_document_versions(conn)
    _migrate_document_taxonomy(conn)
    _repair_metric_fact_periods(conn)
    conn.executescript(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_documents_logical_version
            ON documents(dataset_id, logical_doc_id, version_no);
        CREATE INDEX IF NOT EXISTS idx_documents_current
            ON documents(dataset_id, is_current, deleted_at, status);
        CREATE INDEX IF NOT EXISTS idx_documents_source_path
            ON documents(dataset_id, source_root, source_relpath);
        """
    )


def _migrate_document_taxonomy(conn: sqlite3.Connection) -> None:
    """Collapse legacy primary types into the three-category v2 taxonomy."""

    rows = conn.execute(
        """
        SELECT doc_id, doc_type, classification_metadata_json
        FROM documents
        WHERE COALESCE(classification_taxonomy_version, '') <> ?
           OR COALESCE(doc_type, '') NOT IN ('financial_valuation_data',
                                             'meeting_third_party', 'other')
        """,
        (TAXONOMY_VERSION,),
    ).fetchall()
    changed_at = now_iso()
    for row in rows:
        legacy_type = str(row["doc_type"] or "unknown").strip().lower()
        doc_type = LEGACY_DOCUMENT_TYPE_MAP.get(legacy_type, "other")
        try:
            metadata = json.loads(row["classification_metadata_json"] or "{}")
            if not isinstance(metadata, dict):
                metadata = {}
        except (TypeError, ValueError, json.JSONDecodeError):
            metadata = {}
        metadata.update(
            {
                "doc_type": doc_type,
                "legacy_doc_type": legacy_type,
                "taxonomy_version": TAXONOMY_VERSION,
                "classifier_version": CLASSIFIER_VERSION,
            }
        )
        conn.execute(
            """
            UPDATE documents
            SET doc_type=?, classification_taxonomy_version=?, classifier_version=?,
                classification_metadata_json=?, updated_at=?
            WHERE doc_id=?
            """,
            (
                doc_type,
                TAXONOMY_VERSION,
                CLASSIFIER_VERSION,
                dumps_json(metadata),
                changed_at,
                row["doc_id"],
            ),
        )


def _location_id(chunk_id: str, loc_index: int, display: str) -> str:
    return sha256_text(f"{chunk_id}\0{loc_index}\0{display}")[:40]


def _chunk_id(doc_id: str, index: int, content: str, source_ref: str) -> str:
    return sha256_text(f"{doc_id}\0{index}\0{source_ref}\0{sha256_text(content)}")[:40]


def _token_count(content: str) -> int:
    return len(re.findall(r"\w+", content))


def _copy_to_raw(source: Path, raw_dir: Path) -> Path:
    raw_dir.mkdir(parents=True, exist_ok=True)
    target = raw_dir / source.name
    if source.resolve() == target.resolve():
        return target
    if target.exists():
        src_hash = sha256_file(source)
        dst_hash = sha256_file(target)
        if src_hash == dst_hash:
            return target
        target = raw_dir / f"{source.stem}_{src_hash[:8]}{source.suffix}"
        if target.exists() and sha256_file(target) == src_hash:
            return target
    shutil.copy2(source, target)
    return target


def _path_is_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def _validate_source_output_layout(
    source_dir: Path, workspace_root: Path, dataset_root: Path
) -> None:
    source = source_dir.resolve()
    workspace = workspace_root.resolve()
    dataset = dataset_root.resolve()
    if source == workspace:
        raise ValueError(
            "directory_path must not equal workspace_root; choose a separate source directory "
            "to prevent generated state from being ingested or tombstoned"
        )
    if source == dataset:
        raise ValueError(
            "directory_path must not equal dataset_root; use a dedicated source subdirectory "
            "such as dataset_root/_uploads"
        )


def _is_ignored_input_file(path: Path) -> bool:
    name = path.name.casefold()
    if name.startswith("~$"):
        return True
    return name.endswith(
        (
            ".sqlite",
            ".sqlite3",
            ".sqlite-wal",
            ".sqlite-shm",
            ".sqlite3-wal",
            ".sqlite3-shm",
            ".db-wal",
            ".db-shm",
        )
    )


def _iter_input_files(
    source_dir: Path,
    recursive: bool,
    *,
    excluded_roots: Optional[list[Path]] = None,
) -> tuple[list[Path], list[Path]]:
    iterator = source_dir.rglob("*") if recursive else source_dir.glob("*")
    exclusions = [root.resolve() for root in (excluded_roots or [])]
    all_files = [
        path
        for path in iterator
        if path.is_file()
        and not any(part.startswith(".") for part in path.relative_to(source_dir).parts)
        and not _is_ignored_input_file(path)
        and not any(_path_is_within(path, root) for root in exclusions)
    ]
    all_files.sort(key=lambda path: str(path.relative_to(source_dir)).casefold())
    supported = [path for path in all_files if path.suffix.lower() in SUPPORTED_EXTENSIONS]
    unsupported = [path for path in all_files if path.suffix.lower() not in SUPPORTED_EXTENSIONS]
    return supported, unsupported


def _iter_supported_files(source_dir: Path, recursive: bool) -> list[Path]:
    supported, _ = _iter_input_files(source_dir, recursive)
    return supported


def _delete_document_payload(conn: sqlite3.Connection, doc_id: str) -> None:
    for table in (
        "chunk_locations",
        "chunks",
        "pdf_pages",
        "excel_workbooks",
        "excel_sheets",
        "excel_regions",
        "excel_cells",
        "metric_facts",
    ):
        conn.execute(f"DELETE FROM {table} WHERE doc_id = ?", (doc_id,))


def _register_document(
    conn: sqlite3.Connection,
    *,
    dataset_id: str,
    stored_path: Path,
    original_filename: str,
    checksum: str,
    classification: DocumentClassification,
    source_root: str,
    source_relpath: str,
    logical_doc_id: str,
    version_no: int,
    supersedes_doc_id: Optional[str],
) -> str:
    doc_id = sha256_text(
        f"{dataset_id}\0{logical_doc_id}\0{version_no}\0{checksum}"
    )[:40]
    now = now_iso()
    conn.execute(
        """
        INSERT INTO documents (
            doc_id, dataset_id, logical_doc_id, version_no, supersedes_doc_id,
            is_current, lifecycle_state, title, original_filename, source_root, source_relpath,
            stored_path, file_type, doc_type, doc_subtype, doc_type_confidence,
            classification_status, classification_method, classification_taxonomy_version,
            classifier_version, classification_metadata_json,
            source_type, source_name, company_name, company_ticker, company_confidence,
            company_detection_method, document_date, checksum, file_size, status, chunk_count, error_message,
            metadata_json, parser_name, parser_version, parser_metadata_json,
            created_at, updated_at, deleted_at
        ) VALUES (
            :doc_id, :dataset_id, :logical_doc_id, :version_no, :supersedes_doc_id,
            0, 'pending', :title, :original_filename, :source_root, :source_relpath,
            :stored_path, :file_type, :doc_type, :doc_subtype, :doc_type_confidence,
            :classification_status, :classification_method, :classification_taxonomy_version,
            :classifier_version, :classification_metadata_json,
            'local_directory', :source_name, :company_name, :company_ticker, :company_confidence,
            :company_detection_method, :document_date, :checksum, :file_size, 'parsing', 0, NULL,
            :metadata_json, NULL, NULL, NULL, :created_at, :updated_at, NULL
        )
        """,
        {
            "doc_id": doc_id,
            "dataset_id": dataset_id,
            "logical_doc_id": logical_doc_id,
            "version_no": version_no,
            "supersedes_doc_id": supersedes_doc_id,
            "title": Path(original_filename).stem,
            "original_filename": original_filename,
            "source_root": source_root,
            "source_relpath": _normalized_source_relpath(source_relpath),
            "stored_path": str(stored_path),
            "file_type": stored_path.suffix.lower().lstrip("."),
            "doc_type": classification.doc_type,
            "doc_subtype": classification.doc_subtype or None,
            "doc_type_confidence": classification.confidence,
            "classification_status": classification.classification_status,
            "classification_method": classification.method,
            "classification_taxonomy_version": classification.taxonomy_version,
            "classifier_version": classification.classifier_version,
            "classification_metadata_json": dumps_json(classification.to_metadata()),
            "source_name": stored_path.name,
            "company_name": classification.company_name,
            "company_ticker": classification.company_ticker,
            "company_confidence": classification.company_confidence,
            "company_detection_method": classification.company_method,
            "document_date": _date_from_filename(original_filename),
            "checksum": checksum,
            "file_size": stored_path.stat().st_size,
            "metadata_json": dumps_json(
                {
                    "source": "private_fund_directory_ingest",
                    "source_root": source_root,
                    "source_relpath": _normalized_source_relpath(source_relpath),
                }
            ),
            "created_at": now,
            "updated_at": now,
        },
    )
    return doc_id


def _update_document_classification(
    conn: sqlite3.Connection,
    doc_id: str,
    classification: DocumentClassification,
) -> None:
    conn.execute(
        """
        UPDATE documents
        SET doc_type = ?, doc_subtype = ?, doc_type_confidence = ?,
            classification_status = ?, classification_method = ?,
            classification_taxonomy_version = ?, classifier_version = ?,
            classification_metadata_json = ?, company_name = ?, company_ticker = ?,
            company_confidence = ?, company_detection_method = ?, updated_at = ?
        WHERE doc_id = ?
        """,
        (
            classification.doc_type,
            classification.doc_subtype or None,
            classification.confidence,
            classification.classification_status,
            classification.method,
            classification.taxonomy_version,
            classification.classifier_version,
            dumps_json(classification.to_metadata()),
            classification.company_name,
            classification.company_ticker,
            classification.company_confidence,
            classification.company_method,
            now_iso(),
            doc_id,
        ),
    )


def _next_document_version(conn: sqlite3.Connection, dataset_id: str, logical_doc_id: str) -> int:
    row = conn.execute(
        """
        SELECT COALESCE(MAX(version_no), 0) + 1
        FROM documents
        WHERE dataset_id = ? AND logical_doc_id = ?
        """,
        (dataset_id, logical_doc_id),
    ).fetchone()
    return int(row[0] or 1)


def _current_document(
    conn: sqlite3.Connection, dataset_id: str, logical_doc_id: str
) -> Optional[sqlite3.Row]:
    return conn.execute(
        """
        SELECT * FROM documents
        WHERE dataset_id = ? AND logical_doc_id = ?
          AND is_current = 1 AND lifecycle_state = 'active'
          AND deleted_at IS NULL
        ORDER BY version_no DESC, created_at DESC
        LIMIT 1
        """,
        (dataset_id, logical_doc_id),
    ).fetchone()


def _latest_document_version(
    conn: sqlite3.Connection, dataset_id: str, logical_doc_id: str
) -> Optional[sqlite3.Row]:
    return conn.execute(
        """
        SELECT * FROM documents
        WHERE dataset_id = ? AND logical_doc_id = ?
        ORDER BY version_no DESC, created_at DESC
        LIMIT 1
        """,
        (dataset_id, logical_doc_id),
    ).fetchone()


def _activate_document_version(
    conn: sqlite3.Connection,
    *,
    dataset_id: str,
    logical_doc_id: str,
    doc_id: str,
) -> None:
    changed_at = now_iso()
    conn.execute(
        """
        UPDATE documents
        SET lifecycle_state = 'superseded', is_current = 0,
            deleted_at = ?, updated_at = ?
        WHERE dataset_id = ? AND logical_doc_id = ? AND doc_id <> ?
          AND is_current = 1 AND deleted_at IS NULL
        """,
        (changed_at, changed_at, dataset_id, logical_doc_id, doc_id),
    )
    conn.execute(
        """
        UPDATE documents
        SET lifecycle_state = 'active', is_current = 1,
            deleted_at = NULL, updated_at = ?
        WHERE doc_id = ?
        """,
        (changed_at, doc_id),
    )


def _document_result_from_row(conn: sqlite3.Connection, row: sqlite3.Row) -> DocumentIngestResult:
    doc_id = str(row["doc_id"])

    def count(table: str) -> int:
        return int(conn.execute(f"SELECT COUNT(*) FROM {table} WHERE doc_id = ?", (doc_id,)).fetchone()[0] or 0)

    parser_metadata: dict[str, Any] = {}
    try:
        parser_metadata = json.loads(row["parser_metadata_json"] or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        parser_metadata = {}
    return DocumentIngestResult(
        doc_id=doc_id,
        filename=str(row["original_filename"]),
        file_type=str(row["file_type"]),
        status=str(row["status"]),
        chunk_count=int(row["chunk_count"] or 0),
        location_count=count("chunk_locations"),
        pdf_page_count=count("pdf_pages"),
        excel_sheet_count=count("excel_sheets"),
        excel_region_count=count("excel_regions"),
        excel_cell_count=count("excel_cells"),
        metric_fact_count=count("metric_facts"),
        error_message=row["error_message"],
        logical_doc_id=str(row["logical_doc_id"] or ""),
        version_no=int(row["version_no"] or 0),
        supersedes_doc_id=row["supersedes_doc_id"],
        reused=True,
        parser_name=str(row["parser_name"] or ""),
        parser_version=str(row["parser_version"] or ""),
        parser_metadata=parser_metadata,
        lifecycle_state=str(row["lifecycle_state"] or "active"),
        doc_type=str(row["doc_type"] or "other"),
        doc_subtype=str(row["doc_subtype"] or ""),
        doc_type_confidence=float(row["doc_type_confidence"] or 0),
        classification_status=str(row["classification_status"] or "needs_review"),
        classification_method=str(row["classification_method"] or ""),
        company_name=str(row["company_name"] or ""),
        company_ticker=str(row["company_ticker"] or ""),
        company_confidence=float(row["company_confidence"] or 0),
    )


def _date_from_filename(name: str) -> str:
    text = unicodedata.normalize("NFKC", name)
    for pattern, fmt in (
        (r"(20\d{2})[-_.年]?([01]\d)[-_.月]?([0-3]\d)", "%Y-%m-%d"),
        (r"(20\d{2})([01]\d)([0-3]\d)", "%Y-%m-%d"),
    ):
        match = re.search(pattern, text)
        if match:
            y, m, d = match.groups()
            try:
                return datetime.strptime(f"{y}-{m}-{d}", "%Y-%m-%d").date().isoformat()
            except ValueError:
                pass
    match = re.search(r"(20\d{2})", text)
    return f"{match.group(1)}-01-01" if match else ""


def _write_chunks(
    conn: sqlite3.Connection,
    *,
    dataset_id: str,
    doc_id: str,
    chunks: list[dict[str, Any]],
) -> tuple[int, int]:
    now = now_iso()
    chunk_rows: list[dict[str, Any]] = []
    loc_rows: list[dict[str, Any]] = []
    chunk_ids: list[str] = []

    for index, chunk in enumerate(chunks, start=1):
        content = str(chunk.get("content") or "")
        source_ref = str(chunk.get("source_ref") or "")
        chunk_id = _chunk_id(doc_id, index, content, source_ref)
        chunk_ids.append(chunk_id)
        title_path = chunk.get("title_path")
        if isinstance(title_path, list):
            title_text = " > ".join(str(x) for x in title_path if str(x).strip())
        else:
            title_text = str(title_path or chunk.get("title") or "")
        chunk_rows.append(
            {
                "chunk_id": chunk_id,
                "dataset_id": dataset_id,
                "doc_id": doc_id,
                "chunk_index": index,
                "content": content,
                "content_type": str(chunk.get("content_type") or "text"),
                "title_path": title_text,
                "summary": chunk.get("summary") or None,
                "token_count": _token_count(content),
                "content_hash": sha256_text(content),
                "prev_chunk_id": None,
                "next_chunk_id": None,
                "source_ref": source_ref or None,
                "metadata_json": dumps_json(chunk.get("metadata") or {}),
                "created_at": now,
            }
        )
        locations = chunk.get("locations") or []
        if not locations:
            locations = [{"display_text": source_ref or title_text}]
        for loc_index, loc in enumerate(locations):
            display = str(loc.get("display_text") or loc.get("source_ref") or source_ref or title_text)
            loc_rows.append(
                {
                    "location_id": _location_id(chunk_id, loc_index, display),
                    "chunk_id": chunk_id,
                    "doc_id": doc_id,
                    "location_index": loc_index,
                    "page_start": loc.get("page_start"),
                    "page_end": loc.get("page_end"),
                    "page_numbers_json": dumps_json(loc.get("page_numbers")) if loc.get("page_numbers") else None,
                    "slide_start": loc.get("slide_start"),
                    "slide_end": loc.get("slide_end"),
                    "sheet_name": loc.get("sheet_name"),
                    "cell_range": loc.get("cell_range"),
                    "heading_path": title_text,
                    "bbox_json": dumps_json(loc.get("bbox")) if loc.get("bbox") else None,
                    "source_refs_json": dumps_json([display]),
                    "display_text": display,
                    "metadata_json": dumps_json(loc.get("metadata") or {}),
                }
            )

    for idx, row in enumerate(chunk_rows):
        row["prev_chunk_id"] = chunk_ids[idx - 1] if idx > 0 else None
        row["next_chunk_id"] = chunk_ids[idx + 1] if idx + 1 < len(chunk_ids) else None

    if chunk_rows:
        conn.executemany(
            """
            INSERT INTO chunks (
                chunk_id, dataset_id, doc_id, chunk_index, content, content_type,
                title_path, summary, token_count, content_hash, prev_chunk_id,
                next_chunk_id, source_ref, metadata_json, created_at
            ) VALUES (
                :chunk_id, :dataset_id, :doc_id, :chunk_index, :content, :content_type,
                :title_path, :summary, :token_count, :content_hash, :prev_chunk_id,
                :next_chunk_id, :source_ref, :metadata_json, :created_at
            )
            """,
            chunk_rows,
        )
    if loc_rows:
        conn.executemany(
            """
            INSERT INTO chunk_locations (
                location_id, chunk_id, doc_id, location_index, page_start, page_end,
                page_numbers_json, slide_start, slide_end, sheet_name, cell_range,
                heading_path, bbox_json, source_refs_json, display_text, metadata_json
            ) VALUES (
                :location_id, :chunk_id, :doc_id, :location_index, :page_start, :page_end,
                :page_numbers_json, :slide_start, :slide_end, :sheet_name, :cell_range,
                :heading_path, :bbox_json, :source_refs_json, :display_text, :metadata_json
            )
            """,
            loc_rows,
        )
    conn.execute(
        "UPDATE documents SET chunk_count = ?, updated_at = ? WHERE doc_id = ?",
        (len(chunk_rows), now_iso(), doc_id),
    )
    return len(chunk_rows), len(loc_rows)


def _extract_pdf_pages(path: Path) -> tuple[list[dict[str, Any]], str, str]:
    """Extract PDF text deterministically with the declared PyMuPDF parser.

    Do not switch parsers based on whichever optional package happens to be
    installed on a machine: that changes reading order, chunks and hashes.
    """
    try:
        import fitz  # type: ignore
    except Exception as exc:
        raise RuntimeError("PyMuPDF is required for deterministic PDF ingestion") from exc

    parser_version = str(getattr(fitz, "VersionBind", "") or getattr(fitz, "__version__", "unknown"))
    try:
        pages: list[dict[str, Any]] = []
        with fitz.open(str(path)) as document:
            for page_index in range(document.page_count):
                page = document.load_page(page_index)
                rect = page.rect
                words = [
                    {
                        "text": str(item[4]).strip(),
                        "x_min": float(item[0]),
                        "y_min": float(item[1]),
                        "x_max": float(item[2]),
                        "y_max": float(item[3]),
                    }
                    for item in page.get_text("words", sort=True)
                    if len(item) >= 5 and str(item[4]).strip()
                ]
                text = str(page.get_text("text", sort=True) or "").strip()
                pages.append(
                    {
                        "page_number": page_index + 1,
                        "text": text,
                        "word_count": len(words),
                        "bbox": [0, 0, float(rect.width), float(rect.height)],
                        "lines": _group_pdf_words_into_lines(words),
                    }
                )
        return pages, "pymupdf", parser_version
    except Exception as exc:
        raise RuntimeError(f"Unable to extract PDF text with PyMuPDF from {path}: {exc}") from exc


def _pdf_text_quality(pages: list[dict[str, Any]]) -> dict[str, Any]:
    meaningful_by_page = [
        len(re.findall(r"[A-Za-z0-9\u3400-\u9fff]", str(page.get("text") or "")))
        for page in pages
    ]
    page_count = len(pages)
    meaningful_chars = sum(meaningful_by_page)
    text_page_count = sum(1 for count in meaningful_by_page if count >= MIN_PDF_TEXT_PAGE_CHARS)
    text_page_coverage = text_page_count / max(1, page_count)
    low_text_pages = [
        index
        for index, count in enumerate(meaningful_by_page, start=1)
        if count < MIN_PDF_TEXT_PAGE_CHARS
    ]
    minimum_total = max(
        MIN_PDF_MEANINGFUL_CHARS,
        min(page_count * MIN_PDF_MEANINGFUL_CHARS_PER_PAGE, 300),
    )
    reasons: list[str] = []
    if page_count == 0:
        reasons.append("PDF has no pages")
    if meaningful_chars == 0:
        reasons.append("No extractable text was found")
    elif meaningful_chars < minimum_total:
        reasons.append(
            f"Only {meaningful_chars} meaningful characters were extracted; minimum is {minimum_total}"
        )
    if page_count >= 2 and text_page_coverage < MIN_PDF_TEXT_PAGE_COVERAGE:
        reasons.append(
            f"Only {text_page_count}/{page_count} pages contain enough text "
            f"({text_page_coverage:.0%} coverage)"
        )
    return {
        "status": "needs_ocr" if reasons else "passed",
        "needs_ocr": bool(reasons),
        "reasons": reasons,
        "page_count": page_count,
        "meaningful_chars": meaningful_chars,
        "minimum_meaningful_chars": minimum_total,
        "text_page_count": text_page_count,
        "text_page_coverage": text_page_coverage,
        "low_text_pages": low_text_pages,
    }

def _group_pdf_words_into_lines(words: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sorted_words = sorted(words, key=lambda word: ((word["y_min"] + word["y_max"]) / 2, word["x_min"]))
    grouped: list[list[dict[str, Any]]] = []
    centers: list[float] = []
    for word in sorted_words:
        center = (word["y_min"] + word["y_max"]) / 2
        if grouped and abs(center - centers[-1]) <= 3.5:
            grouped[-1].append(word)
            centers[-1] = (centers[-1] * (len(grouped[-1]) - 1) + center) / len(grouped[-1])
        else:
            grouped.append([word])
            centers.append(center)

    lines: list[dict[str, Any]] = []
    for line_words in grouped:
        words_by_x = sorted(line_words, key=lambda word: word["x_min"])
        lines.append(
            {
                "text": " ".join(word["text"] for word in words_by_x).strip(),
                "bbox": [
                    min(word["x_min"] for word in words_by_x),
                    min(word["y_min"] for word in words_by_x),
                    max(word["x_max"] for word in words_by_x),
                    max(word["y_max"] for word in words_by_x),
                ],
            }
        )
    return lines


def _split_long_text(text: str, max_chars: int = DEFAULT_MAX_PDF_CHARS) -> list[str]:
    text = text.strip()
    if len(text) <= max_chars:
        return [text] if text else []
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n+", text) if p.strip()]
    if len(paragraphs) <= 1:
        paragraphs = [p.strip() for p in re.split(r"(?<=[。！？?!])\s+", text) if p.strip()]
    out: list[str] = []
    current = ""
    for para in paragraphs:
        candidate = f"{current}\n\n{para}".strip() if current else para
        if current and len(candidate) > max_chars:
            out.append(current)
            current = para
        else:
            current = candidate
    if current:
        out.append(current)
    return out


_SPEAKER_SEGMENT_RE = re.compile(r"发[言⾔]人\s*\d+\s+\d{2}:\d{2}:\d{2}")


def _is_pdf_footer_text(text: str) -> bool:
    normalized = unicodedata.normalize("NFKC", text)
    return bool(re.search(r"知识星球|前沿信息收录|VX[:：]", normalized, flags=re.I))


def _union_line_bbox(lines: list[dict[str, Any]]) -> list[float] | None:
    bboxes = [line.get("bbox") for line in lines if isinstance(line.get("bbox"), list) and len(line["bbox"]) == 4]
    if not bboxes:
        return None
    return [
        float(min(bbox[0] for bbox in bboxes)),
        float(min(bbox[1] for bbox in bboxes)),
        float(max(bbox[2] for bbox in bboxes)),
        float(max(bbox[3] for bbox in bboxes)),
    ]


def _speaker_segments_from_lines(page_lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not page_lines:
        return []

    segments: list[dict[str, Any]] = []
    current: list[dict[str, Any]] = []
    prefix: list[dict[str, Any]] = []

    def finish(lines: list[dict[str, Any]]) -> None:
        content_lines = [
            str(line.get("text") or "").strip()
            for line in lines
            if str(line.get("text") or "").strip() and not _is_pdf_footer_text(str(line.get("text") or ""))
        ]
        text = "\n".join(content_lines).strip()
        if len(text) >= 40:
            content_line_set = set(content_lines)
            bbox_lines = [
                line
                for line in lines
                if str(line.get("text") or "").strip() in content_line_set
                and not _is_pdf_footer_text(str(line.get("text") or ""))
            ]
            segments.append({"text": text, "bbox": _union_line_bbox(bbox_lines)})

    for line in page_lines:
        text = unicodedata.normalize("NFKC", str(line.get("text") or "")).strip()
        if not text:
            continue
        if _SPEAKER_SEGMENT_RE.search(text):
            if current:
                finish(current)
            elif prefix:
                finish(prefix)
            current = [line]
            prefix = []
        elif current:
            current.append(line)
        else:
            prefix.append(line)

    if current:
        finish(current)
    elif prefix:
        finish(prefix)
    return segments


def _speaker_segments(page_text: str, page_lines: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    line_segments = _speaker_segments_from_lines(page_lines or [])
    if line_segments:
        return line_segments

    page_text = unicodedata.normalize("NFKC", page_text)
    marker = re.compile(r"(发[言⾔]人\s*\d+\s+\d{2}:\d{2}:\d{2})")
    parts = marker.split(page_text)
    if len(parts) <= 1:
        return []
    segments: list[dict[str, Any]] = []
    prefix = parts[0].strip()
    for i in range(1, len(parts), 2):
        head = parts[i].strip()
        body = parts[i + 1].strip() if i + 1 < len(parts) else ""
        text = f"{head}\n{body}".strip()
        if len(text) >= 40:
            segments.append({"text": text, "bbox": None})
    if prefix and len(prefix) >= 80:
        segments.insert(0, {"text": prefix, "bbox": None})
    return segments


def ingest_pdf(conn: sqlite3.Connection, *, dataset_id: str, doc_id: str, path: Path) -> DocumentIngestResult:
    pages, method, parser_version = _extract_pdf_pages(path)
    text_quality = _pdf_text_quality(pages)
    page_rows = []
    chunks: list[dict[str, Any]] = []
    total_chars = 0
    for page in pages:
        text = page["text"]
        page_no = int(page["page_number"])
        total_chars += len(text)
        page_rows.append(
            {
                "page_id": sha256_text(f"{doc_id}\0page\0{page_no}")[:40],
                "dataset_id": dataset_id,
                "doc_id": doc_id,
                "page_number": page_no,
                "text": text,
                "char_count": len(text),
                "word_count": int(page.get("word_count") or len(text.split())),
                "extraction_method": method,
                "bbox_json": dumps_json(page.get("bbox")) if page.get("bbox") else None,
                "metadata_json": dumps_json(
                    {
                        "source": "private_fund_directory_ingest",
                        "parser_name": method,
                        "parser_version": parser_version,
                        "text_quality": text_quality,
                    }
                ),
            }
        )
        if text:
            page_segments = _split_long_text(text)
            for part_index, segment in enumerate(page_segments, start=1):
                title = f"{path.stem} p.{page_no}" if len(page_segments) == 1 else f"{path.stem} p.{page_no} part {part_index}"
                chunks.append(
                    {
                        "content": segment,
                        "content_type": "pdf_page",
                        "title_path": [path.stem, f"page {page_no}"],
                        "summary": segment[:240],
                        "source_ref": f"{path.name} p.{page_no}",
                        "metadata": {"extraction_method": method, "part_index": part_index},
                        "locations": [
                            {
                                "page_start": page_no,
                                "page_end": page_no,
                                "display_text": f"{path.name} p.{page_no}",
                                "bbox": page.get("bbox"),
                                "metadata": {"part_index": part_index},
                            }
                        ],
                        "title": title,
                    }
                )
            for turn_index, segment in enumerate(_speaker_segments(text, page.get("lines") or []), start=1):
                segment_text = str(segment.get("text") or "").strip()
                if not segment_text:
                    continue
                chunks.append(
                    {
                        "content": segment_text,
                        "content_type": "pdf_speaker_turn",
                        "title_path": [path.stem, f"page {page_no}", f"speaker turn {turn_index}"],
                        "summary": segment_text[:240],
                        "source_ref": f"{path.name} p.{page_no}, turn {turn_index}",
                        "metadata": {"extraction_method": method, "turn_index": turn_index},
                        "locations": [
                            {
                                "page_start": page_no,
                                "page_end": page_no,
                                "display_text": f"{path.name} p.{page_no}",
                                "bbox": segment.get("bbox") or page.get("bbox"),
                                "metadata": {"turn_index": turn_index},
                            }
                        ],
                    }
                )

    if page_rows:
        conn.executemany(
            """
            INSERT INTO pdf_pages (
                page_id, dataset_id, doc_id, page_number, text, char_count,
                word_count, extraction_method, bbox_json, metadata_json
            ) VALUES (
                :page_id, :dataset_id, :doc_id, :page_number, :text, :char_count,
                :word_count, :extraction_method, :bbox_json, :metadata_json
            )
            """,
            page_rows,
        )

    parser_metadata = {
        "parser_name": method,
        "parser_version": parser_version,
        "text_quality": text_quality,
    }
    if text_quality["needs_ocr"]:
        reason = "; ".join(text_quality["reasons"])
        return DocumentIngestResult(
            doc_id=doc_id,
            filename=path.name,
            file_type="pdf",
            status="needs_ocr",
            chunk_count=0,
            location_count=0,
            pdf_page_count=len(pages),
            error_message=f"OCR required: {reason}",
            parser_name=method,
            parser_version=parser_version,
            parser_metadata=parser_metadata,
        )

    summary = (
        f"PDF document: {path.name}\n"
        f"Pages: {len(pages)}\n"
        f"Extraction method: {method}\n"
        f"Total text characters: {total_chars}\n"
        "OCR required: no, deterministic text-quality gate passed.\n"
    )
    chunks.insert(
        0,
        {
            "content": summary,
            "content_type": "pdf_document_summary",
            "title_path": [path.stem, "document summary"],
            "summary": summary,
            "source_ref": path.name,
            "metadata": {"page_count": len(pages), "extraction_method": method, "total_chars": total_chars},
            "locations": [{"page_start": 1 if pages else None, "page_end": len(pages) or None, "display_text": path.name}],
        },
    )

    chunk_count, location_count = _write_chunks(conn, dataset_id=dataset_id, doc_id=doc_id, chunks=chunks)
    return DocumentIngestResult(
        doc_id=doc_id,
        filename=path.name,
        file_type="pdf",
        status="indexed",
        chunk_count=chunk_count,
        location_count=location_count,
        pdf_page_count=len(pages),
        parser_name=method,
        parser_version=parser_version,
        parser_metadata=parser_metadata,
    )


def _col_letter(index: int) -> str:
    letters = ""
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _cell_ref(row: int, col: int) -> str:
    return f"{_col_letter(col)}{row}"


def _range_ref(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    return f"{_cell_ref(min_row, min_col)}:{_cell_ref(max_row, max_col)}"


def _is_formula(value: Any) -> bool:
    if isinstance(value, str):
        return value.startswith("=")
    return type(value).__name__ in {"ArrayFormula", "DataTableFormula"}


def _formula_details(value: Any) -> tuple[bool, Optional[str], Optional[str], dict[str, Any]]:
    if isinstance(value, str) and value.startswith("="):
        return True, "standard", value, {}
    formula_type = type(value).__name__
    if formula_type not in {"ArrayFormula", "DataTableFormula"}:
        return False, None, None, {}
    attributes = json_safe(getattr(value, "__dict__", {}))
    metadata = attributes if isinstance(attributes, dict) else {"attributes": attributes}
    expression = getattr(value, "text", None)
    if isinstance(expression, str) and expression:
        formula_text = expression
    else:
        formula_text = dumps_json({"formula_type": formula_type, **metadata})
    stable_type = "array" if formula_type == "ArrayFormula" else "data_table"
    return True, stable_type, formula_text, metadata


def _formula_cache_status(is_formula: bool, cached: Any) -> str:
    if not is_formula:
        return "not_applicable"
    if cached is None or cached == "":
        return "missing"
    if _is_formula(cached) or not isinstance(cached, (str, int, float, bool, datetime, date)):
        return "unavailable"
    if isinstance(cached, str) and cached.startswith("#"):
        return "error"
    return "present"


def _numeric_value(value: Any) -> Optional[float]:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    text = normalize_text(value).replace(",", "")
    if not text:
        return None
    percent = text.endswith("%")
    if percent:
        text = text[:-1]
    try:
        number = float(text)
        return number / 100.0 if percent else number
    except ValueError:
        return None


def _period_from_label(label: str) -> str:
    label = normalize_text(label)
    # A financial value such as ``12068.32666`` can contain a ``20xx``
    # substring after the decimal point.  Treat only standalone period tokens
    # as years; otherwise long statement rows inherit fictitious periods.
    patterns = [
        r"(?<![\d.])([1-4]Q\s*20\d{2})(?![\d.])",
        r"(?<![\d.])(20\d{2}\s*[1-4]Q)(?![\d.])",
        r"(?<![\d.])(Q[1-4]\s*[-/. ]?\s*20\d{2})(?![\d.])",
        r"(?<![\d.])(Q[1-4]\s*[-/. ]?\s*\d{2})(?![\d.])",
        r"(?<![\d.])(FY\s*20\d{2})(?![\d.])",
        r"(?<![\d.])([1-4]Q\s*\d{2})(?![\d.])",
        r"(?<![\d.])(FY\s*\d{2})(?![\d.])",
        r"(?<![\d.])(20\d{2}\s*[EQAF]?)(?![\d.])",
    ]
    for pattern in patterns:
        match = re.search(pattern, label, flags=re.IGNORECASE)
        if match:
            period = normalize_text(match.group(1))
            year_match = re.search(r"(?<!\d)(20\d{2}|\d{2})(?!\d)", period)
            if year_match:
                year = int(year_match.group(1))
                if year < 100:
                    year = 1900 + year if year >= 70 else 2000 + year
                if not 1990 <= year <= 2050:
                    continue
            return period
    return ""


def _repair_metric_fact_periods(conn: sqlite3.Connection) -> None:
    """Rebuild fact periods from column headers after period-parser upgrades."""

    quarter_headers: dict[tuple[str, str, int], list[tuple[int, str]]] = {}
    for cell in conn.execute(
        """
        SELECT doc_id, sheet_name, row_index, col_index, display_value
        FROM excel_cells
        WHERE display_value IS NOT NULL AND display_value <> ''
        """
    ):
        display = str(cell["display_value"] or "")
        period = _period_from_label(display)
        if not period or not re.search(r"(?:[1-4]\s*Q|Q\s*[1-4])", period, re.IGNORECASE):
            continue
        key = (str(cell["doc_id"]), str(cell["sheet_name"]), int(cell["col_index"]))
        quarter_headers.setdefault(key, []).append((int(cell["row_index"]), period))
    for headers in quarter_headers.values():
        headers.sort()

    rows = conn.execute(
        """
        SELECT f.fact_id, f.doc_id, f.sheet_name, f.period AS fact_period,
               c.cell_id, c.period AS cell_period, c.row_index, c.col_index, c.col_label
        FROM metric_facts f
        JOIN excel_cells c
          ON c.doc_id=f.doc_id AND c.sheet_name=f.sheet_name AND c.cell_ref=f.cell_ref
        """
    ).fetchall()
    for row in rows:
        direct = _period_from_label(str(row["col_label"] or ""))
        period = direct if re.search(r"(?:[1-4]\s*Q|Q\s*[1-4])", direct, re.IGNORECASE) else ""
        if not period:
            key = (str(row["doc_id"]), str(row["sheet_name"]), int(row["col_index"]))
            preceding = [
                value
                for header_row, value in quarter_headers.get(key, [])
                if header_row <= int(row["row_index"])
            ]
            period = preceding[-1] if preceding else ""
        if not period:
            period = direct or _period_from_label(str(row["fact_period"] or ""))
        if str(row["fact_period"] or "") != period:
            conn.execute(
                "UPDATE metric_facts SET period=? WHERE fact_id=?",
                (period or None, row["fact_id"]),
            )
        if str(row["cell_period"] or "") != period:
            conn.execute(
                "UPDATE excel_cells SET period=? WHERE cell_id=?",
                (period or None, row["cell_id"]),
            )


def _looks_like_period_label(label: str) -> bool:
    return bool(_period_from_label(label))


def _unit_from_text(text: str) -> str:
    text = normalize_text(text)
    if "%" in text:
        return "%"
    for unit in ("CNYm", "RMBm", "USDm", "GWh", "MWh", "Wh", "MW", "GW", "元/Wh"):
        if unit.lower() in text.lower():
            return unit
    return ""


def _unit_from_number_format(number_format: str) -> str:
    return "%" if "%" in str(number_format or "") else ""


def _sheet_role(sheet_name: str, sample_text: str) -> str:
    normalized_name = normalize_text(sheet_name).lower()
    if any(
        marker in normalized_name
        for marker in ("upload", "download", "raw data", "raw_data", "bloomberg", "__fdscache__")
    ):
        return "raw_upload"
    text = f"{sheet_name} {sample_text}".lower()
    checks = [
        ("valuation_dcf", ("dcf", "wacc", "terminal value", "valuation")),
        ("sensitivity", ("sensitivity", "敏感")),
        ("driver_model", ("driver", "assumption", "asp", "shipment")),
        ("financial_statement", ("pl_bs_cfs", "income statement", "balance sheet", "cash flow")),
        ("quarterly_results", ("qoq", "results", "quarter")),
        ("output_table", ("table", "snapshot", "breakdown")),
        ("chart_data", ("chart",)),
        ("raw_upload", ("upload", "bloomberg", "@")),
        ("resource_note", ("resource", "products", "产品")),
    ]
    for role, words in checks:
        if any(word in text for word in words):
            return role
    return "worksheet"


def _region_type(sheet_name: str, values: list[Any], formula_count: int) -> str:
    text = normalize_text(" ".join(cell_display(v, 80) for v in values)).lower()
    mapping = [
        ("valuation_dcf", ("dcf", "wacc", "terminal value", "valuation")),
        ("sensitivity", ("sensitivity", "敏感")),
        ("income_statement", ("revenue", "gross profit", "net profit", "eps", "income statement")),
        ("cash_flow", ("free cash flow", "cash flow", "fcf")),
        ("driver_assumption", ("driver", "asp", "shipment", "assumption", "orders")),
        ("business_breakdown", ("breakdown", "segment", "contribution")),
        ("note", ("note", "摘要", "q&a")),
    ]
    for region_type, words in mapping:
        if any(word in text for word in words):
            return region_type
    if formula_count / max(1, len(values)) >= 0.25:
        return "formula_block"
    if len(values) >= 6:
        return "table"
    return "text_block"


def _workbook_type(sheet_summaries: list[dict[str, Any]]) -> str:
    formulas = sum(int(s["formula_count"]) for s in sheet_summaries)
    non_empty = sum(int(s["non_empty_cell_count"]) for s in sheet_summaries)
    density = formulas / max(1, non_empty)
    role_text = " ".join(str(s.get("sheet_role") or "") for s in sheet_summaries)
    if density >= 0.15 or any(role in role_text for role in ("valuation_dcf", "driver_model", "sensitivity")):
        return "valuation_model"
    if len(sheet_summaries) <= 3 and density < 0.1:
        return "simple_table"
    return "financial_workbook"


def _nonempty_cells(ws) -> dict[tuple[int, int], Any]:
    cells: dict[tuple[int, int], Any] = {}
    for key, cell in getattr(ws, "_cells", {}).items():
        value = cell.value
        if value is not None and cell_display(value):
            cells[(cell.row, cell.column)] = value
    return cells


_NON_RESEARCH_SHEET_NAMES = {
    "__fdscache__",
    "db disclaimer",
}


def _is_non_research_sheet(sheet_name: str) -> bool:
    """Return True for vendor cache/legal boilerplate sheets with no research value."""
    return normalize_text(sheet_name).casefold() in _NON_RESEARCH_SHEET_NAMES


def _sheet_bounds(cells: dict[tuple[int, int], Any]) -> Optional[tuple[int, int, int, int]]:
    if not cells:
        return None
    rows = [row for row, _ in cells]
    cols = [col for _, col in cells]
    return min(rows), min(cols), max(rows), max(cols)


def _group_sorted(values: list[int], gap: int = 1) -> list[tuple[int, int]]:
    if not values:
        return []
    values = sorted(set(values))
    groups: list[tuple[int, int]] = []
    start = prev = values[0]
    for value in values[1:]:
        if value - prev <= gap + 1:
            prev = value
        else:
            groups.append((start, prev))
            start = prev = value
    groups.append((start, prev))
    return groups


def _detect_regions(cells: dict[tuple[int, int], Any]) -> list[tuple[int, int, int, int]]:
    row_groups = _group_sorted([row for row, _ in cells], gap=1)
    regions: list[tuple[int, int, int, int]] = []
    for row_start, row_end in row_groups:
        cols = [col for row, col in cells if row_start <= row <= row_end]
        for col_start, col_end in _group_sorted(cols, gap=1):
            region_cells = [
                (row, col)
                for row, col in cells
                if row_start <= row <= row_end and col_start <= col <= col_end
            ]
            if region_cells:
                rows = [row for row, _ in region_cells]
                cols_in_region = [col for _, col in region_cells]
                regions.append((min(rows), min(cols_in_region), max(rows), max(cols_in_region)))
    return regions


def _nearest_left_label(row_text_cols: dict[int, list[tuple[int, str]]], row: int, col: int) -> str:
    cols = row_text_cols.get(row) or []
    indexes = [item[0] for item in cols]
    pos = bisect.bisect_left(indexes, col) - 1
    while pos >= 0:
        label = cols[pos][1]
        if label:
            return label
        pos -= 1
    return ""


def _nearest_top_label(col_text_rows: dict[int, list[tuple[int, str]]], row: int, col: int) -> str:
    rows = col_text_rows.get(col) or []
    indexes = [item[0] for item in rows]
    pos = bisect.bisect_left(indexes, row) - 1
    while pos >= 0:
        label = rows[pos][1]
        if label:
            return label
        pos -= 1
    return ""


def _sample_labels(cells: dict[tuple[int, int], Any], max_items: int = DEFAULT_MAX_REGION_LABELS) -> list[str]:
    labels: list[str] = []
    for _, value in sorted(cells.items(), key=lambda item: item[0]):
        text = cell_display(value, 80)
        if text and not _is_formula(value) and not re.fullmatch(r"[-+]?\d+(\.\d+)?%?", text):
            labels.append(text)
        if len(labels) >= max_items:
            break
    return labels


def ingest_excel(conn: sqlite3.Connection, *, dataset_id: str, doc_id: str, path: Path) -> DocumentIngestResult:
    try:
        import openpyxl  # type: ignore
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required for Excel ingestion") from exc

    parser_name = "openpyxl"
    parser_version = str(getattr(openpyxl, "__version__", "unknown"))

    wb_formula = load_workbook(path, data_only=False, read_only=False, keep_links=False)
    wb_values = load_workbook(path, data_only=True, read_only=False, keep_links=False)

    sheet_rows: list[dict[str, Any]] = []
    region_rows: list[dict[str, Any]] = []
    cell_rows: list[dict[str, Any]] = []
    fact_rows: list[dict[str, Any]] = []
    chunks: list[dict[str, Any]] = []

    for sheet_index, ws in enumerate(wb_formula.worksheets, start=1):
        if _is_non_research_sheet(ws.title):
            continue
        values_ws = wb_values[ws.title] if ws.title in wb_values.sheetnames else None
        cells = _nonempty_cells(ws)
        bounds = _sheet_bounds(cells)
        formula_count = sum(1 for value in cells.values() if _is_formula(value))
        non_empty = len(cells)
        if bounds:
            min_row, min_col, max_row, max_col = bounds
            used_range = _range_ref(min_row, min_col, max_row, max_col)
            row_count = max_row - min_row + 1
            col_count = max_col - min_col + 1
        else:
            min_row = min_col = max_row = max_col = 0
            used_range = ""
            row_count = col_count = 0
        labels = _sample_labels(cells)
        role = _sheet_role(ws.title, " ".join(labels))
        sheet_unit = ""
        for label in labels[:10]:
            sheet_unit = _unit_from_text(label)
            if sheet_unit:
                break
        formula_density = formula_count / max(1, non_empty)
        sheet_summary = (
            f"Excel sheet: {ws.title}\n"
            f"Role: {role}\n"
            f"Used range: {used_range or 'empty'}\n"
            f"Non-empty cells: {non_empty}; formulas: {formula_count}; formula density: {formula_density:.2%}\n"
            f"Key labels: {', '.join(labels[:20])}"
        )
        sheet_id = sha256_text(f"{doc_id}\0sheet\0{ws.title}")[:40]
        sheet_rows.append(
            {
                "sheet_id": sheet_id,
                "dataset_id": dataset_id,
                "doc_id": doc_id,
                "sheet_index": sheet_index,
                "sheet_name": ws.title,
                "sheet_role": role,
                "sheet_state": ws.sheet_state,
                "used_range": used_range,
                "row_count": row_count,
                "col_count": col_count,
                "non_empty_cell_count": non_empty,
                "formula_count": formula_count,
                "formula_density": formula_density,
                "summary": sheet_summary,
                "header_json": dumps_json(labels),
                "metadata_json": dumps_json(
                    {
                        "freeze_panes": str(ws.freeze_panes or ""),
                        "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
                    }
                ),
            }
        )
        if non_empty:
            chunks.append(
                {
                    "content": sheet_summary,
                    "content_type": "excel_sheet_summary",
                    "title_path": [path.stem, ws.title, "sheet summary"],
                    "summary": sheet_summary[:240],
                    "source_ref": f"{path.name} {ws.title}!{used_range}",
                    "metadata": {"sheet_role": role, "used_range": used_range},
                    "locations": [
                        {
                            "sheet_name": ws.title,
                            "cell_range": used_range,
                            "display_text": f"{ws.title}!{used_range}",
                            "metadata": {"sheet_role": role},
                        }
                    ],
                }
            )

        row_text_cols: dict[int, list[tuple[int, str]]] = {}
        col_text_rows: dict[int, list[tuple[int, str]]] = {}
        col_period_rows: dict[int, list[tuple[int, str]]] = {}
        for (row, col), value in cells.items():
            is_formula, _, _, _ = _formula_details(value)
            cached_for_label = values_ws.cell(row, col).value if is_formula and values_ws is not None else None
            cache_status = _formula_cache_status(is_formula, cached_for_label)
            label_value = cached_for_label if cache_status in {"present", "error"} else None
            text = cell_display(label_value if is_formula else value, 120)
            raw_text = cell_display(value, 120)
            if text and not is_formula and _numeric_value(value) is None:
                row_text_cols.setdefault(row, []).append((col, text))
            if text and _looks_like_period_label(text):
                col_period_rows.setdefault(col, []).append((row, text))
            if text and (_numeric_value(text) is None or _looks_like_period_label(text) or row <= max(5, min_row + 4)):
                col_text_rows.setdefault(col, []).append((row, text))
            elif raw_text and not is_formula and _numeric_value(raw_text) is None:
                col_text_rows.setdefault(col, []).append((row, raw_text))
        for items in row_text_cols.values():
            items.sort(key=lambda x: x[0])
        for items in col_text_rows.values():
            items.sort(key=lambda x: x[0])
        for items in col_period_rows.values():
            items.sort(key=lambda x: x[0])

        for (row, col), value in sorted(cells.items(), key=lambda item: item[0]):
            cached = values_ws.cell(row, col).value if values_ws is not None else None
            is_formula, formula_type, formula, formula_metadata = _formula_details(value)
            cache_status = _formula_cache_status(is_formula, cached)
            display_source = cached if is_formula and cache_status in {"present", "error"} else value
            display = cell_display(display_source, 200)
            row_label = _nearest_left_label(row_text_cols, row, col)
            col_label = _nearest_top_label(col_text_rows, row, col)
            period_label = _nearest_top_label(col_period_rows, row, col)
            period = (
                _period_from_label(period_label)
                or _period_from_label(col_label)
                or _period_from_label(display)
            )
            number_format = str(ws.cell(row, col).number_format or "")
            unit = (
                _unit_from_text(row_label)
                or _unit_from_text(col_label)
                or _unit_from_text(display)
                or _unit_from_number_format(number_format)
                or sheet_unit
            )
            numeric = _numeric_value(cached if is_formula else value)
            cell_ref = _cell_ref(row, col)
            value_type = f"formula_{formula_type}" if is_formula else type(value).__name__
            cell_id = sha256_text(f"{doc_id}\0{ws.title}\0{cell_ref}")[:40]
            cell_rows.append(
                {
                    "cell_id": cell_id,
                    "dataset_id": dataset_id,
                    "doc_id": doc_id,
                    "sheet_name": ws.title,
                    "cell_ref": cell_ref,
                    "row_index": row,
                    "col_index": col,
                    "value_type": value_type,
                    "display_value": display,
                    "raw_value": cell_display(value, 500),
                    "numeric_value": numeric,
                    "formula": formula,
                    "cached_value": cell_display(cached, 500) if is_formula and cached is not None else None,
                    "number_format": number_format,
                    "row_label": row_label,
                    "col_label": col_label,
                    "period": period,
                    "unit": unit,
                    "is_formula": 1 if is_formula else 0,
                    "formula_type": formula_type,
                    "formula_cache_status": cache_status,
                    "metadata_json": dumps_json(
                        {
                            "sheet_role": role,
                            "formula_type": formula_type,
                            "formula_cache_status": cache_status,
                            "formula_metadata": formula_metadata,
                        }
                    ),
                }
            )
            if numeric is not None and row_label:
                fact_id = sha256_text(f"{doc_id}\0{ws.title}\0{cell_ref}\0{row_label}\0{period}")[:40]
                quality_issues = ["metric_name_inferred_from_nearest_left_label"]
                if not period:
                    quality_issues.append("period_missing")
                if not unit:
                    quality_issues.append("unit_missing")
                if is_formula and cache_status != "present":
                    quality_issues.append(f"formula_cache_{cache_status}")
                quality_status = (
                    "candidate_complete"
                    if period and unit and (not is_formula or cache_status == "present")
                    else "review_required"
                )
                confidence = 0.75 if quality_status == "candidate_complete" else (0.65 if period else 0.55)
                fact_rows.append(
                    {
                        "fact_id": fact_id,
                        "dataset_id": dataset_id,
                        "doc_id": doc_id,
                        "metric_name": row_label,
                        "metric_alias": normalize_text(row_label).lower(),
                        "period": period,
                        "value_text": display,
                        "value_numeric": numeric,
                        "unit": unit,
                        "sheet_name": ws.title,
                        "cell_ref": cell_ref,
                        "source_range": f"{ws.title}!{cell_ref}",
                        "formula": formula,
                        "confidence": confidence,
                        "fact_status": "candidate",
                        "quality_status": quality_status,
                        "quality_issues_json": dumps_json(quality_issues),
                        "metadata_json": dumps_json(
                            {
                                "col_label": col_label,
                                "sheet_role": role,
                                "extraction_method": "nearest_left_metric_and_nearest_top_period",
                                "fact_status": "candidate",
                                "quality_status": quality_status,
                                "quality_issues": quality_issues,
                                "formula_type": formula_type,
                                "formula_cache_status": cache_status,
                            }
                        ),
                    }
                )

        for region_index, (r1, c1, r2, c2) in enumerate(_detect_regions(cells), start=1):
            region_values = [
                cells[(row, col)]
                for row in range(r1, r2 + 1)
                for col in range(c1, c2 + 1)
                if (row, col) in cells
            ]
            region_formula_count = sum(1 for value in region_values if _is_formula(value))
            region_range = _range_ref(r1, c1, r2, c2)
            region_type = _region_type(ws.title, region_values, region_formula_count)
            region_labels = _sample_labels({(idx, 1): v for idx, v in enumerate(region_values, start=1)})
            region_summary = (
                f"Excel region: {ws.title}!{region_range}\n"
                f"Sheet role: {role}\n"
                f"Region type: {region_type}\n"
                f"Rows: {r2 - r1 + 1}; columns: {c2 - c1 + 1}; non-empty cells: {len(region_values)}; "
                f"formulas: {region_formula_count}\n"
                f"Key labels: {', '.join(region_labels[:20])}"
            )
            region_id = sha256_text(f"{doc_id}\0region\0{ws.title}\0{region_range}\0{region_index}")[:40]
            region_rows.append(
                {
                    "region_id": region_id,
                    "dataset_id": dataset_id,
                    "doc_id": doc_id,
                    "sheet_name": ws.title,
                    "region_index": region_index,
                    "region_type": region_type,
                    "cell_range": region_range,
                    "row_count": r2 - r1 + 1,
                    "col_count": c2 - c1 + 1,
                    "non_empty_cell_count": len(region_values),
                    "formula_count": region_formula_count,
                    "formula_density": region_formula_count / max(1, len(region_values)),
                    "summary": region_summary,
                    "header_json": dumps_json(region_labels),
                    "metadata_json": dumps_json({"sheet_role": role}),
                }
            )
            chunks.append(
                {
                    "content": region_summary,
                    "content_type": "excel_region_summary",
                    "title_path": [path.stem, ws.title, region_type, region_range],
                    "summary": region_summary[:240],
                    "source_ref": f"{path.name} {ws.title}!{region_range}",
                    "metadata": {"sheet_role": role, "region_type": region_type, "cell_range": region_range},
                    "locations": [
                        {
                            "sheet_name": ws.title,
                            "cell_range": region_range,
                            "display_text": f"{ws.title}!{region_range}",
                            "metadata": {"sheet_role": role, "region_type": region_type},
                        }
                    ],
                }
            )

    workbook_type = _workbook_type(sheet_rows)
    total_formulas = sum(int(row["formula_count"]) for row in sheet_rows)
    total_cells = sum(int(row["non_empty_cell_count"]) for row in sheet_rows)
    if total_cells == 0:
        wb_formula.close()
        wb_values.close()
        raise RuntimeError(
            f"{path.name} contains no non-empty cells; empty workbooks are not indexed"
        )
    formula_cache_counts: dict[str, int] = {}
    for row in cell_rows:
        if row["is_formula"]:
            status = str(row["formula_cache_status"])
            formula_cache_counts[status] = formula_cache_counts.get(status, 0) + 1
    fact_quality_counts: dict[str, int] = {}
    for row in fact_rows:
        status = str(row["quality_status"])
        fact_quality_counts[status] = fact_quality_counts.get(status, 0) + 1
    workbook_summary = (
        f"Excel workbook: {path.name}\n"
        f"Workbook type: {workbook_type}\n"
        f"Sheets: {len(sheet_rows)}; non-empty cells: {total_cells}; formulas: {total_formulas}; "
        f"formula density: {total_formulas / max(1, total_cells):.2%}\n"
        f"Formula cache status: {dumps_json(formula_cache_counts)}\n"
        "Metric facts are heuristic candidates and require source-cell review.\n"
        "Sheet roles:\n"
        + "\n".join(f"- {row['sheet_name']}: {row['sheet_role']} ({row['used_range']})" for row in sheet_rows)
    )
    chunks.insert(
        0,
        {
            "content": workbook_summary,
            "content_type": "excel_workbook_summary",
            "title_path": [path.stem, "workbook summary"],
            "summary": workbook_summary[:240],
            "source_ref": path.name,
            "metadata": {
                "workbook_type": workbook_type,
                "sheet_count": len(sheet_rows),
                "parser_name": parser_name,
                "parser_version": parser_version,
                "formula_cache_status_counts": formula_cache_counts,
                "metric_fact_status": "candidate",
                "fact_quality_status_counts": fact_quality_counts,
            },
            "locations": [{"display_text": path.name, "metadata": {"workbook_type": workbook_type}}],
        },
    )

    conn.execute(
        """
        INSERT INTO excel_workbooks (
            workbook_id, dataset_id, doc_id, workbook_type, sheet_count,
            visible_sheet_count, formula_count, non_empty_cell_count,
            formula_density, metadata_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            sha256_text(f"{doc_id}\0workbook")[:40],
            dataset_id,
            doc_id,
            workbook_type,
            len(sheet_rows),
            sum(1 for ws in wb_formula.worksheets if ws.sheet_state == "visible"),
            total_formulas,
            total_cells,
            total_formulas / max(1, total_cells),
            dumps_json(
                {
                    "source": "private_fund_directory_ingest",
                    "parser_name": parser_name,
                    "parser_version": parser_version,
                    "formula_cache_status_counts": formula_cache_counts,
                    "fact_status": "candidate",
                    "fact_quality_status_counts": fact_quality_counts,
                }
            ),
        ),
    )
    if sheet_rows:
        conn.executemany(
            """
            INSERT INTO excel_sheets (
                sheet_id, dataset_id, doc_id, sheet_index, sheet_name, sheet_role,
                sheet_state, used_range, row_count, col_count, non_empty_cell_count,
                formula_count, formula_density, summary, header_json, metadata_json
            ) VALUES (
                :sheet_id, :dataset_id, :doc_id, :sheet_index, :sheet_name, :sheet_role,
                :sheet_state, :used_range, :row_count, :col_count, :non_empty_cell_count,
                :formula_count, :formula_density, :summary, :header_json, :metadata_json
            )
            """,
            sheet_rows,
        )
    if region_rows:
        conn.executemany(
            """
            INSERT INTO excel_regions (
                region_id, dataset_id, doc_id, sheet_name, region_index, region_type,
                cell_range, row_count, col_count, non_empty_cell_count, formula_count,
                formula_density, summary, header_json, metadata_json
            ) VALUES (
                :region_id, :dataset_id, :doc_id, :sheet_name, :region_index, :region_type,
                :cell_range, :row_count, :col_count, :non_empty_cell_count, :formula_count,
                :formula_density, :summary, :header_json, :metadata_json
            )
            """,
            region_rows,
        )
    if cell_rows:
        conn.executemany(
            """
            INSERT INTO excel_cells (
                cell_id, dataset_id, doc_id, sheet_name, cell_ref, row_index, col_index,
                value_type, display_value, raw_value, numeric_value, formula, cached_value,
                number_format, row_label, col_label, period, unit, is_formula,
                formula_type, formula_cache_status, metadata_json
            ) VALUES (
                :cell_id, :dataset_id, :doc_id, :sheet_name, :cell_ref, :row_index, :col_index,
                :value_type, :display_value, :raw_value, :numeric_value, :formula, :cached_value,
                :number_format, :row_label, :col_label, :period, :unit, :is_formula,
                :formula_type, :formula_cache_status, :metadata_json
            )
            """,
            cell_rows,
        )
    if fact_rows:
        conn.executemany(
            """
            INSERT INTO metric_facts (
                fact_id, dataset_id, doc_id, metric_name, metric_alias, period,
                value_text, value_numeric, unit, sheet_name, cell_ref, source_range,
                formula, confidence, fact_status, quality_status, quality_issues_json,
                metadata_json
            ) VALUES (
                :fact_id, :dataset_id, :doc_id, :metric_name, :metric_alias, :period,
                :value_text, :value_numeric, :unit, :sheet_name, :cell_ref, :source_range,
                :formula, :confidence, :fact_status, :quality_status, :quality_issues_json,
                :metadata_json
            )
            """,
            fact_rows,
        )

    chunk_count, location_count = _write_chunks(conn, dataset_id=dataset_id, doc_id=doc_id, chunks=chunks)
    wb_formula.close()
    wb_values.close()
    return DocumentIngestResult(
        doc_id=doc_id,
        filename=path.name,
        file_type=path.suffix.lower().lstrip("."),
        status="indexed",
        chunk_count=chunk_count,
        location_count=location_count,
        excel_sheet_count=len(sheet_rows),
        excel_region_count=len(region_rows),
        excel_cell_count=len(cell_rows),
        metric_fact_count=len(fact_rows),
        parser_name=parser_name,
        parser_version=parser_version,
        parser_metadata={
            "formula_cache_status_counts": formula_cache_counts,
            "fact_status": "candidate",
            "fact_quality_status_counts": fact_quality_counts,
        },
    )


def ingest_adapted_document(
    conn: sqlite3.Connection, *, dataset_id: str, doc_id: str, path: Path
) -> DocumentIngestResult:
    chunks = adapt_document(path)
    meaningful_chunks = [
        chunk
        for chunk in chunks
        if not str(chunk.get("content_type") or "").endswith("_document_summary")
        and normalize_text(chunk.get("content"))
    ]
    if not meaningful_chunks:
        raise RuntimeError(
            f"{path.name} contains no meaningful extractable content; summary-only documents are not indexed"
        )
    parser_name = "private_fund_format_adapters"
    parser_version = "1"
    parser_metadata = {
        "adapter_suffix": path.suffix.lower(),
        "meaningful_chunk_count": len(meaningful_chunks),
        "total_chunk_count": len(chunks),
    }
    for chunk in chunks:
        metadata = dict(chunk.get("metadata") or {})
        metadata.update({"parser_name": parser_name, "parser_version": parser_version})
        chunk["metadata"] = metadata
    chunk_count, location_count = _write_chunks(
        conn, dataset_id=dataset_id, doc_id=doc_id, chunks=chunks
    )
    return DocumentIngestResult(
        doc_id=doc_id,
        filename=path.name,
        file_type=path.suffix.lower().lstrip("."),
        status="indexed",
        chunk_count=chunk_count,
        location_count=location_count,
        parser_name=parser_name,
        parser_version=parser_version,
        parser_metadata=parser_metadata,
    )


def _doc_type_for_path(path: Path) -> str:
    """Compatibility wrapper returning only the controlled primary type."""

    return classify_document(build_document_preview(path)).doc_type


def _fallback_classification(
    *,
    company_name: str,
    company_ticker: str,
    error: str = "",
) -> DocumentClassification:
    classification = DocumentClassification(
        doc_type="other",
        confidence=0.0,
        company_name=company_name,
        company_ticker=company_ticker,
        company_confidence=0.55 if company_name else 0.0,
        classification_status="needs_review",
        method="classifier_fallback",
        company_method="inherited_project" if company_name else "not_detected",
        evidence=["文档分类器未能完成，归入其他并等待复核"],
    )
    classification.llm_error = error[:500]
    return classification


def _apply_classification_result(
    result: DocumentIngestResult,
    classification: DocumentClassification,
) -> None:
    result.doc_type = classification.doc_type
    result.doc_subtype = classification.doc_subtype
    result.doc_type_confidence = classification.confidence
    result.classification_status = classification.classification_status
    result.classification_method = classification.method
    result.company_name = classification.company_name
    result.company_ticker = classification.company_ticker
    result.company_confidence = classification.company_confidence


def _update_document_status(
    conn: sqlite3.Connection,
    doc_id: str,
    status: str,
    error: Optional[str] = None,
    *,
    parser_name: str = "",
    parser_version: str = "",
    parser_metadata: Optional[dict[str, Any]] = None,
) -> None:
    conn.execute(
        """
        UPDATE documents
        SET status = ?, error_message = ?, parser_name = NULLIF(?, ''),
            parser_version = NULLIF(?, ''), parser_metadata_json = ?, updated_at = ?
        WHERE doc_id = ?
        """,
        (
            status,
            error,
            parser_name,
            parser_version,
            dumps_json(parser_metadata or {}),
            now_iso(),
            doc_id,
        ),
    )


def _mark_removed_documents(
    conn: sqlite3.Connection,
    *,
    dataset_id: str,
    source_root: str,
    seen_logical_doc_ids: set[str],
) -> list[str]:
    removed: list[str] = []
    rows = conn.execute(
        """
        SELECT doc_id, logical_doc_id, source_root
        FROM documents
        WHERE dataset_id = ? AND source_type = 'local_directory'
          AND is_current = 1 AND lifecycle_state = 'active'
          AND deleted_at IS NULL
        """,
        (dataset_id,),
    ).fetchall()
    changed_at = now_iso()
    for row in rows:
        logical_id = str(row["logical_doc_id"] or "")
        row_source_root = str(row["source_root"] or "")
        # Legacy rows did not retain their source root.  Treat them
        # conservatively: absence from an arbitrary later directory scan is
        # not enough evidence to tombstone historical active data.
        if not row_source_root:
            continue
        if Path(row_source_root).resolve() != Path(source_root).resolve():
            continue
        if logical_id in seen_logical_doc_ids:
            continue
        conn.execute(
            """
            UPDATE documents
            SET lifecycle_state = 'removed', is_current = 0,
                deleted_at = ?, updated_at = ?
            WHERE doc_id = ?
            """,
            (changed_at, changed_at, row["doc_id"]),
        )
        removed.append(str(row["doc_id"]))
    return removed


def _reset_ingest_artifacts(raw_dir: Path, collection_db_path: Path) -> None:
    """Keep the legacy flag non-destructive.

    A full directory scan plus version/tombstone reconciliation below provides
    logical reset semantics.  Raw version files and the collection ledger must
    remain intact so historical citations and ingest jobs stay reproducible.
    """
    raw_dir.mkdir(parents=True, exist_ok=True)
    collection_db_path.parent.mkdir(parents=True, exist_ok=True)


def _sync_index_registry(conn: sqlite3.Connection, dataset_id: str, source_doc_ids: list[str], chunk_count: int) -> None:
    now = now_iso()
    index_status = "ready" if source_doc_ids and chunk_count > 0 else "empty"
    rows = [
        {
            "index_id": sha256_text(f"{dataset_id}\0sqlite_structured")[:40],
            "dataset_id": dataset_id,
            "index_type": "sqlite_structured",
            "collection_name": dataset_id,
            "index_path": "meta/collection.sqlite3",
            "source_doc_ids_json": dumps_json(source_doc_ids),
            "source_chunk_count": chunk_count,
            "status": index_status,
            "built_at": now if index_status == "ready" else None,
            "error_message": None,
            "metadata_json": dumps_json(
                {
                    "tables": [
                        "documents",
                        "chunks",
                        "chunk_locations",
                        "excel_cells",
                        "metric_facts",
                    ],
                    "active_source_doc_count": len(source_doc_ids),
                    "active_chunk_count": chunk_count,
                }
            ),
        },
        {
            "index_id": sha256_text(f"{dataset_id}\0summary_chunks")[:40],
            "dataset_id": dataset_id,
            "index_type": "summary_chunks",
            "collection_name": dataset_id,
            "index_path": "meta/collection.sqlite3:chunks",
            "source_doc_ids_json": dumps_json(source_doc_ids),
            "source_chunk_count": chunk_count,
            "status": index_status,
            "built_at": now if index_status == "ready" else None,
            "error_message": None,
            "metadata_json": dumps_json(
                {
                    "note": (
                        "Chunks are ready for a later Chroma/BM25 indexing step."
                        if index_status == "ready"
                        else "No active searchable chunks are available."
                    ),
                    "active_source_doc_count": len(source_doc_ids),
                    "active_chunk_count": chunk_count,
                }
            ),
        },
    ]
    conn.executemany(
        """
        INSERT OR REPLACE INTO index_registry (
            index_id, dataset_id, index_type, collection_name, index_path,
            source_doc_ids_json, source_chunk_count, status, built_at,
            error_message, metadata_json
        ) VALUES (
            :index_id, :dataset_id, :index_type, :collection_name, :index_path,
            :source_doc_ids_json, :source_chunk_count, :status, :built_at,
            :error_message, :metadata_json
        )
        """,
        rows,
    )


def ingest_directory(
    *,
    directory_path: str | Path,
    workspace_root: str | Path | None = None,
    dataset_id: Optional[str] = None,
    dataset_name: Optional[str] = None,
    company_name: str = "",
    company_ticker: str = "",
    recursive: bool = True,
    reset: bool = False,
    job_id: Optional[str] = None,
    classification_llm: ClassificationChatClient | None = None,
) -> IngestResult:
    source_dir = Path(directory_path).expanduser().resolve()
    if not source_dir.is_dir():
        raise FileNotFoundError(f"directory_path is not a directory: {source_dir}")

    workspace = Path(workspace_root).expanduser().resolve() if workspace_root else default_workspace_root()
    dataset_id = safe_slug(dataset_id or source_dir.name)
    dataset_name = dataset_name or source_dir.name
    dataset_root = workspace / dataset_id
    raw_dir = dataset_root / "raw"
    meta_dir = dataset_root / "meta"
    collection_db_path = meta_dir / "collection.sqlite3"
    global_db_path = workspace / "datasets.sqlite3"
    started = now_iso()
    job_id = job_id or sha256_text(f"{dataset_id}\0{source_dir}\0{started}")[:16]

    _validate_source_output_layout(source_dir, workspace, dataset_root)

    if reset:
        _reset_ingest_artifacts(raw_dir, collection_db_path)

    workspace.mkdir(parents=True, exist_ok=True)
    raw_dir.mkdir(parents=True, exist_ok=True)
    meta_dir.mkdir(parents=True, exist_ok=True)

    excluded_roots: list[Path] = []
    if dataset_root != source_dir and _path_is_within(dataset_root, source_dir):
        excluded_roots.append(dataset_root)
    if workspace != source_dir and _path_is_within(workspace, source_dir):
        excluded_roots.append(workspace)
    files, unsupported_files = _iter_input_files(
        source_dir,
        recursive=recursive,
        excluded_roots=excluded_roots,
    )
    discovered_file_count = len(files) + len(unsupported_files)
    result = IngestResult(
        job_id=job_id,
        dataset_id=dataset_id,
        dataset_name=dataset_name,
        source_dir=str(source_dir),
        workspace_root=str(workspace),
        dataset_root=str(dataset_root),
        collection_db_path=str(collection_db_path),
        global_db_path=str(global_db_path),
        status="running",
        file_count=len(files),
        discovered_file_count=discovered_file_count,
        supported_file_count=len(files),
        unsupported_file_count=len(unsupported_files),
        started_at=started,
    )
    for source_path in unsupported_files:
        relpath = _normalized_source_relpath(str(source_path.relative_to(source_dir)))
        suffix = source_path.suffix.lower() or "(none)"
        result.documents.append(
            DocumentIngestResult(
                doc_id=sha256_text(f"{dataset_id}\0unsupported\0{relpath}")[:40],
                filename=relpath,
                file_type=source_path.suffix.lower().lstrip(".") or "unknown",
                status="unsupported",
                error_message=(
                    f"Unsupported file type {suffix}. Supported extensions: "
                    f"{', '.join(sorted(SUPPORTED_EXTENSIONS))}"
                ),
                logical_doc_id=_logical_doc_id(dataset_id, relpath),
                lifecycle_state="unsupported",
            )
        )

    with connect_sqlite(global_db_path) as global_conn:
        ensure_global_schema(global_conn)
        global_conn.execute(
            """
            INSERT OR REPLACE INTO datasets (
                dataset_id, name, status, source_dir, dataset_root, company_name,
                company_ticker, file_count, created_at, updated_at, metadata_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, COALESCE((SELECT created_at FROM datasets WHERE dataset_id = ?), ?), ?, ?)
            """,
            (
                dataset_id,
                dataset_name,
                "indexing",
                str(source_dir),
                str(dataset_root),
                company_name,
                company_ticker,
                len(files),
                dataset_id,
                started,
                started,
                dumps_json(
                    {
                        "pipeline": "private_fund_directory_ingest",
                        "discovered_file_count": discovered_file_count,
                        "supported_file_count": len(files),
                        "unsupported_file_count": len(unsupported_files),
                    }
                ),
            ),
        )
        global_conn.execute(
            "INSERT OR REPLACE INTO dataset_state (id, active_dataset_id, updated_at) VALUES (1, ?, ?)",
            (dataset_id, now_iso()),
        )
        global_conn.commit()

    doc_ids: list[str] = []
    seen_logical_doc_ids: set[str] = set()
    try:
        with connect_sqlite(collection_db_path) as conn:
            ensure_collection_schema(conn)
            conn.execute(
                """
                INSERT OR REPLACE INTO ingest_jobs (
                    job_id, dataset_id, job_type, status, doc_ids_json, file_count,
                    log_path, message, returncode, created_at, started_at, finished_at,
                    metadata_json
                ) VALUES (?, ?, ?, ?, ?, ?, NULL, ?, NULL, ?, ?, NULL, ?)
                """,
                (
                    job_id,
                    dataset_id,
                    "private_fund_directory",
                    "running",
                    dumps_json([]),
                    len(files),
                    "Directory ingestion started.",
                    started,
                    started,
                    dumps_json(
                        {
                            "source_dir": str(source_dir),
                            "recursive": recursive,
                            "reset_requested": reset,
                            "reset_mode": "non_destructive_full_reconciliation",
                            "discovered_file_count": discovered_file_count,
                            "supported_file_count": len(files),
                            "unsupported_file_count": len(unsupported_files),
                        }
                    ),
                ),
            )
            conn.commit()

            for source_path in files:
                source_relpath = _normalized_source_relpath(
                    str(source_path.relative_to(source_dir))
                )
                logical_id = _logical_doc_id(dataset_id, source_relpath)
                seen_logical_doc_ids.add(logical_id)
                current = _current_document(conn, dataset_id, logical_id)
                latest = _latest_document_version(conn, dataset_id, logical_id)
                try:
                    checksum = sha256_file(source_path)
                except Exception as exc:
                    result.documents.append(
                        DocumentIngestResult(
                            doc_id=sha256_text(f"{dataset_id}\0{logical_id}\0read_failed")[:40],
                            filename=source_relpath,
                            file_type=source_path.suffix.lower().lstrip("."),
                            status="failed",
                            error_message=f"Unable to read source file: {exc}",
                            logical_doc_id=logical_id,
                            lifecycle_state="failed_attempt",
                        )
                    )
                    continue

                current_stored_file_is_valid = False
                if current is not None:
                    current_stored_path = Path(str(current["stored_path"]))
                    try:
                        current_stored_file_is_valid = (
                            current_stored_path.is_file()
                            and sha256_file(current_stored_path) == str(current["checksum"])
                        )
                    except OSError:
                        current_stored_file_is_valid = False
                if (
                    not reset
                    and current is not None
                    and str(current["checksum"]) == checksum
                    and str(current["status"]) in {"indexed", "needs_ocr"}
                    and current_stored_file_is_valid
                ):
                    if str(current["classifier_version"] or "") != CLASSIFIER_VERSION:
                        try:
                            refreshed_classification = classify_document(
                                build_document_preview(current_stored_path),
                                expected_company=company_name,
                                expected_ticker=company_ticker,
                                llm_client=classification_llm,
                            )
                        except Exception as classification_exc:  # noqa: BLE001
                            refreshed_classification = _fallback_classification(
                                company_name=company_name,
                                company_ticker=company_ticker,
                                error=str(classification_exc),
                            )
                        _update_document_classification(
                            conn,
                            str(current["doc_id"]),
                            refreshed_classification,
                        )
                    conn.execute(
                        """
                        UPDATE documents
                        SET source_root = ?, source_relpath = ?, updated_at = ?
                        WHERE doc_id = ?
                        """,
                        (str(source_dir), source_relpath, now_iso(), current["doc_id"]),
                    )
                    current = _current_document(conn, dataset_id, logical_id) or current
                    reused_result = _document_result_from_row(conn, current)
                    reused_result.filename = source_path.name
                    result.documents.append(reused_result)
                    doc_ids.append(str(current["doc_id"]))
                    conn.commit()
                    continue

                stored_path: Optional[Path] = None
                version_no = _next_document_version(conn, dataset_id, logical_id)
                supersedes_doc_id = str(latest["doc_id"]) if latest is not None else None
                doc_id = ""
                savepoint_open = False
                classification = _fallback_classification(
                    company_name=company_name,
                    company_ticker=company_ticker,
                )
                try:
                    stored_path = _copy_to_raw(source_path, raw_dir)
                    checksum = sha256_file(stored_path)
                    try:
                        preview = build_document_preview(stored_path)
                        classification = classify_document(
                            preview,
                            expected_company=company_name,
                            expected_ticker=company_ticker,
                            llm_client=classification_llm,
                        )
                    except Exception as classification_exc:  # noqa: BLE001
                        classification = _fallback_classification(
                            company_name=company_name,
                            company_ticker=company_ticker,
                            error=str(classification_exc),
                        )
                    conn.execute("SAVEPOINT ingest_document")
                    savepoint_open = True
                    doc_id = _register_document(
                        conn,
                        dataset_id=dataset_id,
                        stored_path=stored_path,
                        original_filename=source_path.name,
                        checksum=checksum,
                        classification=classification,
                        source_root=str(source_dir),
                        source_relpath=source_relpath,
                        logical_doc_id=logical_id,
                        version_no=version_no,
                        supersedes_doc_id=supersedes_doc_id,
                    )
                    suffix = stored_path.suffix.lower()
                    if classification.classification_status == "company_conflict":
                        doc_result = DocumentIngestResult(
                            doc_id=doc_id,
                            filename=stored_path.name,
                            file_type=suffix.lstrip("."),
                            status="classification_review_required",
                            error_message=(
                                "Company classification conflicts with the active project; "
                                "the document was preserved but not indexed. "
                                + "; ".join(classification.evidence[-3:])
                            ),
                        )
                    elif suffix == ".pdf":
                        doc_result = ingest_pdf(
                            conn, dataset_id=dataset_id, doc_id=doc_id, path=stored_path
                        )
                    elif suffix in {".xlsx", ".xlsm"}:
                        doc_result = ingest_excel(
                            conn, dataset_id=dataset_id, doc_id=doc_id, path=stored_path
                        )
                    elif suffix in ADAPTER_EXTENSIONS:
                        doc_result = ingest_adapted_document(
                            conn, dataset_id=dataset_id, doc_id=doc_id, path=stored_path
                        )
                    else:
                        raise ValueError(f"Unsupported file type: {stored_path.suffix}")
                    doc_result.logical_doc_id = logical_id
                    doc_result.version_no = version_no
                    doc_result.supersedes_doc_id = supersedes_doc_id
                    _apply_classification_result(doc_result, classification)
                    _update_document_status(
                        conn,
                        doc_id,
                        doc_result.status,
                        doc_result.error_message,
                        parser_name=doc_result.parser_name,
                        parser_version=doc_result.parser_version,
                        parser_metadata=doc_result.parser_metadata,
                    )
                    _activate_document_version(
                        conn,
                        dataset_id=dataset_id,
                        logical_doc_id=logical_id,
                        doc_id=doc_id,
                    )
                    conn.execute("RELEASE SAVEPOINT ingest_document")
                    savepoint_open = False
                    result.documents.append(doc_result)
                    doc_ids.append(doc_id)
                except Exception as exc:
                    if savepoint_open:
                        conn.execute("ROLLBACK TO SAVEPOINT ingest_document")
                        conn.execute("RELEASE SAVEPOINT ingest_document")
                    error_message = str(exc)
                    failed_doc_id = doc_id or sha256_text(
                        f"{dataset_id}\0{logical_id}\0{version_no}\0{checksum}"
                    )[:40]
                    # Persist the failed attempt itself, but not any partially
                    # written pages/cells/chunks from the rolled-back parser.
                    if stored_path is not None:
                        try:
                            conn.execute("SAVEPOINT record_failed_document")
                            failed_doc_id = _register_document(
                                conn,
                                dataset_id=dataset_id,
                                stored_path=stored_path,
                                original_filename=source_path.name,
                                checksum=checksum,
                                classification=classification,
                                source_root=str(source_dir),
                                source_relpath=source_relpath,
                                logical_doc_id=logical_id,
                                version_no=version_no,
                                supersedes_doc_id=supersedes_doc_id,
                            )
                            _update_document_status(
                                conn, failed_doc_id, "failed", error_message
                            )
                            if current is None:
                                _activate_document_version(
                                    conn,
                                    dataset_id=dataset_id,
                                    logical_doc_id=logical_id,
                                    doc_id=failed_doc_id,
                                )
                            else:
                                failed_at = now_iso()
                                conn.execute(
                                    """
                                    UPDATE documents
                                    SET lifecycle_state = 'failed_attempt',
                                        is_current = 0, deleted_at = ?, updated_at = ?
                                    WHERE doc_id = ?
                                    """,
                                    (failed_at, failed_at, failed_doc_id),
                                )
                            conn.execute("RELEASE SAVEPOINT record_failed_document")
                            doc_ids.append(failed_doc_id)
                        except Exception:
                            conn.execute("ROLLBACK TO SAVEPOINT record_failed_document")
                            conn.execute("RELEASE SAVEPOINT record_failed_document")
                    result.documents.append(
                        DocumentIngestResult(
                            doc_id=failed_doc_id,
                            filename=source_path.name,
                            file_type=source_path.suffix.lower().lstrip("."),
                            status="failed",
                            error_message=error_message,
                            logical_doc_id=logical_id,
                            version_no=version_no,
                            supersedes_doc_id=supersedes_doc_id,
                            lifecycle_state=(
                                "failed_attempt" if current is not None else "active"
                            ),
                            doc_type=classification.doc_type,
                            doc_subtype=classification.doc_subtype,
                            doc_type_confidence=classification.confidence,
                            classification_status=classification.classification_status,
                            classification_method=classification.method,
                            company_name=classification.company_name,
                            company_ticker=classification.company_ticker,
                            company_confidence=classification.company_confidence,
                        )
                    )
                conn.commit()

            removed_doc_ids = _mark_removed_documents(
                conn,
                dataset_id=dataset_id,
                source_root=str(source_dir),
                seen_logical_doc_ids=seen_logical_doc_ids,
            )
            result.removed_file_count = len(removed_doc_ids)
            active_doc_ids = [
                str(row[0])
                for row in conn.execute(
                    """
                    SELECT doc_id FROM documents
                    WHERE dataset_id = ? AND is_current = 1
                      AND lifecycle_state = 'active'
                      AND deleted_at IS NULL AND status = 'indexed'
                    ORDER BY source_relpath, version_no
                    """,
                    (dataset_id,),
                )
            ]
            total_chunks = conn.execute(
                """
                SELECT COUNT(*)
                FROM chunks c
                JOIN documents d ON d.doc_id = c.doc_id
                WHERE c.dataset_id = ? AND d.is_current = 1
                  AND d.lifecycle_state = 'active'
                  AND d.deleted_at IS NULL AND d.status = 'indexed'
                """,
                (dataset_id,),
            ).fetchone()[0]
            _sync_index_registry(conn, dataset_id, active_doc_ids, int(total_chunks or 0))
            failures = [doc for doc in result.documents if doc.status == "failed"]
            warnings = [
                doc
                for doc in result.documents
                if doc.status
                in {"unsupported", "needs_ocr", "classification_review_required"}
            ]
            result.warning_count = len(warnings)
            if not files or failures:
                result.status = "failed"
            elif warnings:
                result.status = "completed_with_warnings"
            else:
                result.status = "completed"
            result.finished_at = now_iso()
            result.message = (
                f"Processed {len(files)} supported files: {len(failures)} failed, "
                f"{sum(doc.status == 'needs_ocr' for doc in warnings)} need OCR, "
                f"{sum(doc.status == 'classification_review_required' for doc in warnings)} need company review, "
                f"{len(unsupported_files)} unsupported, {len(removed_doc_ids)} removed."
            )
            returncode = 1 if result.status == "failed" else (2 if warnings else 0)
            conn.execute(
                """
                UPDATE ingest_jobs
                SET status = ?, doc_ids_json = ?, message = ?, returncode = ?,
                    finished_at = ?, metadata_json = ?
                WHERE job_id = ?
                """,
                (
                    result.status,
                    dumps_json(doc_ids),
                    result.message,
                    returncode,
                    result.finished_at,
                    dumps_json(asdict(result)),
                    job_id,
                ),
            )
            conn.commit()
    except Exception as exc:
        result.status = "failed"
        result.finished_at = now_iso()
        result.message = f"Directory ingestion failed before completion: {exc}"
        try:
            with connect_sqlite(collection_db_path) as failed_conn:
                failed_conn.execute(
                    """
                    UPDATE ingest_jobs
                    SET status = 'failed', message = ?, returncode = 1,
                        finished_at = ?, metadata_json = ?
                    WHERE job_id = ?
                    """,
                    (
                        result.message,
                        result.finished_at,
                        dumps_json(asdict(result)),
                        job_id,
                    ),
                )
                failed_conn.commit()
        except Exception:
            # Preserve the original ingestion exception; job finalization is
            # best-effort when the collection database itself is unavailable.
            pass
        raise
    finally:
        with connect_sqlite(global_db_path) as global_conn:
            ensure_global_schema(global_conn)
            global_conn.execute(
                "UPDATE datasets SET status = ?, updated_at = ?, file_count = ?, metadata_json = ? WHERE dataset_id = ?",
                (result.status, now_iso(), len(files), dumps_json(asdict(result)), dataset_id),
            )
            global_conn.commit()

    return result


def result_to_dict(result: IngestResult) -> dict[str, Any]:
    return asdict(result)


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest a private-fund research directory into structured SQLite.")
    parser.add_argument("--directory", required=True, help="Directory containing PDF/Excel documents.")
    parser.add_argument("--workspace-root", default="", help="Output workspace root. Defaults to ./output/private_fund_datasets.")
    parser.add_argument("--dataset-id", default="", help="Dataset id. Defaults to directory name.")
    parser.add_argument("--dataset-name", default="", help="Human readable dataset name.")
    parser.add_argument("--company-name", default="", help="Company name metadata.")
    parser.add_argument("--company-ticker", default="", help="Ticker metadata.")
    parser.add_argument("--no-recursive", action="store_true", help="Only scan direct children.")
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Run a full non-destructive reconciliation; history and prior report artifacts are preserved.",
    )
    args = parser.parse_args()

    result = ingest_directory(
        directory_path=args.directory,
        workspace_root=args.workspace_root or None,
        dataset_id=args.dataset_id or None,
        dataset_name=args.dataset_name or None,
        company_name=args.company_name,
        company_ticker=args.company_ticker,
        recursive=not args.no_recursive,
        reset=args.reset,
    )
    print(dumps_json(result_to_dict(result)))
    if result.status == "failed":
        raise SystemExit(1)
    if result.status == "completed_with_warnings" or result.warning_count:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
