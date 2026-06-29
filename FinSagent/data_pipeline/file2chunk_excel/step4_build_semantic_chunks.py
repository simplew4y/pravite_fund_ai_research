#!/usr/bin/env python3
"""Step 4: merge source regions into RAG-ready semantic Excel chunks."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from common import (
    bounds_to_range,
    cell_display,
    dump_json,
    ensure_supported_excel,
    is_formula,
    load_json,
    normalize_text,
    sha256_text,
    union_bounds,
)


MAJOR_TYPES = {
    "table",
    "formula_block",
    "assumptions",
    "income_statement",
    "balance_sheet",
    "cash_flow",
    "dcf",
    "comps",
    "sensitivity",
}

MODEL_TYPES = {
    "assumptions",
    "income_statement",
    "balance_sheet",
    "cash_flow",
    "dcf",
    "comps",
    "sensitivity",
    "formula_block",
}


def _load_value_workbook(excel_path: Path):
    try:
        return load_workbook(excel_path, data_only=True, read_only=False, keep_links=False)
    except Exception:
        return None


def _bounds_gap(a: dict[str, int], b: dict[str, int]) -> tuple[int, int]:
    if a["max_row"] < b["min_row"]:
        row_gap = b["min_row"] - a["max_row"] - 1
    elif b["max_row"] < a["min_row"]:
        row_gap = a["min_row"] - b["max_row"] - 1
    else:
        row_gap = 0
    if a["max_col"] < b["min_col"]:
        col_gap = b["min_col"] - a["max_col"] - 1
    elif b["max_col"] < a["min_col"]:
        col_gap = a["min_col"] - b["max_col"] - 1
    else:
        col_gap = 0
    return row_gap, col_gap


def _is_major(region: dict[str, Any]) -> bool:
    region_type = region.get("region_type")
    non_empty = int(region.get("non_empty_cell_count") or 0)
    row_count = int(region.get("row_count") or 0)
    col_count = int(region.get("col_count") or 0)
    if region_type in MAJOR_TYPES:
        return True
    return non_empty >= 10 or (row_count >= 3 and col_count >= 2 and non_empty >= 6)


def _region_text(region: dict[str, Any]) -> str:
    return normalize_text(" ".join(str(item) for item in region.get("row_texts", [])[:12]))


def _region_title(region: dict[str, Any]) -> str:
    for line in region.get("row_texts", [])[:8]:
        text = normalize_text(str(line))
        if text and len(text) >= 4:
            return text[:120]
    preview = " ".join(str(item) for item in region.get("value_preview", [])[:4])
    return normalize_text(preview)[:120] or f"{region.get('sheet_name')} {region.get('cell_range')}"


def _source_location(region: dict[str, Any]) -> dict[str, Any]:
    return {
        "source_region_id": region.get("region_id"),
        "sheet_name": region.get("sheet_name"),
        "cell_range": region.get("cell_range"),
        "bounds": region.get("bounds"),
        "region_type": region.get("region_type"),
        "non_empty_cell_count": region.get("non_empty_cell_count"),
        "formula_count": region.get("formula_count"),
    }


def _attach_small_regions(regions: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    sorted_regions = sorted(regions, key=lambda r: (r["bounds"]["min_row"], r["bounds"]["min_col"]))
    major_groups = [[region] for region in sorted_regions if _is_major(region)]
    small_regions = [region for region in sorted_regions if not _is_major(region)]
    if not major_groups:
        return _group_text_regions(small_regions)

    for small in small_regions:
        best_index = None
        best_score = 10**9
        for idx, group in enumerate(major_groups):
            group_bounds = union_bounds(region["bounds"] for region in group)
            row_gap, col_gap = _bounds_gap(small["bounds"], group_bounds)
            above_or_overlap = small["bounds"]["max_row"] <= group_bounds["max_row"] + 4
            near = row_gap <= 4 or (row_gap <= 10 and col_gap == 0)
            if not (above_or_overlap and near):
                continue
            score = row_gap * 20 + col_gap
            if score < best_score:
                best_score = score
                best_index = idx
        if best_index is not None:
            major_groups[best_index].append(small)
        elif len(_region_text(small)) >= 30:
            major_groups.append([small])

    merged = [_merge_overlapping_groups(group) for group in major_groups]
    merged.sort(key=lambda group: (min(r["bounds"]["min_row"] for r in group), min(r["bounds"]["min_col"] for r in group)))
    return merged


def _group_text_regions(regions: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    groups: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    current_bounds: dict[str, int] | None = None
    for region in sorted(regions, key=lambda r: (r["bounds"]["min_row"], r["bounds"]["min_col"])):
        if not current:
            current = [region]
            current_bounds = region["bounds"]
            continue
        row_gap, _ = _bounds_gap(current_bounds or region["bounds"], region["bounds"])
        if row_gap <= 4:
            current.append(region)
            current_bounds = union_bounds(item["bounds"] for item in current)
        else:
            groups.append(current)
            current = [region]
            current_bounds = region["bounds"]
    if current:
        groups.append(current)
    return groups


def _merge_overlapping_groups(group: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # Within one semantic group, keep source regions ordered and unique.
    seen = set()
    ordered: list[dict[str, Any]] = []
    for region in sorted(group, key=lambda r: (r["bounds"]["min_row"], r["bounds"]["min_col"])):
        key = (region.get("sheet_name"), region.get("cell_range"), region.get("region_id"))
        if key in seen:
            continue
        seen.add(key)
        ordered.append(region)
    return ordered


def _group_regions_by_sheet(regions: list[dict[str, Any]], workbook_type: str) -> list[list[dict[str, Any]]]:
    groups: list[list[dict[str, Any]]] = []
    by_sheet: dict[str, list[dict[str, Any]]] = {}
    for region in regions:
        by_sheet.setdefault(str(region.get("sheet_name") or ""), []).append(region)

    for _, sheet_regions in by_sheet.items():
        sheet_regions = sorted(sheet_regions, key=lambda r: (r["bounds"]["min_row"], r["bounds"]["min_col"]))
        if workbook_type == "simple_table":
            groups.append(sheet_regions)
        else:
            groups.extend(_attach_small_regions(sheet_regions))
    return [group for group in groups if _group_has_meaningful_text(group)]


def _group_has_meaningful_text(group: list[dict[str, Any]]) -> bool:
    text = normalize_text(" ".join(_region_text(region) for region in group))
    if len(text) < 8:
        return False
    if len(group) == 1 and int(group[0].get("non_empty_cell_count") or 0) <= 2 and len(text) < 30:
        return False
    return True


def _matrix_for_bounds(ws_formula, ws_values, bounds: dict[str, int], *, max_rows: int, max_cols: int) -> tuple[list[list[str]], list[dict[str, Any]], bool]:
    row_stop = min(bounds["max_row"], bounds["min_row"] + max_rows - 1)
    col_stop = min(bounds["max_col"], bounds["min_col"] + max_cols - 1)
    truncated = row_stop < bounds["max_row"] or col_stop < bounds["max_col"]
    matrix: list[list[str]] = []
    formula_samples: list[dict[str, Any]] = []

    header = ["row"]
    header.extend(get_column_letter(col) for col in range(bounds["min_col"], col_stop + 1))
    matrix.append(header)
    for row_idx in range(bounds["min_row"], row_stop + 1):
        row_values = [str(row_idx)]
        has_value = False
        for col_idx in range(bounds["min_col"], col_stop + 1):
            f_cell = ws_formula.cell(row_idx, col_idx)
            v_cell = ws_values.cell(row_idx, col_idx) if ws_values is not None else None
            value = f_cell.value
            cached = v_cell.value if v_cell is not None else None
            if is_formula(value):
                row_values.append(cell_display(cached))
                has_value = has_value or cached is not None
                if len(formula_samples) < 40:
                    formula_samples.append(
                        {
                            "cell": f_cell.coordinate,
                            "formula": value,
                            "cached_value": cached,
                            "formula_value_missing": cached is None,
                        }
                    )
            else:
                display = cell_display(value)
                row_values.append(display)
                has_value = has_value or bool(display)
        if has_value:
            matrix.append(row_values)
    return matrix, formula_samples, truncated


def _markdown_table(rows: list[list[str]], *, max_cell_len: int = 80) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized: list[list[str]] = []
    for row in rows:
        values = []
        for idx in range(width):
            value = row[idx] if idx < len(row) else ""
            value = value.replace("|", "\\|")
            if len(value) > max_cell_len:
                value = value[: max_cell_len - 1] + "..."
            values.append(value)
        normalized.append(values)
    sep = ["---"] * width
    lines = ["| " + " | ".join(normalized[0]) + " |", "| " + " | ".join(sep) + " |"]
    lines.extend("| " + " | ".join(row) + " |" for row in normalized[1:])
    return "\n".join(lines)


def _formula_block(formulas: list[dict[str, Any]]) -> str:
    if not formulas:
        return "None"
    lines = []
    for item in formulas[:20]:
        cached = "" if item.get("cached_value") is None else f", cached={item.get('cached_value')}"
        missing = " [cached value missing]" if item.get("formula_value_missing") else ""
        lines.append(f"- {item.get('cell')}: {item.get('formula')}{cached}{missing}")
    return "\n".join(lines)


def _content_type_for_group(group: list[dict[str, Any]], workbook_type: str) -> str:
    types = {str(region.get("region_type") or "") for region in group}
    if workbook_type == "valuation_model" and types & MODEL_TYPES:
        return "excel_model_section"
    if "table" in types or any(_is_major(region) for region in group):
        return "excel_table"
    if workbook_type != "financial_supplement" and types & MODEL_TYPES:
        return "excel_model_section"
    return "excel_section"


def _section_title(group: list[dict[str, Any]], workbook_name: str) -> str:
    major = [region for region in group if _is_major(region)]
    candidates = major or group
    title = _region_title(candidates[0])
    if len(title) > 140:
        title = title[:137] + "..."
    return title or workbook_name


def _chunk_from_group(
    *,
    excel_path: Path,
    wb_formula,
    wb_values,
    group: list[dict[str, Any]],
    workbook_type: str,
    doc_type: str,
    max_preview_rows: int,
    max_preview_cols: int,
) -> dict[str, Any]:
    sheet_name = str(group[0].get("sheet_name") or "")
    ws_formula = wb_formula[sheet_name]
    ws_values = wb_values[sheet_name] if wb_values is not None and sheet_name in wb_values.sheetnames else None
    bounds = union_bounds(region["bounds"] for region in group)
    matrix, formulas, truncated = _matrix_for_bounds(
        ws_formula,
        ws_values,
        bounds,
        max_rows=max_preview_rows,
        max_cols=max_preview_cols,
    )
    table = _markdown_table(matrix)
    locations = [_source_location(region) for region in group]
    ranges = [f"{loc['sheet_name']}!{loc['cell_range']}" for loc in locations]
    region_types = sorted({str(region.get("region_type") or "unknown") for region in group})
    title = _section_title(group, excel_path.stem)
    content_type = _content_type_for_group(group, workbook_type)
    context_lines = []
    for region in group:
        if not _is_major(region):
            text = _region_text(region)
            if text:
                context_lines.append(text)
    context = "\n".join(dict.fromkeys(context_lines[:8]))
    content_parts = [
        f"Workbook: {excel_path.stem}",
        f"Workbook type: {workbook_type}",
        f"Document type: {doc_type}",
        f"Sheet: {sheet_name}",
        f"Section: {title}",
        f"Source ranges: {', '.join(ranges)}",
        f"Region types: {', '.join(region_types)}",
    ]
    if context:
        content_parts.extend(["", f"Context:\n{context}"])
    content_parts.extend(["", f"Table preview:\n{table}"])
    if truncated:
        content_parts.append("Note: table preview is truncated; source ranges preserve full Excel location.")
    content_parts.extend(["", f"Formula samples:\n{_formula_block(formulas)}"])
    content = "\n".join(content_parts).strip()
    metadata = {
        "source": "file2chunk_excel",
        "workbook_type": workbook_type,
        "doc_type": doc_type,
        "sheet_name": sheet_name,
        "union_cell_range": bounds_to_range(bounds),
        "region_types": region_types,
        "source_regions": locations,
        "formula_samples": formulas,
        "preview_truncated": truncated,
    }
    return {
        "content": content,
        "content_type": content_type,
        "type": content_type,
        "title": title,
        "title_path": [excel_path.stem, sheet_name, title],
        "summary": "",
        "sheet_name": sheet_name,
        "cell_range": bounds_to_range(bounds),
        "source_locations": locations,
        "source_ref": "; ".join(ranges),
        "metadata": metadata,
        "content_hash": sha256_text(content),
    }


def _split_chunk(chunk: dict[str, Any], max_chunk_chars: int) -> list[dict[str, Any]]:
    content = str(chunk.get("content") or "")
    if len(content) <= max_chunk_chars:
        return [chunk]
    lines = content.splitlines()
    header: list[str] = []
    body_start = 0
    for idx, line in enumerate(lines):
        header.append(line)
        if line.startswith("Table preview:"):
            body_start = idx + 1
            break
    if body_start == 0:
        body_start = min(len(lines), 8)
        header = lines[:body_start]
    body = lines[body_start:]
    parts: list[dict[str, Any]] = []
    current: list[str] = []
    for line in body:
        candidate = "\n".join([*header, f"Part: {len(parts) + 1}", *current, line])
        if current and len(candidate) > max_chunk_chars:
            parts.append(_chunk_part(chunk, header, current, len(parts) + 1))
            current = [line]
        else:
            current.append(line)
    if current:
        parts.append(_chunk_part(chunk, header, current, len(parts) + 1))
    return parts


def _chunk_part(chunk: dict[str, Any], header: list[str], body: list[str], part_index: int) -> dict[str, Any]:
    item = dict(chunk)
    item["title"] = f"{chunk.get('title')} part {part_index}"
    item["title_path"] = [*chunk.get("title_path", [])[:-1], item["title"]]
    item["content"] = "\n".join([*header, f"Part: {part_index}", *body]).strip()
    item["content_hash"] = sha256_text(item["content"])
    metadata = dict(chunk.get("metadata") or {})
    metadata["part_index"] = part_index
    metadata["split_from_large_chunk"] = True
    item["metadata"] = metadata
    return item


def build_semantic_chunks(
    excel_path: str | Path,
    regions_path: str | Path,
    classification_path: str | Path,
    *,
    doc_type: str,
    max_preview_rows: int = 80,
    max_preview_cols: int = 24,
    max_chunk_chars: int = 6000,
) -> list[dict[str, Any]]:
    excel_path = ensure_supported_excel(excel_path)
    regions_doc = load_json(regions_path)
    classification = load_json(classification_path)
    workbook_type = classification.get("workbook_type") or regions_doc.get("workbook_type") or "unknown"
    regions = list(regions_doc.get("regions", []))
    wb_formula = load_workbook(excel_path, data_only=False, read_only=False, keep_links=False)
    wb_values = _load_value_workbook(excel_path)

    semantic_groups = _group_regions_by_sheet(regions, workbook_type)
    chunks: list[dict[str, Any]] = []
    for group in semantic_groups:
        chunk = _chunk_from_group(
            excel_path=excel_path,
            wb_formula=wb_formula,
            wb_values=wb_values,
            group=group,
            workbook_type=workbook_type,
            doc_type=doc_type,
            max_preview_rows=max_preview_rows,
            max_preview_cols=max_preview_cols,
        )
        chunks.extend(_split_chunk(chunk, max_chunk_chars))

    for idx, chunk in enumerate(chunks, start=1):
        chunk["id"] = idx
        chunk["chunk_index"] = idx
    return chunks


def main() -> None:
    parser = argparse.ArgumentParser(description="Build semantic Excel chunks.")
    parser.add_argument("--excel", required=True, help="Path to .xlsx/.xlsm file")
    parser.add_argument("--regions", required=True, help="Path to regions.json")
    parser.add_argument("--classification", required=True, help="Path to classification.json")
    parser.add_argument("--output", required=True, help="Path to base_final.json")
    parser.add_argument("--doc-type", default="valuation_model")
    parser.add_argument("--max-preview-rows", type=int, default=80)
    parser.add_argument("--max-preview-cols", type=int, default=24)
    parser.add_argument("--max-chunk-chars", type=int, default=6000)
    args = parser.parse_args()
    chunks = build_semantic_chunks(
        args.excel,
        args.regions,
        args.classification,
        doc_type=args.doc_type,
        max_preview_rows=args.max_preview_rows,
        max_preview_cols=args.max_preview_cols,
        max_chunk_chars=args.max_chunk_chars,
    )
    dump_json(chunks, args.output)


if __name__ == "__main__":
    main()
