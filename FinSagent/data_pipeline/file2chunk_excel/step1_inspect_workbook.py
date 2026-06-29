#!/usr/bin/env python3
"""Step 1: inspect workbook structure without producing RAG chunks."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from common import cell_display, dump_json, ensure_supported_excel, is_formula, sha256_file


def _defined_names(wb) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    try:
        iterable = wb.defined_names.values()
    except AttributeError:
        iterable = []
    for item in iterable:
        try:
            destinations = [f"{sheet}!{coord}" for sheet, coord in item.destinations]
        except Exception:
            destinations = []
        items.append(
            {
                "name": str(getattr(item, "name", "")),
                "attr_text": str(getattr(item, "attr_text", "") or ""),
                "destinations": destinations,
            }
        )
    return items


def _used_bounds(ws) -> dict[str, int] | None:
    min_row = min_col = 10**9
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            min_row = min(min_row, cell.row)
            min_col = min(min_col, cell.column)
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
    if max_row == 0:
        return None
    return {"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col}


def _sample_rows(ws, bounds: dict[str, int] | None, *, max_rows: int = 12, max_cols: int = 16) -> list[list[str]]:
    if bounds is None:
        return []
    row_stop = min(bounds["max_row"], bounds["min_row"] + max_rows - 1)
    col_stop = min(bounds["max_col"], bounds["min_col"] + max_cols - 1)
    rows: list[list[str]] = []
    header = ["row"]
    header.extend(get_column_letter(col) for col in range(bounds["min_col"], col_stop + 1))
    rows.append(header)
    for row_idx in range(bounds["min_row"], row_stop + 1):
        values = [str(row_idx)]
        for col_idx in range(bounds["min_col"], col_stop + 1):
            values.append(cell_display(ws.cell(row_idx, col_idx).value, max_len=80))
        rows.append(values)
    return rows


def _sheet_manifest(ws, index: int) -> dict[str, Any]:
    bounds = _used_bounds(ws)
    non_empty = 0
    formula_count = 0
    text_values: list[str] = []
    if bounds is not None:
        for row in ws.iter_rows(
            min_row=bounds["min_row"],
            max_row=bounds["max_row"],
            min_col=bounds["min_col"],
            max_col=bounds["max_col"],
        ):
            for cell in row:
                if cell.value is None:
                    continue
                non_empty += 1
                if is_formula(cell.value):
                    formula_count += 1
                if len(text_values) < 80:
                    value = cell_display(cell.value, max_len=100)
                    if value:
                        text_values.append(value)
    cell_range = None
    if bounds is not None:
        cell_range = (
            f"{get_column_letter(bounds['min_col'])}{bounds['min_row']}:"
            f"{get_column_letter(bounds['max_col'])}{bounds['max_row']}"
        )
    return {
        "sheet_index": index,
        "sheet_name": ws.title,
        "sheet_state": ws.sheet_state,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "used_bounds": bounds,
        "used_range": cell_range,
        "non_empty_cell_count": non_empty,
        "formula_count": formula_count,
        "formula_density": round(formula_count / max(1, non_empty), 4),
        "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
        "sample_rows": _sample_rows(ws, bounds),
        "sample_text": text_values,
    }


def inspect_workbook(excel_path: str | Path) -> dict[str, Any]:
    excel_path = ensure_supported_excel(excel_path)
    wb = load_workbook(excel_path, data_only=False, read_only=False, keep_links=False)
    sheets = [_sheet_manifest(ws, idx) for idx, ws in enumerate(wb.worksheets)]
    return {
        "source_path": str(excel_path),
        "source_filename": excel_path.name,
        "checksum": sha256_file(excel_path),
        "workbook": {
            "sheet_count": len(wb.worksheets),
            "sheet_names": [ws.title for ws in wb.worksheets],
            "has_macros": excel_path.suffix.lower() == ".xlsm",
            "defined_names": _defined_names(wb),
        },
        "sheets": sheets,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect an Excel workbook.")
    parser.add_argument("--excel", required=True, help="Path to .xlsx/.xlsm file")
    parser.add_argument("--output", required=True, help="Path to workbook_manifest.json")
    args = parser.parse_args()
    dump_json(inspect_workbook(args.excel), args.output)


if __name__ == "__main__":
    main()
