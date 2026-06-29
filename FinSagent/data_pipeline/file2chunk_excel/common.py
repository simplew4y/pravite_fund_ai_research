#!/usr/bin/env python3
"""Shared helpers for the standalone Excel semantic chunk pipeline."""

from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl.utils import get_column_letter, range_boundaries


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
DEFAULT_DOC_TYPE = "valuation_model"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def ensure_supported_excel(path: str | Path) -> Path:
    excel_path = Path(path).expanduser().resolve()
    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if excel_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Unsupported Excel type {excel_path.suffix!r}; supported: {supported}")
    return excel_path


def load_json(path: str | Path) -> Any:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def dump_json(data: Any, path: str | Path) -> None:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_stem(path: str | Path) -> str:
    stem = Path(path).stem.strip()
    stem = re.sub(r"[^0-9A-Za-z._\-\u4e00-\u9fff]+", "_", stem)
    return stem[:120] or "workbook"


def cell_display(value: Any, *, max_len: int = 120) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        text = f"{value:.6g}"
    else:
        text = str(value)
    text = " ".join(text.replace("\r", "\n").split())
    if len(text) > max_len:
        return text[: max_len - 1] + "..."
    return text


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def bounds_to_range(bounds: dict[str, int]) -> str:
    return (
        f"{get_column_letter(bounds['min_col'])}{bounds['min_row']}:"
        f"{get_column_letter(bounds['max_col'])}{bounds['max_row']}"
    )


def range_to_bounds(cell_range: str) -> dict[str, int]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return {
        "min_row": min_row,
        "min_col": min_col,
        "max_row": max_row,
        "max_col": max_col,
    }


def union_bounds(bounds_list: Iterable[dict[str, int]]) -> dict[str, int]:
    items = list(bounds_list)
    if not items:
        return {"min_row": 1, "min_col": 1, "max_row": 1, "max_col": 1}
    return {
        "min_row": min(item["min_row"] for item in items),
        "min_col": min(item["min_col"] for item in items),
        "max_row": max(item["max_row"] for item in items),
        "max_col": max(item["max_col"] for item in items),
    }


def relative_datasets_path(path: str | Path) -> str:
    resolved = Path(path).resolve()
    parts = resolved.parts
    if "datasets" in parts:
        idx = parts.index("datasets")
        return "/".join(parts[idx:])
    return str(resolved)


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def text_contains_any(text: str, words: Iterable[str]) -> bool:
    lowered = text.lower()
    return any(word.lower() in lowered for word in words)


def connect_collection(dataset_root: str | Path) -> sqlite3.Connection:
    db_path = Path(dataset_root).resolve() / "meta" / "collection.sqlite3"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path, timeout=30)
    conn.row_factory = sqlite3.Row
    ensure_collection_schema(conn)
    return conn


def ensure_collection_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS documents (
            doc_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            title TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            file_type TEXT NOT NULL,
            doc_type TEXT,
            source_type TEXT,
            source_name TEXT,
            company_name TEXT,
            company_ticker TEXT,
            document_date TEXT,
            checksum TEXT NOT NULL,
            file_size INTEGER NOT NULL,
            status TEXT NOT NULL,
            chunk_count INTEGER NOT NULL DEFAULT 0,
            error_message TEXT,
            metadata_json TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            deleted_at TEXT
        );

        CREATE TABLE IF NOT EXISTS chunks (
            chunk_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            chunk_index INTEGER NOT NULL,
            content TEXT NOT NULL,
            content_type TEXT NOT NULL,
            title_path TEXT,
            summary TEXT,
            token_count INTEGER,
            content_hash TEXT NOT NULL,
            prev_chunk_id TEXT,
            next_chunk_id TEXT,
            source_ref TEXT,
            metadata_json TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS chunk_locations (
            location_id TEXT PRIMARY KEY,
            chunk_id TEXT NOT NULL,
            doc_id TEXT NOT NULL,
            location_index INTEGER NOT NULL DEFAULT 0,
            page_start INTEGER,
            page_end INTEGER,
            page_numbers_json TEXT,
            slide_start INTEGER,
            slide_end INTEGER,
            sheet_name TEXT,
            cell_range TEXT,
            heading_path TEXT,
            bbox_json TEXT,
            source_refs_json TEXT,
            display_text TEXT NOT NULL,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS index_registry (
            index_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            index_type TEXT NOT NULL,
            collection_name TEXT,
            index_path TEXT NOT NULL,
            source_doc_ids_json TEXT,
            source_chunk_count INTEGER,
            status TEXT NOT NULL,
            built_at TEXT,
            error_message TEXT,
            metadata_json TEXT
        );

        CREATE TABLE IF NOT EXISTS ingest_jobs (
            job_id TEXT PRIMARY KEY,
            dataset_id TEXT NOT NULL,
            job_type TEXT NOT NULL,
            status TEXT NOT NULL,
            doc_ids_json TEXT,
            file_count INTEGER NOT NULL DEFAULT 0,
            log_path TEXT,
            message TEXT,
            returncode INTEGER,
            created_at TEXT NOT NULL,
            started_at TEXT,
            finished_at TEXT,
            metadata_json TEXT
        );
        """
    )
    conn.commit()


def update_document_failure(dataset_root: str | Path, doc_id: str, message: str) -> None:
    try:
        with connect_collection(dataset_root) as conn:
            conn.execute(
                """
                UPDATE documents
                SET status = 'failed', error_message = ?, updated_at = ?
                WHERE doc_id = ?
                """,
                (message[:4000], now_iso(), doc_id),
            )
            conn.commit()
    except sqlite3.Error:
        return
