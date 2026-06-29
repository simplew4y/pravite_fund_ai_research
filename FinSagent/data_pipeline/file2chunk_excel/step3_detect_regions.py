#!/usr/bin/env python3
"""Step 3: detect fine-grained Excel source regions for later semantic merging."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import bounds_to_range, cell_display, dump_json, ensure_supported_excel, is_formula, load_json, normalize_text


MODEL_KEYWORDS = {
    "assumptions": ("assumption", "input", "driver", "scenario", "case", "假设", "输入", "驱动"),
    "income_statement": ("income statement", "p&l", "profit and loss", "revenue", "ebitda", "net income", "收入", "利润表"),
    "balance_sheet": ("balance sheet", "assets", "liabilities", "equity", "资产负债表"),
    "cash_flow": ("cash flow", "cashflow", "operating cash", "free cash flow", "现金流"),
    "dcf": ("dcf", "wacc", "terminal value", "discount", "npv", "折现", "估值"),
    "comps": ("comps", "comparable", "peer", "multiple", "ev/ebitda", "peers", "可比"),
    "sensitivity": ("sensitivity", "敏感性", "敏感度"),
}

UNIT_WORDS = ("$", "usd", "millions", "thousands", "except per share", "人民币", "百万元", "千元")
NOTE_WORDS = ("note:", "notes:", "source:", "来源", "注：", "备注")


def _used_cells(ws) -> dict[tuple[int, int], Any]:
    cells: dict[tuple[int, int], Any] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells[(cell.row, cell.column)] = cell.value
    return cells


def _bands_with_tolerance(indices: set[int], *, max_gap: int) -> list[tuple[int, int]]:
    ordered = sorted(indices)
    if not ordered:
        return []
    bands: list[tuple[int, int]] = []
    start = prev = ordered[0]
    for item in ordered[1:]:
        if item - prev <= max_gap + 1:
            prev = item
            continue
        bands.append((start, prev))
        start = prev = item
    bands.append((start, prev))
    return bands


def _rectangles(cells: dict[tuple[int, int], Any], *, row_gap: int, col_gap: int) -> list[dict[str, int]]:
    row_bands = _bands_with_tolerance({row for row, _ in cells}, max_gap=row_gap)
    rects: list[dict[str, int]] = []
    for min_row, max_row in row_bands:
        cols = {col for row, col in cells if min_row <= row <= max_row}
        for min_col, max_col in _bands_with_tolerance(cols, max_gap=col_gap):
            if any(min_row <= row <= max_row and min_col <= col <= max_col for row, col in cells):
                rects.append({"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col})
    return rects


def _cell_values(ws, bounds: dict[str, int]) -> list[Any]:
    values: list[Any] = []
    for row in ws.iter_rows(
        min_row=bounds["min_row"],
        max_row=bounds["max_row"],
        min_col=bounds["min_col"],
        max_col=bounds["max_col"],
    ):
        for cell in row:
            if cell.value is not None:
                values.append(cell.value)
    return values


def _region_type(sheet_name: str, values: list[Any], bounds: dict[str, int], formula_count: int) -> str:
    text = normalize_text(" ".join(cell_display(v, max_len=120) for v in values if v is not None)).lower()
    row_count = bounds["max_row"] - bounds["min_row"] + 1
    col_count = bounds["max_col"] - bounds["min_col"] + 1
    non_empty = len(values)

    for region_type, words in MODEL_KEYWORDS.items():
        if any(word.lower() in text for word in words):
            return region_type
    if formula_count / max(1, non_empty) >= 0.25:
        return "formula_block"
    if any(word in text for word in UNIT_WORDS) and non_empty <= 8:
        return "unit_or_header"
    if any(word in text for word in NOTE_WORDS):
        return "note"
    if row_count >= 2 and col_count >= 2 and non_empty >= 6:
        return "table"
    if non_empty <= 6:
        return "text_block"
    return "table"


def _preview(ws, bounds: dict[str, int], *, max_items: int = 32) -> list[str]:
    out: list[str] = []
    for row in ws.iter_rows(
        min_row=bounds["min_row"],
        max_row=bounds["max_row"],
        min_col=bounds["min_col"],
        max_col=bounds["max_col"],
    ):
        for cell in row:
            if cell.value is None:
                continue
            out.append(cell_display(cell.value, max_len=140))
            if len(out) >= max_items:
                return out
    return out


def _row_texts(ws, bounds: dict[str, int], *, max_rows: int = 80) -> list[str]:
    rows: list[str] = []
    row_stop = min(bounds["max_row"], bounds["min_row"] + max_rows - 1)
    for row_idx in range(bounds["min_row"], row_stop + 1):
        values = [
            cell_display(ws.cell(row_idx, col_idx).value, max_len=80)
            for col_idx in range(bounds["min_col"], bounds["max_col"] + 1)
        ]
        text = normalize_text(" | ".join(value for value in values if value))
        if text:
            rows.append(text)
    return rows


def detect_regions(
    excel_path: str | Path,
    manifest_path: str | Path,
    classification_path: str | Path,
) -> dict[str, Any]:
    excel_path = ensure_supported_excel(excel_path)
    manifest = load_json(manifest_path)
    classification = load_json(classification_path)
    workbook_type = classification.get("workbook_type") or "unknown"

    if workbook_type == "financial_supplement":
        row_gap, col_gap = 1, 2
    elif workbook_type == "valuation_model":
        row_gap, col_gap = 1, 1
    else:
        row_gap, col_gap = 1, 1

    wb = load_workbook(excel_path, data_only=False, read_only=False, keep_links=False)
    regions: list[dict[str, Any]] = []
    for sheet_index, ws in enumerate(wb.worksheets):
        cells = _used_cells(ws)
        if not cells:
            continue
        rects = _rectangles(cells, row_gap=row_gap, col_gap=col_gap)
        rects.sort(key=lambda item: (item["min_row"], item["min_col"], item["max_row"], item["max_col"]))
        for region_index, bounds in enumerate(rects, start=1):
            values = _cell_values(ws, bounds)
            if not values:
                continue
            formula_count = sum(1 for value in values if is_formula(value))
            region_type = _region_type(ws.title, values, bounds, formula_count)
            region = {
                "region_id": f"{sheet_index + 1}:{region_index}",
                "sheet_index": sheet_index,
                "sheet_name": ws.title,
                "sheet_state": ws.sheet_state,
                "bounds": bounds,
                "cell_range": bounds_to_range(bounds),
                "row_count": bounds["max_row"] - bounds["min_row"] + 1,
                "col_count": bounds["max_col"] - bounds["min_col"] + 1,
                "non_empty_cell_count": len(values),
                "formula_count": formula_count,
                "formula_density": round(formula_count / max(1, len(values)), 4),
                "region_type": region_type,
                "value_preview": _preview(ws, bounds),
                "row_texts": _row_texts(ws, bounds),
                "metadata": {
                    "detector": "tolerant_row_col_rectangles",
                    "row_gap_tolerance": row_gap,
                    "col_gap_tolerance": col_gap,
                    "hidden": ws.sheet_state != "visible",
                    "workbook_type": workbook_type,
                },
            }
            regions.append(region)

    return {
        "source_path": str(excel_path),
        "source_filename": excel_path.name,
        "manifest_path": str(manifest_path),
        "classification_path": str(classification_path),
        "workbook_type": workbook_type,
        "workbook": manifest.get("workbook", {}),
        "regions": regions,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Detect Excel source regions.")
    parser.add_argument("--excel", required=True, help="Path to .xlsx/.xlsm file")
    parser.add_argument("--manifest", required=True, help="Path to workbook_manifest.json")
    parser.add_argument("--classification", required=True, help="Path to classification.json")
    parser.add_argument("--output", required=True, help="Path to regions.json")
    args = parser.parse_args()
    dump_json(detect_regions(args.excel, args.manifest, args.classification), args.output)


if __name__ == "__main__":
    main()
