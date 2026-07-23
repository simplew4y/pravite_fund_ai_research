"""Evidence-grounded Agent analysis and controlled valuation-workbook derivation.

The Agent is deliberately separated from deterministic snapshot/diff logic. It
plans an evidence retrieval pass, synthesizes a structured analysis, and may
propose model changes. A derived workbook is always a new file. Only high-
confidence proposals mapped to non-formula input cells are written; every other
proposal remains visible in the Agent_Analysis sheet for human review.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import sqlite3
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Protocol

from omnigent.server import private_fund_valuation_tracking as valuation

VALUATION_AGENT_VERSION = "valuation-agent-v1"
MAX_MODEL_EVIDENCE = 180
MAX_RESEARCH_EVIDENCE = 120
MAX_SELECTED_EVIDENCE = 24
MIN_AUTO_APPLY_CONFIDENCE = 0.8


class ValuationAgentChatClient(Protocol):
    def chat(
        self,
        messages: list[dict[str, str]],
        *,
        max_tokens: int | None = None,
        temperature: float | None = None,
    ) -> str: ...


def _connect(collection_db: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(collection_db), timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _decode(value: Any, default: Any) -> Any:
    if value in (None, ""):
        return default
    try:
        return json.loads(str(value))
    except (TypeError, ValueError, json.JSONDecodeError):
        return default


def _digest(*parts: Any, length: int = 24) -> str:
    payload = "\0".join(str(part or "") for part in parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length]


def _safe_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _clamp_confidence(value: Any) -> float:
    number = _safe_float(value)
    return min(1.0, max(0.0, number if number is not None else 0.5))


def _parse_json_object(text: str) -> dict[str, Any]:
    candidate = str(text or "").strip()
    fenced = re.search(r"```(?:json)?\s*([\s\S]*?)```", candidate, flags=re.IGNORECASE)
    if fenced:
        candidate = fenced.group(1).strip()
    try:
        payload = json.loads(candidate)
    except json.JSONDecodeError:
        start = candidate.find("{")
        end = candidate.rfind("}")
        if start < 0 or end <= start:
            raise ValueError("Agent response did not contain a JSON object") from None
        payload = json.loads(candidate[start : end + 1])
    if not isinstance(payload, dict):
        raise ValueError("Agent response must be a JSON object")
    return payload


def _chat_json(
    llm_client: ValuationAgentChatClient,
    messages: list[dict[str, str]],
    *,
    max_tokens: int,
    temperature: float,
    attempts: int = 3,
) -> tuple[dict[str, Any], str]:
    """Retry transient provider failures and malformed/truncated JSON outputs."""

    last_error: Exception | None = None
    active_messages = messages
    active_temperature = temperature
    for attempt in range(max(1, attempts)):
        raw = ""
        try:
            raw = llm_client.chat(
                active_messages,
                max_tokens=max_tokens,
                temperature=active_temperature,
            )
            return _parse_json_object(raw), raw
        except ValueError as exc:
            last_error = exc
            if attempt + 1 >= attempts:
                raise
            if raw:
                active_messages = [
                    {
                        "role": "system",
                        "content": (
                            "Repair the supplied truncated or malformed JSON. Return one valid, "
                            "compact JSON object only. Preserve completed content and drop any "
                            "incomplete trailing item if needed."
                        ),
                    },
                    {"role": "user", "content": raw[:60_000]},
                ]
            active_temperature = 0.0
            time.sleep(0.5)
        except Exception as exc:
            last_error = exc
            if attempt + 1 >= attempts:
                raise
            time.sleep(1.5 * (2**attempt))
    raise RuntimeError("Valuation Agent did not return JSON") from last_error


def _tables(conn: sqlite3.Connection) -> set[str]:
    return {
        str(row[0]) for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
    }


def _analysis_payload(row: sqlite3.Row | None) -> dict[str, Any]:
    if row is None:
        return {}
    payload = dict(row)
    analysis = _decode(payload.pop("analysis_json"), {})
    for recommendation in analysis.get("recommended_changes") or []:
        target_evidence_id = str(recommendation.get("target_evidence_id") or "")
        if (
            _looks_like_period_header(recommendation)
            or not target_evidence_id
            or target_evidence_id not in (recommendation.get("evidence_ids") or [])
        ):
            recommendation["writable"] = False
    payload["analysis"] = analysis
    payload["planner"] = _decode(payload.pop("planner_json"), {})
    payload["evidence_ids"] = _decode(payload.pop("evidence_ids_json"), [])
    return payload


def _derived_payload(row: sqlite3.Row | None) -> dict[str, Any]:
    if row is None:
        return {}
    payload = dict(row)
    payload["applied_changes"] = _decode(payload.pop("applied_changes_json"), [])
    payload["skipped_changes"] = _decode(payload.pop("skipped_changes_json"), [])
    payload["resource_status"] = str(payload.get("resource_status") or "not_added")
    return payload


def get_analysis(collection_db: Path, dataset_id: str, analysis_id: str) -> dict[str, Any]:
    with _connect(collection_db) as conn:
        valuation.ensure_valuation_schema(conn, dataset_id)
        row = conn.execute(
            "SELECT * FROM valuation_agent_analyses WHERE dataset_id=? AND analysis_id=?",
            (dataset_id, analysis_id),
        ).fetchone()
        if row is None:
            raise KeyError(analysis_id)
        return _analysis_payload(row)


def list_analyses(
    collection_db: Path, dataset_id: str, *, series_id: str = "", limit: int = 50
) -> list[dict[str, Any]]:
    with _connect(collection_db) as conn:
        valuation.ensure_valuation_schema(conn, dataset_id)
        sql = "SELECT * FROM valuation_agent_analyses WHERE dataset_id=?"
        params: list[Any] = [dataset_id]
        if series_id:
            sql += " AND series_id=?"
            params.append(series_id)
        sql += " ORDER BY created_at DESC LIMIT ?"
        params.append(max(1, min(limit, 200)))
        return [_analysis_payload(row) for row in conn.execute(sql, params)]


def list_derived_models(
    collection_db: Path, dataset_id: str, *, series_id: str = "", limit: int = 50
) -> list[dict[str, Any]]:
    with _connect(collection_db) as conn:
        valuation.ensure_valuation_schema(conn, dataset_id)
        sql = "SELECT * FROM valuation_derived_models WHERE dataset_id=?"
        params: list[Any] = [dataset_id]
        if series_id:
            sql += " AND series_id=?"
            params.append(series_id)
        sql += " ORDER BY created_at DESC LIMIT ?"
        params.append(max(1, min(limit, 200)))
        return [_derived_payload(row) for row in conn.execute(sql, params)]


def _find_imported_document(
    conn: sqlite3.Connection,
    dataset_id: str,
    file_name: str,
    checksum: str,
) -> dict[str, Any] | None:
    if "documents" not in _tables(conn):
        return None
    columns = {
        str(row[1]) for row in conn.execute("PRAGMA table_info(documents)")
    }
    required = {"doc_id", "dataset_id", "original_filename", "checksum"}
    if not required.issubset(columns):
        return None
    filters = ["dataset_id=?", "original_filename=?", "checksum=?"]
    if "deleted_at" in columns:
        filters.append("deleted_at IS NULL")
    order = []
    if "is_current" in columns:
        order.append("is_current DESC")
    if "version_no" in columns:
        order.append("version_no DESC")
    if "created_at" in columns:
        order.append("created_at DESC")
    sql = f"SELECT * FROM documents WHERE {' AND '.join(filters)}"
    if order:
        sql += f" ORDER BY {', '.join(order)}"
    sql += " LIMIT 1"
    row = conn.execute(sql, (dataset_id, file_name, checksum)).fetchone()
    return dict(row) if row is not None else None


def find_imported_document(
    collection_db: Path,
    dataset_id: str,
    file_name: str,
    checksum: str,
) -> dict[str, Any] | None:
    with _connect(collection_db) as conn:
        return _find_imported_document(conn, dataset_id, file_name, checksum)


def mark_resource_import_requested(
    collection_db: Path,
    dataset_id: str,
    derived_model_id: str,
    *,
    file_name: str,
    pipeline_job_id: str,
) -> dict[str, Any]:
    with _connect(collection_db) as conn:
        valuation.ensure_valuation_schema(conn, dataset_id)
        result = conn.execute(
            """
            UPDATE valuation_derived_models
            SET resource_file_name=?, resource_pipeline_job_id=?,
                resource_status='queued', resource_doc_id=NULL,
                resource_added_at=?, resource_error=NULL
            WHERE dataset_id=? AND derived_model_id=?
            """,
            (file_name, pipeline_job_id, _now_iso(), dataset_id, derived_model_id),
        )
        if result.rowcount != 1:
            raise KeyError(derived_model_id)
        conn.commit()
        row = conn.execute(
            "SELECT * FROM valuation_derived_models WHERE derived_model_id=?",
            (derived_model_id,),
        ).fetchone()
        return _derived_payload(row)


def mark_resource_import_completed(
    collection_db: Path,
    dataset_id: str,
    derived_model_id: str,
    *,
    file_name: str,
    doc_id: str,
) -> dict[str, Any]:
    with _connect(collection_db) as conn:
        valuation.ensure_valuation_schema(conn, dataset_id)
        result = conn.execute(
            """
            UPDATE valuation_derived_models
            SET resource_file_name=?, resource_status='completed',
                resource_doc_id=?, resource_added_at=?, resource_error=NULL
            WHERE dataset_id=? AND derived_model_id=?
            """,
            (file_name, doc_id, _now_iso(), dataset_id, derived_model_id),
        )
        if result.rowcount != 1:
            raise KeyError(derived_model_id)
        conn.commit()
        row = conn.execute(
            "SELECT * FROM valuation_derived_models WHERE derived_model_id=?",
            (derived_model_id,),
        ).fetchone()
        return _derived_payload(row)


def update_resource_import_for_pipeline(
    collection_db: Path,
    dataset_id: str,
    pipeline_job_id: str,
    pipeline_status: str,
    *,
    error: str = "",
) -> None:
    """Project a shared pipeline job state onto its originating derived model."""

    with _connect(collection_db) as conn:
        valuation.ensure_valuation_schema(conn, dataset_id)
        derived = conn.execute(
            """
            SELECT * FROM valuation_derived_models
            WHERE dataset_id=? AND resource_pipeline_job_id=?
            """,
            (dataset_id, pipeline_job_id),
        ).fetchone()
        if derived is None:
            return
        normalized = str(pipeline_status or "").lower()
        if normalized in {"completed", "completed_with_warnings"}:
            document = _find_imported_document(
                conn,
                dataset_id,
                str(derived["resource_file_name"] or ""),
                str(derived["checksum"] or ""),
            )
            if document is not None:
                status = "completed"
                doc_id = str(document["doc_id"])
                resource_error = None
            else:
                status = "failed"
                doc_id = None
                resource_error = "Pipeline completed but the imported document was not found."
        elif normalized == "failed":
            status = "failed"
            doc_id = None
            resource_error = str(
                error or "Pipeline failed while indexing the derived model."
            )[:2000]
        elif normalized == "running":
            status = "running"
            doc_id = None
            resource_error = None
        else:
            status = "queued"
            doc_id = None
            resource_error = None
        conn.execute(
            """
            UPDATE valuation_derived_models
            SET resource_status=?, resource_doc_id=?, resource_error=?
            WHERE derived_model_id=?
            """,
            (status, doc_id, resource_error, derived["derived_model_id"]),
        )
        conn.commit()


def enqueue_analysis(
    collection_db: Path,
    dataset_id: str,
    series_id: str,
    *,
    base_model_version_id: str = "",
    comparison_model_version_id: str = "",
    focus: str = "",
) -> dict[str, Any]:
    normalized_focus = str(focus or "").strip()[:2000]
    with _connect(collection_db) as conn:
        valuation.ensure_valuation_schema(conn, dataset_id)
        series = conn.execute(
            "SELECT * FROM valuation_model_series WHERE dataset_id=? AND series_id=?",
            (dataset_id, series_id),
        ).fetchone()
        if series is None:
            raise KeyError(series_id)
        base_id = base_model_version_id or str(series["current_model_version_id"] or "")
        base = conn.execute(
            "SELECT * FROM valuation_model_versions WHERE series_id=? AND model_version_id=?",
            (series_id, base_id),
        ).fetchone()
        if base is None:
            raise KeyError(base_id or "valuation model version")
        comparison_id = comparison_model_version_id
        if not comparison_id:
            previous = conn.execute(
                """
                SELECT model_version_id FROM valuation_model_versions
                WHERE series_id=? AND document_version_no<?
                ORDER BY document_version_no DESC, created_at DESC LIMIT 1
                """,
                (series_id, int(base["document_version_no"])),
            ).fetchone()
            comparison_id = str(previous["model_version_id"]) if previous else ""
        elif (
            conn.execute(
                "SELECT 1 FROM valuation_model_versions WHERE series_id=? AND model_version_id=?",
                (series_id, comparison_id),
            ).fetchone()
            is None
        ):
            raise KeyError(comparison_id)

        analysis_id = (
            f"vaa_{_digest(base_id, comparison_id, normalized_focus, VALUATION_AGENT_VERSION)}"
        )
        now = _now_iso()
        conn.execute(
            """
            INSERT OR IGNORE INTO valuation_agent_analyses
                (analysis_id, dataset_id, series_id, base_model_version_id,
                 comparison_model_version_id, status, focus, agent_version,
                 created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, 'pending', ?, ?, ?, ?)
            """,
            (
                analysis_id,
                dataset_id,
                series_id,
                base_id,
                comparison_id or None,
                normalized_focus,
                VALUATION_AGENT_VERSION,
                now,
                now,
            ),
        )
        conn.execute(
            """
            UPDATE valuation_agent_analyses
            SET status='pending', error_message=NULL, updated_at=?
            WHERE analysis_id=? AND status='failed'
            """,
            (now, analysis_id),
        )
        conn.commit()

    job = valuation.enqueue_job(
        collection_db,
        dataset_id,
        job_type="agent_analysis",
        source_id=analysis_id,
        payload={
            "analysis_id": analysis_id,
            "series_id": series_id,
            "base_model_version_id": base_id,
            "comparison_model_version_id": comparison_id,
            "focus": normalized_focus,
        },
        priority=10,
        requeue_failed=True,
    )
    return {"analysis": get_analysis(collection_db, dataset_id, analysis_id), "job": job}


def _writable_node(row: sqlite3.Row) -> bool:
    return (
        str(row["node_kind"] or "") in {"assumption", "forecast"}
        and not str(row["formula"] or "").strip()
        and bool(str(row["sheet_name"] or "").strip())
        and bool(str(row["cell_ref"] or "").strip())
        and not _looks_like_period_header(row)
    )


def _looks_like_period_header(row: sqlite3.Row | dict[str, Any]) -> bool:
    def value(key: str) -> Any:
        return row.get(key) if isinstance(row, dict) else row[key]

    period_match = re.search(r"(?:19|20)\d{2}", str(value("period") or ""))
    current = _safe_float(value("value_numeric"))
    return bool(
        period_match
        and current is not None
        and current == float(period_match.group(0))
        and not str(value("unit") or "").strip()
    )


def _load_evidence_context(
    collection_db: Path, dataset_id: str, analysis: dict[str, Any]
) -> dict[str, Any]:
    base_id = str(analysis["base_model_version_id"])
    with _connect(collection_db) as conn:
        version = conn.execute(
            """
            SELECT v.*, d.stored_path, d.file_type, s.name AS series_name,
                   s.company_name, s.company_ticker
            FROM valuation_model_versions v
            JOIN valuation_model_series s ON s.series_id=v.series_id
            JOIN documents d ON d.doc_id=v.doc_id
            WHERE v.dataset_id=? AND v.model_version_id=?
            """,
            (dataset_id, base_id),
        ).fetchone()
        if version is None:
            raise KeyError(base_id)
        node_rows = conn.execute(
            """
            SELECT n.node_id, n.canonical_key, n.node_kind, n.metric_key,
                   n.display_name, n.scope, n.period, n.scenario,
                   v.value_numeric, v.value_text, v.unit, v.formula,
                   v.sheet_name, v.cell_ref, v.evidence_id, v.confidence
            FROM valuation_model_node_values v
            JOIN valuation_model_nodes n ON n.node_id=v.node_id
            WHERE v.model_version_id=?
            ORDER BY CASE n.node_kind
                WHEN 'output' THEN 0 WHEN 'assumption' THEN 1
                WHEN 'forecast' THEN 2 ELSE 3 END,
                n.metric_key, n.period, n.display_name
            LIMIT ?
            """,
            (base_id, MAX_MODEL_EVIDENCE),
        ).fetchall()

        catalog: dict[str, dict[str, Any]] = {}
        node_map: dict[str, dict[str, Any]] = {}
        for row in node_rows:
            node = dict(row)
            node["writable"] = _writable_node(row)
            node_map[str(row["node_id"])] = node
            evidence_id = str(row["evidence_id"])
            catalog[evidence_id] = {
                "evidence_id": evidence_id,
                "kind": "model_node",
                "label": str(row["display_name"]),
                "source": f"{row['sheet_name']}!{row['cell_ref']}",
                "detail": (
                    f"{row['node_kind']} {row['metric_key']} period={row['period']} "
                    f"scenario={row['scenario']} value={row['value_numeric']!r} "
                    f"text={row['value_text']!r} unit={row['unit']!r} "
                    f"formula={str(row['formula'] or '')[:500]!r}"
                ),
                "node_id": str(row["node_id"]),
                "writable": node["writable"],
            }

        tables = _tables(conn)
        if {"chunks", "documents"}.issubset(tables):
            for row in conn.execute(
                """
                SELECT c.chunk_id, c.content, d.original_filename,
                       d.doc_type, d.document_date
                FROM chunks c JOIN documents d ON d.doc_id=c.doc_id
                WHERE c.dataset_id=? AND COALESCE(d.is_current,1)=1
                  AND d.doc_id<>?
                ORDER BY COALESCE(d.document_date,'' ) DESC, c.chunk_index
                LIMIT ?
                """,
                (dataset_id, version["doc_id"], MAX_RESEARCH_EVIDENCE),
            ):
                evidence_id = f"chunk:{row['chunk_id']}"
                catalog.setdefault(
                    evidence_id,
                    {
                        "evidence_id": evidence_id,
                        "kind": "research_chunk",
                        "label": str(row["original_filename"]),
                        "source": str(row["original_filename"]),
                        "detail": str(row["content"] or "")[:1600],
                    },
                )

        if {"metric_facts", "documents"}.issubset(tables):
            for row in conn.execute(
                """
                SELECT f.fact_id, f.metric_name, f.period, f.value_numeric,
                       f.value_text, f.unit, f.sheet_name, f.cell_ref,
                       d.original_filename
                FROM metric_facts f JOIN documents d ON d.doc_id=f.doc_id
                WHERE f.dataset_id=? AND COALESCE(d.is_current,1)=1
                  AND d.doc_id<>?
                ORDER BY COALESCE(d.document_date,'') DESC, f.metric_name
                LIMIT ?
                """,
                (dataset_id, version["doc_id"], MAX_RESEARCH_EVIDENCE),
            ):
                evidence_id = f"fact:{row['fact_id']}"
                catalog.setdefault(
                    evidence_id,
                    {
                        "evidence_id": evidence_id,
                        "kind": "research_metric",
                        "label": str(row["metric_name"]),
                        "source": (
                            f"{row['original_filename']} {row['sheet_name']}!{row['cell_ref']}"
                        ),
                        "detail": (
                            f"period={row['period']} value={row['value_numeric']!r} "
                            f"text={row['value_text']!r} unit={row['unit']!r}"
                        ),
                    },
                )

    changes: list[dict[str, Any]] = []
    comparison_id = str(analysis.get("comparison_model_version_id") or "")
    if comparison_id:
        comparison = valuation.compare_versions(
            collection_db,
            dataset_id,
            str(analysis["series_id"]),
            comparison_id,
            base_id,
        )
        changes = comparison["changes"][:100]
    return {
        "version": dict(version),
        "catalog": catalog,
        "node_map": node_map,
        "changes": changes,
    }


def _plan_evidence(
    llm_client: ValuationAgentChatClient,
    analysis: dict[str, Any],
    context: dict[str, Any],
) -> dict[str, Any]:
    catalog = list(context["catalog"].values())
    catalog_text = "\n".join(
        f"[{item['evidence_id']}] {item['kind']} | {item['label']} | "
        f"{item['source']} | {item['detail'][:320]}"
        for item in catalog
    )[:80_000]
    prompt = f"""
You are the planning stage of a valuation-model analysis Agent.
Choose the evidence needed to analyze the model across valuation method,
operating forecasts, assumptions, formula logic, sensitivities, risks, and
investment impact. When a predecessor exists, prioritize evidence that explains
why each material input, output, formula, period, or scenario changed instead of
merely repeating the numeric delta. Return JSON only with:
{{"selected_evidence_ids": ["..."], "analysis_dimensions": ["..."],
  "comparison_questions": ["..."]}}
Select at most {MAX_SELECTED_EVIDENCE} IDs and use only IDs in the catalog.
User focus: {analysis.get("focus") or "全面分析"}

Evidence catalog:
{catalog_text}
""".strip()
    payload, _raw = _chat_json(
        llm_client,
        [
            {
                "role": "system",
                "content": "Plan evidence retrieval for an auditable valuation Agent. JSON only.",
            },
            {"role": "user", "content": prompt},
        ],
        # Evidence IDs are deliberately long, stable identifiers. A full
        # selection plus dimensions can exceed small completion limits and
        # leave otherwise valid JSON truncated mid-string.
        max_tokens=2200,
        temperature=0.0,
    )
    valid_ids = set(context["catalog"])
    selected = [
        str(item) for item in payload.get("selected_evidence_ids") or [] if str(item) in valid_ids
    ][:MAX_SELECTED_EVIDENCE]
    if not selected:
        selected = list(context["catalog"])[:MAX_SELECTED_EVIDENCE]
    return {
        "selected_evidence_ids": list(dict.fromkeys(selected)),
        "analysis_dimensions": [
            str(item)[:200] for item in payload.get("analysis_dimensions") or []
        ][:12],
        "comparison_questions": [
            str(item)[:500] for item in payload.get("comparison_questions") or []
        ][:12],
    }


def _normalize_claims(items: Any, valid_evidence_ids: set[str]) -> list[dict[str, Any]]:
    normalized = []
    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            continue
        evidence_ids = [
            str(value)
            for value in item.get("evidence_ids") or []
            if str(value) in valid_evidence_ids
        ]
        normalized.append(
            {
                "title": str(item.get("title") or item.get("claim") or "分析结论")[:300],
                "detail": str(item.get("detail") or item.get("reasoning") or "")[:4000],
                "impact": str(item.get("impact") or "medium")[:40],
                "confidence": _clamp_confidence(item.get("confidence")),
                "evidence_ids": list(dict.fromkeys(evidence_ids)),
            }
        )
    return normalized[:30]


def _normalize_recommendations(
    items: Any,
    *,
    valid_evidence_ids: set[str],
    node_map: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    normalized = []
    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            continue
        node_id = str(item.get("node_id") or "")
        node = node_map.get(node_id)
        if node is None:
            continue
        proposed_numeric = _safe_float(item.get("proposed_value_numeric"))
        evidence_ids = [
            str(value)
            for value in item.get("evidence_ids") or []
            if str(value) in valid_evidence_ids
        ]
        normalized.append(
            {
                "node_id": node_id,
                "display_name": node["display_name"],
                "metric_key": node["metric_key"],
                "period": node["period"],
                "scenario": node["scenario"],
                "current_value_numeric": node["value_numeric"],
                "current_value_text": node["value_text"],
                "unit": node["unit"],
                "proposed_value_numeric": proposed_numeric,
                "proposed_value_text": str(item.get("proposed_value_text") or "")[:500],
                "rationale": str(item.get("rationale") or "")[:4000],
                "confidence": _clamp_confidence(item.get("confidence")),
                "evidence_ids": list(dict.fromkeys(evidence_ids)),
                "writable": bool(node["writable"] and str(node["evidence_id"]) in evidence_ids),
                "target_evidence_id": node["evidence_id"],
                "sheet_name": node["sheet_name"],
                "cell_ref": node["cell_ref"],
                "formula": node["formula"],
            }
        )
    return normalized[:80]


def _synthesize_analysis(
    llm_client: ValuationAgentChatClient,
    analysis: dict[str, Any],
    context: dict[str, Any],
    planner: dict[str, Any],
) -> tuple[dict[str, Any], str]:
    selected_ids = planner["selected_evidence_ids"]
    selected = [context["catalog"][item] for item in selected_ids]
    evidence_text = "\n\n".join(
        f"[{item['evidence_id']}] {item['kind']} | {item['label']} | "
        f"{item['source']}\n{item['detail']}"
        for item in selected
    )[:100_000]
    writable_targets = [
        {
            key: node.get(key)
            for key in (
                "node_id",
                "display_name",
                "metric_key",
                "period",
                "scenario",
                "value_numeric",
                "value_text",
                "unit",
                "sheet_name",
                "cell_ref",
                "evidence_id",
            )
        }
        for node in context["node_map"].values()
        if node["writable"]
    ][:80]
    change_text = _json(
        [
            {
                key: change.get(key)
                for key in (
                    "display_name",
                    "metric_key",
                    "period",
                    "change_type",
                    "materiality",
                    "summary",
                    "relative_change",
                    "evidence_ids",
                )
            }
            for change in context["changes"]
        ]
    )[:30_000]
    version = context["version"]
    analysis_prompt = f"""
Analyze this valuation model as a senior buy-side valuation Agent. Use only the
provided evidence. Explain the valuation method, operating/financial forecast
changes, assumptions, formula or structure risks, sensitivity, and investment
impact. Distinguish facts from inference. Every finding must cite evidence IDs
from the selected evidence.

Return a compact JSON object only with this shape:
{{
  "valuation_method": "...",
  "executive_summary": "...",
  "investment_conclusion": "...",
  "version_change_summary": "...",
  "key_findings": [
    {{"title":"...", "detail":"...", "impact":"low|medium|high|critical",
      "confidence":0.0, "evidence_ids":["..."]}}
  ],
  "risks": [
    {{"title":"...", "detail":"...", "impact":"...", "confidence":0.0,
      "evidence_ids":["..."]}}
  ],
  "open_questions": ["..."]
}}

Keep the response decision-dense: 5-6 key findings, no more than 3 risks, and
no more than 3 open questions. Keep each detail within 160 Chinese characters.
State whether an explanation is deterministic from the workbook or an inference.
Treat a matching historical snapshot as rollback, not as a new economic change.
Do not invent evidence, values, formulas, or reasons for changes.

Model: {version["series_name"]} v{version["document_version_no"]}
Company: {version["company_name"]} {version["company_ticker"]}
Model type: {version["model_type"]}
User focus: {analysis.get("focus") or "全面分析"}
Analysis dimensions: {_json(planner.get("analysis_dimensions") or [])}
Comparison questions: {_json(planner.get("comparison_questions") or [])}
Version changes: {change_text}

Selected evidence:
{evidence_text}
""".strip()
    analysis_payload, analysis_raw = _chat_json(
        llm_client,
        [
            {
                "role": "system",
                "content": (
                    "You are an evidence-grounded valuation Agent. Return auditable JSON only."
                ),
            },
            {"role": "user", "content": analysis_prompt},
        ],
        max_tokens=1800,
        temperature=0.1,
    )
    valid_ids = set(selected_ids)
    executive_summary = str(analysis_payload.get("executive_summary") or "").strip()[:12000]
    if not executive_summary:
        raise ValueError("Agent response omitted executive_summary")
    action_prompt = f"""
Build the auditable evidence chain and implementation recommendations for this
valuation analysis. Use only selected evidence IDs and writable node IDs below.
Return one compact JSON object only:
{{
  "evidence_chain": [
    {{"claim":"...", "reasoning":"...", "confidence":0.0,
      "evidence_ids":["..."]}}
  ],
  "recommended_changes": [
    {{"node_id":"...", "proposed_value_numeric":0.0,
      "proposed_value_text":"...", "rationale":"...", "confidence":0.0,
      "evidence_ids":["..."]}}
  ]
}}
Return 3-5 evidence-chain items and no more than 5 changes. Keep each reasoning
or rationale within 160 Chinese characters. Recommend a numeric change only
when evidence supports a specific value. Do not invent evidence, values,
formulas, or node IDs. Findings about formula outputs belong in the evidence
chain, not as a direct cell-write recommendation. Every recommended change
must cite both its target evidence_id from Writable targets and the supporting
evidence for the new value.

Executive summary: {executive_summary}
Investment conclusion: {str(analysis_payload.get("investment_conclusion") or "")[:2000]}
Version changes: {change_text}
Writable targets: {_json(writable_targets)}

Selected evidence:
{evidence_text}
""".strip()
    action_payload, action_raw = _chat_json(
        llm_client,
        [
            {
                "role": "system",
                "content": (
                    "You map valuation conclusions to evidence and safe model inputs. "
                    "Return auditable JSON only."
                ),
            },
            {"role": "user", "content": action_prompt},
        ],
        max_tokens=1600,
        temperature=0.0,
    )
    normalized = {
        "valuation_method": str(analysis_payload.get("valuation_method") or "unknown")[:300],
        "executive_summary": executive_summary,
        "investment_conclusion": str(analysis_payload.get("investment_conclusion") or "")[:12000],
        "version_change_summary": str(
            analysis_payload.get("version_change_summary") or ""
        )[:12000],
        "key_findings": _normalize_claims(analysis_payload.get("key_findings"), valid_ids),
        "evidence_chain": _normalize_claims(action_payload.get("evidence_chain"), valid_ids),
        "recommended_changes": _normalize_recommendations(
            action_payload.get("recommended_changes"),
            valid_evidence_ids=valid_ids,
            node_map=context["node_map"],
        ),
        "risks": _normalize_claims(analysis_payload.get("risks"), valid_ids),
        "open_questions": [
            str(item)[:1000] for item in analysis_payload.get("open_questions") or []
        ][:30],
        "selected_evidence": selected,
        "comparison_change_count": len(context["changes"]),
    }
    return normalized, f"{analysis_raw}\n\n--- ACTIONS ---\n\n{action_raw}"


def run_agent_analysis(
    collection_db: Path,
    dataset_id: str,
    analysis_id: str,
    *,
    llm_client: ValuationAgentChatClient | None,
) -> dict[str, Any]:
    if llm_client is None:
        raise RuntimeError("Valuation Agent LLM is not configured")
    analysis = get_analysis(collection_db, dataset_id, analysis_id)
    if analysis["status"] == "completed":
        return analysis
    now = _now_iso()
    with _connect(collection_db) as conn:
        conn.execute(
            """
            UPDATE valuation_agent_analyses
            SET status='running', updated_at=? WHERE analysis_id=?
            """,
            (now, analysis_id),
        )
        conn.commit()
    try:
        context = _load_evidence_context(collection_db, dataset_id, analysis)
        stored_planner = analysis.get("planner") or {}
        stored_ids = stored_planner.get("selected_evidence_ids") or []
        planner = (
            stored_planner
            if stored_ids and all(str(item) in context["catalog"] for item in stored_ids)
            else _plan_evidence(llm_client, analysis, context)
        )
        with _connect(collection_db) as conn:
            conn.execute(
                """
                UPDATE valuation_agent_analyses
                SET planner_json=?, updated_at=? WHERE analysis_id=?
                """,
                (_json(planner), _now_iso(), analysis_id),
            )
            conn.commit()
        result, raw_response = _synthesize_analysis(llm_client, analysis, context, planner)
        evidence_ids = list(
            dict.fromkeys(
                evidence_id
                for group in (
                    result["key_findings"],
                    result["evidence_chain"],
                    result["recommended_changes"],
                    result["risks"],
                )
                for item in group
                for evidence_id in item.get("evidence_ids") or []
            )
        )
        model_name = str(getattr(getattr(llm_client, "config", None), "model_name", ""))
        completed_at = _now_iso()
        with _connect(collection_db) as conn:
            conn.execute(
                """
                UPDATE valuation_agent_analyses
                SET status='completed', valuation_method=?, executive_summary=?,
                    investment_conclusion=?, analysis_json=?, planner_json=?,
                    evidence_ids_json=?, raw_response=?, model_name=?,
                    error_message=NULL, updated_at=?, completed_at=?
                WHERE analysis_id=?
                """,
                (
                    result["valuation_method"],
                    result["executive_summary"],
                    result["investment_conclusion"],
                    _json(result),
                    _json(planner),
                    _json(evidence_ids),
                    raw_response[:100_000],
                    model_name,
                    completed_at,
                    completed_at,
                    analysis_id,
                ),
            )
            conn.commit()
        return get_analysis(collection_db, dataset_id, analysis_id)
    except Exception as exc:
        with _connect(collection_db) as conn:
            conn.execute(
                """
                UPDATE valuation_agent_analyses
                SET status='failed', error_message=?, updated_at=?, completed_at=?
                WHERE analysis_id=?
                """,
                (str(exc)[:4000], _now_iso(), _now_iso(), analysis_id),
            )
            conn.commit()
        raise


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._\-㐀-鿿]+", "_", value).strip("._")
    return cleaned[:160] or "valuation_model"


def _file_checksum(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _auto_apply_guard(recommendation: dict[str, Any]) -> str:
    if not recommendation.get("writable"):
        return "target is a formula or is not a writable assumption/forecast input"
    if float(recommendation.get("confidence") or 0.0) < MIN_AUTO_APPLY_CONFIDENCE:
        return f"confidence is below {MIN_AUTO_APPLY_CONFIDENCE:.0%}"
    proposed = _safe_float(recommendation.get("proposed_value_numeric"))
    if proposed is None:
        return "proposal has no finite numeric value"
    current = _safe_float(recommendation.get("current_value_numeric"))
    metric_key = str(recommendation.get("metric_key") or "")
    if metric_key in {"wacc", "terminal_growth", "risk_free_rate", "cost_of_debt", "tax_rate"}:
        if not -1.0 <= proposed <= 1.0:
            return "rate proposal is outside the safe [-100%, 100%] range"
    if current not in (None, 0.0) and abs(proposed / current) > 10:
        return "proposal exceeds the 10x change guard"
    return ""


def _analysis_sheet(
    workbook: Any, analysis: dict[str, Any], decisions: list[dict[str, Any]]
) -> str:
    from openpyxl.styles import Alignment, Font, PatternFill

    base_name = "Agent_Analysis"
    sheet_name = base_name
    suffix = 2
    while sheet_name in workbook.sheetnames:
        sheet_name = f"{base_name}_{suffix}"
        suffix += 1
    sheet = workbook.create_sheet(sheet_name)
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 32
    sheet.column_dimensions["C"].width = 78
    sheet.column_dimensions["D"].width = 42
    sheet.column_dimensions["E"].width = 14
    header_fill = PatternFill("solid", fgColor="1F4E3D")
    sub_fill = PatternFill("solid", fgColor="DCE9E2")

    rows: list[list[Any]] = [
        ["Valuation Agent Analysis", "", "", "", ""],
        ["Analysis ID", analysis["analysis_id"], "", "", ""],
        ["Agent version", analysis["agent_version"], "", "", ""],
        ["Model", analysis.get("model_name") or "", "", "", ""],
        ["Valuation method", analysis.get("valuation_method") or "", "", "", ""],
        ["Executive summary", "", analysis.get("executive_summary") or "", "", ""],
        ["Investment conclusion", "", analysis.get("investment_conclusion") or "", "", ""],
        ["", "", "", "", ""],
        ["Section", "Title / Target", "Detail / Rationale", "Evidence IDs", "Confidence"],
    ]
    payload = analysis.get("analysis") or {}
    for finding in payload.get("key_findings") or []:
        rows.append(
            [
                "Key finding",
                finding.get("title"),
                finding.get("detail"),
                ", ".join(finding.get("evidence_ids") or []),
                finding.get("confidence"),
            ]
        )
    for chain in payload.get("evidence_chain") or []:
        rows.append(
            [
                "Evidence chain",
                chain.get("title"),
                chain.get("detail"),
                ", ".join(chain.get("evidence_ids") or []),
                chain.get("confidence"),
            ]
        )
    decision_by_node = {str(item.get("node_id")): item for item in decisions}
    for recommendation in payload.get("recommended_changes") or []:
        decision = decision_by_node.get(str(recommendation.get("node_id")), {})
        detail = (
            f"Current={recommendation.get('current_value_numeric')!r}; "
            f"Proposed={recommendation.get('proposed_value_numeric')!r}; "
            f"Decision={decision.get('status', 'skipped')}"
        )
        if decision.get("reason"):
            detail += f"; Reason={decision['reason']}"
        if recommendation.get("rationale"):
            detail += f"\n{recommendation['rationale']}"
        rows.append(
            [
                "Recommended change",
                (
                    f"{recommendation.get('display_name')} "
                    f"({recommendation.get('sheet_name')}!{recommendation.get('cell_ref')})"
                ),
                detail,
                ", ".join(recommendation.get("evidence_ids") or []),
                recommendation.get("confidence"),
            ]
        )
    for row in rows:
        sheet.append(row)
    sheet.merge_cells("A1:E1")
    sheet["A1"].fill = header_fill
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=15)
    sheet["A1"].alignment = Alignment(vertical="center")
    for cell in sheet[9]:
        cell.fill = sub_fill
        cell.font = Font(bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A10"
    return sheet_name


def derive_model_version(
    collection_db: Path,
    dataset_id: str,
    analysis_id: str,
    dataset_root: Path,
) -> dict[str, Any]:
    analysis = get_analysis(collection_db, dataset_id, analysis_id)
    if analysis["status"] != "completed":
        raise ValueError("Agent analysis must be completed before deriving a model")
    existing = list_derived_models(collection_db, dataset_id, series_id=analysis["series_id"])
    for derived in existing:
        if derived["analysis_id"] == analysis_id and Path(derived["output_path"]).is_file():
            return derived

    with _connect(collection_db) as conn:
        source = conn.execute(
            """
            SELECT v.*, d.stored_path, d.original_filename, d.file_type,
                   s.current_version_no
            FROM valuation_model_versions v
            JOIN documents d ON d.doc_id=v.doc_id
            JOIN valuation_model_series s ON s.series_id=v.series_id
            WHERE v.dataset_id=? AND v.model_version_id=?
            """,
            (dataset_id, analysis["base_model_version_id"]),
        ).fetchone()
        if source is None:
            raise KeyError(analysis["base_model_version_id"])
        node_rows = {
            str(row["node_id"]): dict(row)
            for row in conn.execute(
                """
                SELECT n.node_id, n.node_kind, n.metric_key, n.display_name,
                       n.period,
                       v.value_numeric, v.value_text, v.unit, v.formula,
                       v.sheet_name, v.cell_ref, v.evidence_id
                FROM valuation_model_node_values v
                JOIN valuation_model_nodes n ON n.node_id=v.node_id
                WHERE v.model_version_id=?
                """,
                (analysis["base_model_version_id"],),
            )
        }
        next_version = (
            max(
                int(source["current_version_no"] or 0),
                int(source["document_version_no"] or 0),
            )
            + 1
        )

    source_path = Path(str(source["stored_path"] or "")).expanduser().resolve()
    if not source_path.is_file():
        raise FileNotFoundError(source_path)
    suffix = source_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Only XLSX and XLSM models can be derived safely")

    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill

    output_dir = dataset_root / "derived_models"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = (
        f"{_safe_filename(Path(str(source['original_filename'])).stem)}"
        f"_agent_v{next_version}{suffix}"
    )
    output_path = (output_dir / output_name).resolve()
    temporary_path = output_path.with_name(f".{output_path.stem}.tmp{suffix}")
    workbook = load_workbook(
        source_path,
        data_only=False,
        read_only=False,
        keep_vba=suffix == ".xlsm",
        keep_links=True,
    )
    decisions: list[dict[str, Any]] = []
    applied: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    recommendations = (analysis.get("analysis") or {}).get("recommended_changes") or []
    for recommendation in recommendations:
        node_id = str(recommendation.get("node_id") or "")
        node = node_rows.get(node_id)
        decision = {"node_id": node_id, "status": "skipped", "reason": ""}
        reason = _auto_apply_guard(recommendation)
        if node is None:
            reason = "node is not present in the base model version"
        elif str(node["formula"] or "").strip():
            reason = "database snapshot identifies the target as a formula"
        elif _looks_like_period_header(node):
            reason = "target cell appears to be a period header, not a model input"
        elif str(node["evidence_id"]) not in (recommendation.get("evidence_ids") or []):
            reason = "recommendation does not cite the target cell evidence"
        elif not reason:
            sheet_name = str(node["sheet_name"] or "")
            cell_ref = str(node["cell_ref"] or "")
            if sheet_name not in workbook.sheetnames:
                reason = "target sheet is missing in the source workbook"
            else:
                cell = workbook[sheet_name][cell_ref]
                if isinstance(cell, MergedCell):
                    reason = "target is a non-anchor merged cell"
                elif isinstance(cell.value, str) and cell.value.startswith("="):
                    reason = "live workbook target contains a formula"
                else:
                    proposed = float(recommendation["proposed_value_numeric"])
                    cell.value = proposed
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    cell.comment = Comment(
                        (
                            f"Applied by {VALUATION_AGENT_VERSION}.\n"
                            f"Rationale: {recommendation.get('rationale') or ''}\n"
                            f"Evidence: {', '.join(recommendation.get('evidence_ids') or [])}"
                        )[:32000],
                        "Valuation Agent",
                    )
                    decision.update(
                        {
                            "status": "applied",
                            "sheet_name": sheet_name,
                            "cell_ref": cell_ref,
                            "old_value": recommendation.get("current_value_numeric"),
                            "new_value": proposed,
                        }
                    )
        if reason:
            decision["reason"] = reason
            skipped.append(decision)
        else:
            applied.append(decision)
        decisions.append(decision)

    analysis_sheet_name = _analysis_sheet(workbook, analysis, decisions)
    if getattr(workbook, "calculation", None) is not None:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    workbook.save(temporary_path)
    temporary_path.replace(output_path)
    checksum = _file_checksum(output_path)
    derived_model_id = f"vdm_{_digest(analysis_id, checksum)}"
    created_at = _now_iso()
    applied_payload = [{**item, "analysis_sheet_name": analysis_sheet_name} for item in applied]
    with _connect(collection_db) as conn:
        conn.execute(
            """
            INSERT INTO valuation_derived_models
                (derived_model_id, dataset_id, series_id, analysis_id,
                 base_model_version_id, derived_version_no, output_filename,
                 output_path, checksum, applied_changes_json,
                 skipped_changes_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(analysis_id) DO UPDATE SET
                derived_model_id=excluded.derived_model_id,
                output_filename=excluded.output_filename,
                output_path=excluded.output_path,
                checksum=excluded.checksum,
                applied_changes_json=excluded.applied_changes_json,
                skipped_changes_json=excluded.skipped_changes_json,
                created_at=excluded.created_at
            """,
            (
                derived_model_id,
                dataset_id,
                analysis["series_id"],
                analysis_id,
                analysis["base_model_version_id"],
                next_version,
                output_name,
                str(output_path),
                checksum,
                _json(applied_payload),
                _json(skipped),
                created_at,
            ),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM valuation_derived_models WHERE analysis_id=?", (analysis_id,)
        ).fetchone()
    return _derived_payload(row)


def get_derived_model(
    collection_db: Path, dataset_id: str, derived_model_id: str
) -> dict[str, Any]:
    with _connect(collection_db) as conn:
        valuation.ensure_valuation_schema(conn, dataset_id)
        row = conn.execute(
            "SELECT * FROM valuation_derived_models WHERE dataset_id=? AND derived_model_id=?",
            (dataset_id, derived_model_id),
        ).fetchone()
        if row is None:
            raise KeyError(derived_model_id)
        return _derived_payload(row)
