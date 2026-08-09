from __future__ import annotations

import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, cast

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from omnigent.server import private_fund_valuation_agent as valuation_agent
from omnigent.server import private_fund_valuation_tracking as valuation
from omnigent.server import private_fund_valuation_worker as valuation_worker
from omnigent.server.routes import private_fund_pdf


def _create_collection(path: Path) -> None:
    with sqlite3.connect(path) as conn:
        conn.executescript(
            """
            CREATE TABLE documents (
                doc_id TEXT PRIMARY KEY,
                dataset_id TEXT NOT NULL,
                logical_doc_id TEXT,
                version_no INTEGER NOT NULL,
                is_current INTEGER NOT NULL,
                lifecycle_state TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                doc_type TEXT,
                doc_subtype TEXT,
                classification_status TEXT,
                company_name TEXT,
                company_ticker TEXT,
                document_date TEXT,
                stored_path TEXT,
                file_type TEXT,
                checksum TEXT NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            CREATE TABLE excel_sheets (
                doc_id TEXT NOT NULL,
                sheet_name TEXT NOT NULL,
                sheet_role TEXT NOT NULL
            );
            CREATE TABLE metric_facts (
                fact_id TEXT PRIMARY KEY,
                dataset_id TEXT NOT NULL,
                doc_id TEXT NOT NULL,
                metric_name TEXT NOT NULL,
                period TEXT,
                value_text TEXT,
                value_numeric REAL,
                unit TEXT,
                sheet_name TEXT NOT NULL,
                cell_ref TEXT NOT NULL,
                source_range TEXT,
                formula TEXT,
                confidence REAL,
                fact_status TEXT,
                quality_status TEXT,
                quality_issues_json TEXT,
                metadata_json TEXT
            );
            """
        )


def _insert_model(
    path: Path,
    *,
    doc_id: str,
    logical_doc_id: str,
    version_no: int,
    current: bool,
    checksum: str,
    target_price: float,
    wacc: float,
    revenue: float,
    stored_path: str = "",
) -> None:
    with sqlite3.connect(path) as conn:
        if current:
            conn.execute(
                """
                UPDATE documents SET is_current=0, lifecycle_state='superseded'
                WHERE logical_doc_id=?
                """,
                (logical_doc_id,),
            )
        conn.execute(
            """
            INSERT INTO documents
                (doc_id, dataset_id, logical_doc_id, version_no, is_current,
                 lifecycle_state, original_filename, doc_type, doc_subtype,
                 classification_status, company_name, company_ticker,
                 document_date, stored_path, file_type, checksum, status, created_at)
            VALUES (?, 'demo', ?, ?, ?, ?, ?, 'valuation_model', 'dcf_model',
                    'accepted', 'Demo Corp', 'DEMO', '2026-07-15', ?, 'xlsx', ?,
                    'indexed', ?)
            """,
            (
                doc_id,
                logical_doc_id,
                version_no,
                int(current),
                "active" if current else "superseded",
                f"Demo_Model_v{version_no}.xlsx",
                stored_path,
                checksum,
                f"2026-07-15T00:00:0{version_no}+00:00",
            ),
        )
        conn.execute(
            """
            INSERT INTO excel_sheets (doc_id, sheet_name, sheet_role)
            VALUES (?, 'DCF', 'valuation_dcf')
            """,
            (doc_id,),
        )
        facts = (
            ("Target Price", "Current", target_price, "USD", "E10", "=E8/E9"),
            ("WACC", "Long term", wacc, "%", "E5", ""),
            ("Revenue", "2027E", revenue, "USDm", "E20", "=D20*(1+E19)"),
            ("Decorative label", "", 1.0, "", "Z99", ""),
        )
        for index, (name, period, value, unit, cell_ref, formula) in enumerate(facts):
            conn.execute(
                """
                INSERT INTO metric_facts
                    (fact_id, dataset_id, doc_id, metric_name, period, value_text,
                     value_numeric, unit, sheet_name, cell_ref, source_range,
                     formula, confidence, fact_status, quality_status,
                     quality_issues_json, metadata_json)
                VALUES (?, 'demo', ?, ?, ?, ?, ?, ?, 'DCF', ?, ?, ?, 0.9,
                        'candidate', 'candidate_complete', '[]', '{}')
                """,
                (
                    f"fact-{doc_id}-{index}",
                    doc_id,
                    name,
                    period,
                    str(value),
                    value,
                    unit,
                    cell_ref,
                    f"DCF!{cell_ref}",
                    formula,
                ),
            )


def _drain(path: Path) -> list[dict[str, object]]:
    results = []
    while result := valuation.process_next_job(path, "demo"):
        results.append(result)
    return results


class _FakeValuationAgent:
    class _Config:
        model_name = "test-valuation-agent"

    config = _Config()

    def __init__(self, responses: list[dict[str, Any]]) -> None:
        self._responses = [json.dumps(item, ensure_ascii=False) for item in responses]

    def chat(
        self,
        messages: list[dict[str, str]],
        *,
        max_tokens: int | None = None,
        temperature: float | None = None,
    ) -> str:
        del messages, max_tokens, temperature
        return self._responses.pop(0)


def test_period_header_is_never_treated_as_a_writable_model_input() -> None:
    assert valuation_agent._looks_like_period_header(
        {"period": "2025E", "value_numeric": 2025, "unit": ""}
    )
    assert not valuation_agent._looks_like_period_header(
        {"period": "2025E", "value_numeric": 3683.9, "unit": "RMBm"}
    )


def test_builds_series_versions_changes_and_alerts(tmp_path: Path) -> None:
    database = tmp_path / "collection.sqlite3"
    _create_collection(database)
    _insert_model(
        database,
        doc_id="doc-v1",
        logical_doc_id="logical-demo",
        version_no=1,
        current=False,
        checksum="checksum-v1",
        target_price=100,
        wacc=0.10,
        revenue=1000,
    )
    _insert_model(
        database,
        doc_id="doc-v2",
        logical_doc_id="logical-demo",
        version_no=2,
        current=True,
        checksum="checksum-v2",
        target_price=120,
        wacc=0.11,
        revenue=1110,
    )

    jobs = valuation.enqueue_model_documents(database, "demo", include_history=True)
    assert [job["status"] for job in jobs] == ["queued", "queued"]
    completed = _drain(database)
    assert len(completed) == 2
    assert all(job["status"] == "completed" for job in completed)

    overview = valuation.tracking_overview(database, "demo")
    assert len(overview["series"]) == 1
    series = overview["series"][0]
    assert series["version_count"] == 2
    assert series["current_version_no"] == 2
    assert series["current_version"]["node_count"] == 3
    assert overview["unread_alert_count"] >= 2

    versions = list(reversed(series["versions"]))
    comparison = valuation.compare_versions(
        database,
        "demo",
        series["series_id"],
        versions[0]["model_version_id"],
        versions[1]["model_version_id"],
    )
    by_metric = {change["metric_key"]: change for change in comparison["changes"]}
    assert by_metric["target_price"]["materiality"] == "high"
    assert by_metric["wacc"]["materiality"] == "high"
    assert by_metric["revenue"]["materiality"] == "high"
    assert all(change["evidence_ids"] for change in comparison["changes"])


def test_detects_revert_and_keeps_model_series_separate(tmp_path: Path) -> None:
    database = tmp_path / "collection.sqlite3"
    _create_collection(database)
    _insert_model(
        database,
        doc_id="first-a",
        logical_doc_id="logical-a",
        version_no=1,
        current=False,
        checksum="checksum-a-v1",
        target_price=100,
        wacc=0.10,
        revenue=1000,
    )
    _insert_model(
        database,
        doc_id="second-a",
        logical_doc_id="logical-a",
        version_no=2,
        current=False,
        checksum="checksum-a-v2",
        target_price=120,
        wacc=0.11,
        revenue=1100,
    )
    _insert_model(
        database,
        doc_id="third-a",
        logical_doc_id="logical-a",
        version_no=3,
        current=True,
        checksum="checksum-a-v1",
        target_price=100,
        wacc=0.10,
        revenue=1000,
    )
    _insert_model(
        database,
        doc_id="first-b",
        logical_doc_id="logical-b",
        version_no=1,
        current=True,
        checksum="checksum-b-v1",
        target_price=100,
        wacc=0.10,
        revenue=1000,
    )

    valuation.enqueue_model_documents(database, "demo", include_history=True)
    _drain(database)
    overview = valuation.tracking_overview(database, "demo")
    assert len(overview["series"]) == 2
    series_a = next(item for item in overview["series"] if item["series_key"] == "logical-a")
    assert (
        series_a["current_version"]["reverted_to_version_id"]
        == series_a["versions"][-1]["model_version_id"]
    )


def test_rule_and_alert_lifecycle(tmp_path: Path) -> None:
    database = tmp_path / "collection.sqlite3"
    _create_collection(database)
    _insert_model(
        database,
        doc_id="doc-v1",
        logical_doc_id="logical-demo",
        version_no=1,
        current=False,
        checksum="checksum-v1",
        target_price=100,
        wacc=0.10,
        revenue=1000,
    )
    _insert_model(
        database,
        doc_id="doc-v2",
        logical_doc_id="logical-demo",
        version_no=2,
        current=True,
        checksum="checksum-v2",
        target_price=120,
        wacc=0.11,
        revenue=1100,
    )
    valuation.enqueue_model_documents(database, "demo", include_history=True)
    _drain(database)

    overview = valuation.tracking_overview(database, "demo")
    rule = overview["watch_rules"][0]
    updated_rule = valuation.update_rule(
        database,
        "demo",
        rule["rule_id"],
        active=False,
        min_materiality="high",
    )
    assert updated_rule["active"] == 0
    assert updated_rule["min_materiality"] == "high"

    alert = overview["alerts"][0]
    updated_alert = valuation.update_alert_status(
        database,
        "demo",
        alert["alert_id"],
        status="acknowledged",
    )
    assert updated_alert["status"] == "acknowledged"


def test_valuation_tracking_http_api_exposes_comparison_and_lifecycle(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    database = tmp_path / "collection.sqlite3"
    _create_collection(database)
    _insert_model(
        database,
        doc_id="doc-v1",
        logical_doc_id="logical-demo",
        version_no=1,
        current=False,
        checksum="checksum-v1",
        target_price=100,
        wacc=0.10,
        revenue=1000,
    )
    _insert_model(
        database,
        doc_id="doc-v2",
        logical_doc_id="logical-demo",
        version_no=2,
        current=True,
        checksum="checksum-v2",
        target_price=120,
        wacc=0.11,
        revenue=1100,
    )
    valuation.enqueue_model_documents(database, "demo", include_history=True)
    _drain(database)

    monkeypatch.setattr(
        private_fund_pdf,
        "_require_project_row",
        lambda dataset_id: {"dataset_id": dataset_id},
    )
    monkeypatch.setattr(private_fund_pdf, "_collection_db_path", lambda dataset_id: database)
    monkeypatch.setattr(
        private_fund_pdf,
        "_project_dataset_root",
        lambda dataset_id: tmp_path / "demo",
    )
    app = FastAPI()
    app.include_router(
        private_fund_pdf.create_private_fund_pdf_router(workspace=cast(Any, object())),
        prefix="/v1",
    )
    client = TestClient(app)

    overview_response = client.get("/v1/private-fund/projects/demo/valuation-tracking")
    assert overview_response.status_code == 200
    overview = overview_response.json()
    assert len(overview["series"]) == 1
    assert overview["analyzer_version"] == "valuation-tracking-v1"

    series = overview["series"][0]
    versions = list(reversed(series["versions"]))
    comparison = client.get(
        f"/v1/private-fund/projects/demo/valuation-models/{series['series_id']}/compare",
        params={
            "from_version": versions[0]["model_version_id"],
            "to_version": versions[1]["model_version_id"],
        },
    )
    assert comparison.status_code == 200
    assert {change["metric_key"] for change in comparison.json()["changes"]} == {
        "target_price",
        "wacc",
        "revenue",
    }

    model_overview = client.get(
        f"/v1/private-fund/projects/demo/valuation-models/{series['series_id']}"
        f"/versions/{versions[1]['model_version_id']}/overview"
    )
    assert model_overview.status_code == 200
    assert model_overview.json()["overview"]["model_version_no"] == 2
    assert model_overview.json()["html"].startswith("<!DOCTYPE html>")

    queued = client.post("/v1/private-fund/projects/demo/valuation-tracking/run")
    assert queued.status_code == 202
    assert len(queued.json()["jobs"]) == 2

    alert = overview["alerts"][0]
    acknowledged = client.patch(
        f"/v1/private-fund/projects/demo/valuation-alerts/{alert['alert_id']}",
        json={"status": "acknowledged"},
    )
    assert acknowledged.status_code == 200
    assert acknowledged.json()["alert"]["status"] == "acknowledged"

    rule = overview["watch_rules"][0]
    disabled = client.patch(
        f"/v1/private-fund/projects/demo/valuation-watch-rules/{rule['rule_id']}",
        json={"active": False, "min_materiality": "high"},
    )
    assert disabled.status_code == 200
    assert disabled.json()["watch_rule"]["active"] == 0
    assert disabled.json()["watch_rule"]["min_materiality"] == "high"


def test_standalone_valuation_worker_discovers_and_drains_models(tmp_path: Path) -> None:
    workspace = tmp_path / "datasets"
    database = workspace / "demo" / "meta" / "collection.sqlite3"
    database.parent.mkdir(parents=True)
    _create_collection(database)
    _insert_model(
        database,
        doc_id="doc-v1",
        logical_doc_id="logical-demo",
        version_no=1,
        current=True,
        checksum="checksum-v1",
        target_price=100,
        wacc=0.10,
        revenue=1000,
    )

    processed = valuation_worker.run_cycle(workspace, max_jobs_per_db=4)

    assert processed == 1
    assert valuation.list_jobs(database, "demo")[0]["status"] == "completed"
    health = (workspace / ".valuation-tracking-worker.json").read_text(encoding="utf-8")
    assert '"status": "online"' in health
    assert '"dataset_count": 1' in health


def test_market_refresh_bucket_is_hourly_and_uses_local_timezone(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("PRIVATE_FUND_MARKET_REFRESH_TIMEZONE", "Asia/Shanghai")
    monkeypatch.setenv("PRIVATE_FUND_MARKET_REFRESH_INTERVAL_MINUTES", "60")

    first_hour = datetime(2026, 7, 20, 2, 34, tzinfo=timezone.utc)
    next_hour = datetime(2026, 7, 20, 3, 1, tzinfo=timezone.utc)

    assert valuation_worker._market_refresh_bucket(first_hour) == "2026-07-20T10:00+08:00"
    assert valuation_worker._market_refresh_bucket(next_hour) == "2026-07-20T11:00+08:00"


def test_market_refresh_job_is_idempotent_per_hourly_bucket(tmp_path: Path) -> None:
    database = tmp_path / "collection.sqlite3"
    _create_collection(database)

    first = valuation.enqueue_market_data_refresh(
        database, "demo", refresh_bucket="2026-07-20T11:00+08:00"
    )
    second = valuation.enqueue_market_data_refresh(
        database, "demo", refresh_bucket="2026-07-20T11:00+08:00"
    )

    assert first["job_id"] == second["job_id"]
    jobs = [
        job
        for job in valuation.list_jobs(database, "demo")
        if job["job_type"] == "market_data_refresh"
    ]
    assert len(jobs) == 1


def test_agent_analysis_and_one_click_derived_model_are_auditable(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source_path = tmp_path / "Demo_Model_v1.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DCF"
    sheet["E5"] = 0.10
    sheet["E10"] = "=E8/E9"
    sheet["E20"] = "=D20*(1+E19)"
    workbook.save(source_path)

    database = tmp_path / "collection.sqlite3"
    _create_collection(database)
    _insert_model(
        database,
        doc_id="doc-v1",
        logical_doc_id="logical-demo",
        version_no=1,
        current=True,
        checksum="checksum-v1",
        target_price=100,
        wacc=0.10,
        revenue=1000,
        stored_path=str(source_path),
    )
    valuation.enqueue_model_documents(database, "demo", include_history=True)
    _drain(database)
    overview = valuation.tracking_overview(database, "demo")
    series = overview["series"][0]
    with sqlite3.connect(database) as conn:
        node_rows = conn.execute(
            """
            SELECT n.metric_key, n.node_id
            FROM valuation_model_nodes n
            JOIN valuation_model_node_values v ON v.node_id=n.node_id
            WHERE v.model_version_id=?
            """,
            (series["current_model_version_id"],),
        ).fetchall()
    node_ids = dict(node_rows)

    planner = {
        "selected_evidence_ids": [
            "fact:fact-doc-v1-0",
            "fact:fact-doc-v1-1",
            "missing:evidence",
        ],
        "analysis_dimensions": ["估值假设", "目标价逻辑"],
        "comparison_questions": [],
    }
    synthesis = {
        "valuation_method": "DCF",
        "executive_summary": "WACC 有下调空间，但目标价仍应由模型公式计算。",
        "investment_conclusion": "建议先验证折现率假设，再观察模型重算结果。",
        "key_findings": [
            {
                "title": "WACC 可下调",
                "detail": "证据支持从 10% 调整到 9%。",
                "impact": "high",
                "confidence": 0.95,
                "evidence_ids": ["fact:fact-doc-v1-1", "missing:evidence"],
            }
        ],
        "evidence_chain": [
            {
                "claim": "折现率变化影响估值结果",
                "reasoning": "WACC 是 DCF 的关键输入。",
                "confidence": 0.9,
                "evidence_ids": ["fact:fact-doc-v1-1"],
            }
        ],
        "recommended_changes": [
            {
                "node_id": node_ids["wacc"],
                "proposed_value_numeric": 0.09,
                "rationale": "基于假设复核结果下调 100bp。",
                "confidence": 0.95,
                "evidence_ids": ["fact:fact-doc-v1-1"],
            },
            {
                "node_id": node_ids["target_price"],
                "proposed_value_numeric": 125,
                "rationale": "用作公式结果复核，不应直接覆盖。",
                "confidence": 0.95,
                "evidence_ids": ["fact:fact-doc-v1-0"],
            },
        ],
        "risks": [],
        "open_questions": ["WACC 下调依据是否通过投委会复核？"],
    }
    actions = {
        "evidence_chain": synthesis["evidence_chain"],
        "recommended_changes": synthesis["recommended_changes"],
    }
    fake_agent = _FakeValuationAgent([planner, synthesis, actions])

    monkeypatch.setattr(
        private_fund_pdf,
        "_require_project_row",
        lambda dataset_id: {"dataset_id": dataset_id},
    )
    monkeypatch.setattr(private_fund_pdf, "_collection_db_path", lambda dataset_id: database)
    monkeypatch.setattr(
        private_fund_pdf,
        "_project_dataset_root",
        lambda dataset_id: tmp_path / "demo",
    )
    app = FastAPI()
    app.include_router(
        private_fund_pdf.create_private_fund_pdf_router(workspace=cast(Any, object())),
        prefix="/v1",
    )
    client = TestClient(app)

    queued = client.post(
        f"/v1/private-fund/projects/demo/valuation-models/{series['series_id']}/agent-analysis",
        json={"focus": "重点分析折现率与目标价变化"},
    )
    assert queued.status_code == 202
    analysis_id = queued.json()["analysis"]["analysis_id"]
    processed = valuation.process_next_job(database, "demo", llm_client=fake_agent)
    assert processed is not None
    assert processed["status"] == "completed"

    analysis_response = client.get(
        f"/v1/private-fund/projects/demo/valuation-agent-analyses/{analysis_id}"
    )
    assert analysis_response.status_code == 200
    analysis = analysis_response.json()["analysis"]
    assert analysis["status"] == "completed"
    assert analysis["model_name"] == "test-valuation-agent"
    assert analysis["evidence_ids"] == ["fact:fact-doc-v1-1", "fact:fact-doc-v1-0"]
    assert analysis["analysis"]["key_findings"][0]["evidence_ids"] == ["fact:fact-doc-v1-1"]
    recommendations = {
        item["metric_key"]: item for item in analysis["analysis"]["recommended_changes"]
    }
    assert recommendations["wacc"]["writable"] is True
    assert recommendations["target_price"]["writable"] is False

    derived_response = client.post(
        f"/v1/private-fund/projects/demo/valuation-agent-analyses/{analysis_id}/derive-model"
    )
    assert derived_response.status_code == 201
    derived = derived_response.json()["derived_model"]
    assert len(derived["applied_changes"]) == 1
    assert len(derived["skipped_changes"]) == 1
    assert "formula" in derived["skipped_changes"][0]["reason"]

    output_path = Path(derived["output_path"])
    generated = load_workbook(output_path, data_only=False)
    assert generated["DCF"]["E5"].value == pytest.approx(0.09)
    assert generated["DCF"]["E10"].value == "=E8/E9"
    assert "Agent_Analysis" in generated.sheetnames
    original = load_workbook(source_path, data_only=False)
    assert original["DCF"]["E5"].value == pytest.approx(0.10)

    download = client.get(
        f"/v1/private-fund/projects/demo/valuation-derived-models/"
        f"{derived['derived_model_id']}/file"
    )
    assert download.status_code == 200
    assert download.content == output_path.read_bytes()

    uploads_dir = tmp_path / "uploads"
    uploads_dir.mkdir()
    queued_pipeline = {
        "job_id": "pipeline-derived-v2",
        "dataset_id": "demo",
        "status": "queued",
    }
    monkeypatch.setattr(
        private_fund_pdf,
        "_seed_uploads_from_raw",
        lambda dataset_id: uploads_dir,
    )
    monkeypatch.setattr(
        private_fund_pdf,
        "_mark_project_uploads_changed",
        lambda dataset_id, directory: None,
    )
    monkeypatch.setattr(
        private_fund_pdf,
        "_queue_project_pipeline_job",
        lambda dataset_id, background_tasks, request=None: queued_pipeline,
    )
    monkeypatch.setattr(
        private_fund_pdf,
        "_get_project_pipeline_job_payload",
        lambda job_id: queued_pipeline if job_id == queued_pipeline["job_id"] else None,
    )

    imported_response = client.post(
        f"/v1/private-fund/projects/demo/valuation-derived-models/"
        f"{derived['derived_model_id']}/add-to-resources"
    )
    assert imported_response.status_code == 202
    imported = imported_response.json()
    assert imported["job"]["job_id"] == "pipeline-derived-v2"
    assert imported["resource_import"] == {
        "status": "queued",
        "file_name": "Demo_Model_v1.xlsx",
        "already_added": False,
        "copied": True,
    }
    resource_copy = uploads_dir / "Demo_Model_v1.xlsx"
    assert resource_copy.read_bytes() == output_path.read_bytes()
    assert imported["derived_model"]["resource_status"] == "queued"

    duplicate_response = client.post(
        f"/v1/private-fund/projects/demo/valuation-derived-models/"
        f"{derived['derived_model_id']}/add-to-resources"
    )
    assert duplicate_response.status_code == 202
    assert duplicate_response.json()["resource_import"]["copied"] is False
    assert duplicate_response.json()["job"]["job_id"] == "pipeline-derived-v2"

    with sqlite3.connect(database) as conn:
        conn.execute(
            """
            UPDATE documents SET is_current=0, lifecycle_state='superseded'
            WHERE logical_doc_id='logical-demo'
            """
        )
        conn.execute(
            """
            INSERT INTO documents
                (doc_id, dataset_id, logical_doc_id, version_no, is_current,
                 lifecycle_state, original_filename, doc_type, doc_subtype,
                 classification_status, company_name, company_ticker,
                 document_date, stored_path, file_type, checksum, status, created_at)
            VALUES ('doc-derived-v2', 'demo', 'logical-demo', 2, 1, 'active',
                    'Demo_Model_v1.xlsx', 'valuation_model', 'dcf_model', 'accepted',
                    'Demo Corp', 'DEMO', '2026-07-15', ?, 'xlsx', ?, 'indexed',
                    '2026-07-15T00:00:02+00:00')
            """,
            (str(resource_copy), derived["checksum"]),
        )
        conn.commit()
    valuation_agent.update_resource_import_for_pipeline(
        database,
        "demo",
        "pipeline-derived-v2",
        "completed",
    )
    completed_import = valuation_agent.get_derived_model(
        database,
        "demo",
        derived["derived_model_id"],
    )
    assert completed_import["resource_status"] == "completed"
    assert completed_import["resource_doc_id"] == "doc-derived-v2"
