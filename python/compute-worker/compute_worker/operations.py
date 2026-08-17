"""Pure compute operation implementations."""

import base64
import datetime as dt
import math
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple

from .artifacts import artifact_descriptor, sha256_file, write_ndjson_atomic
from .document_extract import extract_document
from .errors import (
    ComputeOperationError,
    DependencyUnavailableError,
    UnsupportedOperationError,
)
from .paths import (
    PathValidationError,
    generated_output_path,
    prepare_output_directory,
    require_input_file,
)
from .market_data import fetch_market_data
from .pdf_render import render_pdf_page
from .protocol import PROTOCOL_VERSION, failed_response
from .report_render import render_report
from .workbook_derive import derive_workbook


def _load_fitz() -> Any:
    try:
        import fitz  # type: ignore
    except ImportError as exc:
        raise DependencyUnavailableError(
            "extract_pdf requires PyMuPDF; install the compute-worker dependencies"
        ) from exc
    return fitz


def _load_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise DependencyUnavailableError(
            "extract_workbook requires openpyxl; install the compute-worker dependencies"
        ) from exc
    return openpyxl


def _bool_option(
    options: Mapping[str, Any], name: str, default: bool
) -> bool:
    value = options.get(name, default)
    if not isinstance(value, bool):
        raise ComputeOperationError(
            "options.{} must be boolean".format(name), "invalid_options"
        )
    return value


def _positive_int_option(
    options: Mapping[str, Any],
    name: str,
    default: int,
    maximum: Optional[int] = None,
) -> int:
    value = options.get(name, default)
    if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
        raise ComputeOperationError(
            "options.{} must be a positive integer".format(name),
            "invalid_options",
        )
    if maximum is not None and value > maximum:
        raise ComputeOperationError(
            "options.{} may not exceed {}".format(name, maximum),
            "invalid_options",
        )
    return value


def _finite_float(value: Any) -> float:
    converted = float(value)
    if not math.isfinite(converted):
        raise ComputeOperationError(
            "document geometry contains a non-finite value", "invalid_document"
        )
    return converted


def _page_blocks(page: Any) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []
    raw_blocks = page.get_text("blocks", sort=True)
    for index, raw in enumerate(raw_blocks):
        if not isinstance(raw, (tuple, list)) or len(raw) < 5:
            continue
        text = str(raw[4] or "")
        block_number = raw[5] if len(raw) > 5 else index
        block_type = raw[6] if len(raw) > 6 else 0
        blocks.append(
            {
                "blockNumber": int(block_number),
                "blockType": int(block_type),
                "bbox": [
                    _finite_float(raw[0]),
                    _finite_float(raw[1]),
                    _finite_float(raw[2]),
                    _finite_float(raw[3]),
                ],
                "text": text,
            }
        )
    return blocks


def _extract_pdf(
    input_path: Path, output_directory: Path, options: Mapping[str, Any]
) -> Tuple[str, List[Dict[str, Any]], Dict[str, Any]]:
    fitz = _load_fitz()
    include_blocks = _bool_option(options, "includeBlocks", True)
    max_text_chars = _positive_int_option(
        options, "maxTextCharsPerPage", 5_000_000, 50_000_000
    )
    password = options.get("password")
    if password is not None and not isinstance(password, str):
        raise ComputeOperationError(
            "options.password must be a string", "invalid_options"
        )

    try:
        document = fitz.open(str(input_path))
    except Exception as exc:
        raise ComputeOperationError(
            "PDF could not be opened: {}".format(exc), "invalid_pdf"
        ) from exc

    try:
        if not bool(getattr(document, "is_pdf", True)):
            raise ComputeOperationError(
                "inputPath is not a PDF document", "invalid_pdf"
            )
        needs_password = bool(getattr(document, "needs_pass", False))
        if needs_password:
            if not password:
                raise ComputeOperationError(
                    "PDF is encrypted and options.password was not supplied",
                    "encrypted_pdf",
                )
            if not bool(document.authenticate(password)):
                raise ComputeOperationError(
                    "PDF password authentication failed", "encrypted_pdf"
                )

        page_count = int(document.page_count)
        start_page = _positive_int_option(options, "startPage", 1)
        end_page = _positive_int_option(options, "endPage", page_count or 1)
        if page_count == 0:
            start_page = 1
            end_page = 0
        elif start_page > page_count or end_page > page_count:
            raise ComputeOperationError(
                "requested PDF page range exceeds page count",
                "invalid_options",
            )
        if end_page < start_page and page_count > 0:
            raise ComputeOperationError(
                "options.endPage must be greater than or equal to startPage",
                "invalid_options",
            )

        extracted_pages = max(0, end_page - start_page + 1)

        def page_records() -> Iterable[Dict[str, Any]]:
            for page_index in range(start_page - 1, end_page):
                page = document.load_page(page_index)
                text = str(page.get_text("text", sort=True))
                if len(text) > max_text_chars:
                    raise ComputeOperationError(
                        "PDF page {} exceeds maxTextCharsPerPage".format(
                            page_index + 1
                        ),
                        "document_limit_exceeded",
                    )
                rect = page.rect
                record: Dict[str, Any] = {
                    "recordType": "pdf_page",
                    "sourceName": input_path.name,
                    "pageNumber": page_index + 1,
                    "width": _finite_float(rect.width),
                    "height": _finite_float(rect.height),
                    "rotation": int(getattr(page, "rotation", 0)),
                    "text": text,
                }
                if include_blocks:
                    record["blocks"] = _page_blocks(page)
                yield record

        records_path = generated_output_path(
            output_directory, "pdf-records.ndjson"
        )
        record_count, byte_size = write_ndjson_atomic(
            page_records(), records_path
        )
    finally:
        close = getattr(document, "close", None)
        if callable(close):
            close()

    artifact = artifact_descriptor(
        records_path, output_directory, "application/x-ndjson"
    )
    metrics = {
        "inputChecksum": sha256_file(input_path),
        "pageCount": page_count,
        "extractedPageCount": extracted_pages,
        "recordCount": record_count,
        "recordsBytes": byte_size,
    }
    return records_path.name, [artifact], metrics


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if math.isfinite(value):
            return value
        return str(value)
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    if isinstance(value, bytes):
        return {"encoding": "base64", "data": base64.b64encode(value).decode("ascii")}
    return str(value)


def _workbook_options(
    options: Mapping[str, Any],
) -> Tuple[bool, int, Optional[List[str]]]:
    include_empty = _bool_option(options, "includeEmptyCells", False)
    max_cells = _positive_int_option(
        options, "maxCells", 1_000_000, 10_000_000
    )
    selected = options.get("sheets")
    if selected is None:
        sheet_names = None
    else:
        if (
            not isinstance(selected, list)
            or not selected
            or any(not isinstance(item, str) or not item for item in selected)
        ):
            raise ComputeOperationError(
                "options.sheets must be a non-empty array of sheet names",
                "invalid_options",
            )
        sheet_names = list(dict.fromkeys(selected))
    return include_empty, max_cells, sheet_names


def _extract_workbook(
    input_path: Path, output_directory: Path, options: Mapping[str, Any]
) -> Tuple[str, List[Dict[str, Any]], Dict[str, Any]]:
    if input_path.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise ComputeOperationError(
            "extract_workbook supports xlsx, xlsm, xltx, and xltm files",
            "unsupported_workbook_format",
        )

    openpyxl = _load_openpyxl()
    include_empty, max_cells, selected_sheets = _workbook_options(options)
    include_cached = _bool_option(options, "includeCachedValues", True)
    keep_vba = input_path.suffix.lower() in (".xlsm", ".xltm")

    load_options = {
        "read_only": True,
        "keep_vba": keep_vba,
        "keep_links": True,
    }
    formula_workbook = None
    cached_workbook = None
    try:
        formula_workbook = openpyxl.load_workbook(
            filename=str(input_path), data_only=False, **load_options
        )
        cached_workbook = (
            openpyxl.load_workbook(
                filename=str(input_path), data_only=True, **load_options
            )
            if include_cached
            else None
        )
    except Exception as exc:
        if formula_workbook is not None:
            formula_workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()
        raise ComputeOperationError(
            "Workbook could not be opened: {}".format(exc), "invalid_workbook"
        ) from exc

    cell_count = 0
    visited_cell_count = 0
    sheet_count = 0
    try:
        available_sheets = list(formula_workbook.sheetnames)
        sheets = selected_sheets or available_sheets
        missing_sheets = [name for name in sheets if name not in available_sheets]
        if missing_sheets:
            raise ComputeOperationError(
                "Workbook does not contain sheet(s): {}".format(
                    ", ".join(missing_sheets)
                ),
                "invalid_options",
            )

        def workbook_records() -> Iterable[Dict[str, Any]]:
            nonlocal cell_count, visited_cell_count, sheet_count
            yield {
                "recordType": "workbook",
                "sourceName": input_path.name,
                "sheetNames": sheets,
                "keepVba": keep_vba,
            }
            for sheet_name in sheets:
                formula_sheet = formula_workbook[sheet_name]
                cached_sheet = (
                    cached_workbook[sheet_name]
                    if cached_workbook is not None
                    else None
                )
                sheet_count += 1
                yield {
                    "recordType": "worksheet",
                    "sheet": sheet_name,
                    "maxRow": int(formula_sheet.max_row or 0),
                    "maxColumn": int(formula_sheet.max_column or 0),
                }
                cached_rows = (
                    iter(cached_sheet.iter_rows())
                    if cached_sheet is not None
                    else None
                )
                for row in formula_sheet.iter_rows():
                    cached_row = (
                        next(cached_rows, ())
                        if cached_rows is not None
                        else ()
                    )
                    cached_by_coordinate = {
                        candidate.coordinate: candidate
                        for candidate in cached_row
                    }
                    for cell in row:
                        visited_cell_count += 1
                        if visited_cell_count > max_cells:
                            raise ComputeOperationError(
                                "Workbook exceeds options.maxCells",
                                "document_limit_exceeded",
                            )
                        value = cell.value
                        if value is None and not include_empty:
                            continue
                        is_formula = bool(
                            getattr(cell, "data_type", None) == "f"
                            or (
                                isinstance(value, str)
                                and value.startswith("=")
                            )
                        )
                        cached_value = None
                        if is_formula and cached_sheet is not None:
                            cached_cell = cached_by_coordinate.get(
                                cell.coordinate
                            )
                            if cached_cell is not None:
                                cached_value = cached_cell.value
                        cell_count += 1
                        yield {
                            "recordType": "cell",
                            "sheet": sheet_name,
                            "coordinate": str(cell.coordinate),
                            "row": int(cell.row),
                            "column": int(cell.column),
                            "value": _json_value(value),
                            "formula": str(value) if is_formula else None,
                            "cachedValue": _json_value(cached_value),
                            "dataType": str(getattr(cell, "data_type", "") or ""),
                            "numberFormat": str(
                                getattr(cell, "number_format", "") or ""
                            ),
                        }

        records_path = generated_output_path(
            output_directory, "workbook-records.ndjson"
        )
        record_count, byte_size = write_ndjson_atomic(
            workbook_records(), records_path
        )
    finally:
        formula_workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()

    artifact = artifact_descriptor(
        records_path, output_directory, "application/x-ndjson"
    )
    metrics = {
        "inputChecksum": sha256_file(input_path),
        "sheetCount": sheet_count,
        "cellCount": cell_count,
        "visitedCellCount": visited_cell_count,
        "recordCount": record_count,
        "recordsBytes": byte_size,
        "macrosPresent": keep_vba,
    }
    return records_path.name, [artifact], metrics


def execute_request(request: Mapping[str, Any]) -> Dict[str, Any]:
    request_id = str(request["requestId"])
    operation = str(request["operation"])
    started = time.monotonic()
    try:
        handlers = {
            "extract_pdf": _extract_pdf,
            "extract_document": extract_document,
            "render_pdf_page": render_pdf_page,
            "extract_workbook": _extract_workbook,
            "derive_workbook": derive_workbook,
            "fetch_market_data": fetch_market_data,
            "render_report": render_report,
        }
        handler = handlers.get(operation)
        if handler is None:
            raise UnsupportedOperationError(operation)

        input_path = require_input_file(str(request["inputPath"]))
        output_directory = prepare_output_directory(
            str(request["outputDirectory"])
        )
        options = request.get("options", {})
        if not isinstance(options, Mapping):
            raise ComputeOperationError(
                "options must be a JSON object", "invalid_options"
            )

        records_file, artifacts, metrics = handler(
            input_path, output_directory, options
        )

        metrics["durationMs"] = round((time.monotonic() - started) * 1000, 3)
        return {
            "protocolVersion": PROTOCOL_VERSION,
            "requestId": request_id,
            "status": "completed",
            "recordsFile": records_file,
            "artifacts": artifacts,
            "metrics": metrics,
            "error": None,
        }
    except (ComputeOperationError, PathValidationError) as exc:
        code = getattr(exc, "code", "invalid_path")
        return failed_response(request_id, str(exc), code)
    except Exception as exc:
        return failed_response(
            request_id,
            "Compute operation failed: {}: {}".format(type(exc).__name__, exc),
            "compute_failed",
        )
